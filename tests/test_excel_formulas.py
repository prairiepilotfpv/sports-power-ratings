import math
import re

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.pipelines.excel_formulas import (
    apply_model_prob_formulas_for_bets_sheet,
    build_header_letter_map,
    resolve_market_inputs,
    validate_bets_formulas,
)


def _evaluate_excel_formula(formula: str, cell_values: dict[str, float | str | None]) -> float:
    expr = formula
    if expr.startswith("=@"):
        expr = expr[2:]
    elif expr.startswith("="):
        expr = expr[1:]

    expr = re.sub(
        r"\b[A-Z]{1,3}\d+\b",
        lambda match: f"cell_values['{match.group(0)}']",
        expr,
    )
    expr = re.sub(r"(?<![<>=!])=(?!=)", "==", expr)

    def _if(condition, when_true, when_false):
        return when_true if condition else when_false

    env = {
        "IF": _if,
        "OR": lambda *args: any(args),
        "ERF": math.erf,
        "SQRT": math.sqrt,
        "ABS": abs,
        "cell_values": cell_values,
    }

    return eval(expr, {"__builtins__": {}}, env)


def _norm_cdf(x: float, mean: float, sd: float) -> float:
    return 0.5 * (1 + math.erf((x - mean) / (sd * math.sqrt(2))))


def _evaluate_model_prob(ws: Worksheet, row: int) -> float:
    headers = [cell.value for cell in ws[1]]
    model_col = headers.index("model_prob") + 1
    model_cell = f"{get_column_letter(model_col)}{row}"
    formula = ws[model_cell].value
    assert isinstance(formula, str)

    cell_values = {
        f"{get_column_letter(col)}{row}": ws.cell(row=row, column=col).value
        for col in range(1, ws.max_column + 1)
    }

    return _evaluate_excel_formula(formula, cell_values)


def test_bets_model_prob_formula_uses_erf():
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"

    headers = [
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
        "home_win_prob",
        "away_win_prob",
        "margin_mean",
        "margin_sd",
        "total",
        "total_sd",
    ]
    ws.append(headers)
    ws.append(
        [
            "spread",
            "Home",
            -3.5,
            "",
            "Home",
            "Away",
            0.6,
            0.4,
            4.0,
            12.0,
            210.5,
            15.0,
        ]
    )

    apply_model_prob_formulas_for_bets_sheet(ws)

    model_col = headers.index("model_prob") + 1
    model_cell = f"{get_column_letter(model_col)}2"
    formula = ws[model_cell].value

    assert "ERF(" in formula
    assert "NORM.DIST" not in formula
    assert "NORMDIST" not in formula


def test_header_letter_map_and_resolve_market_inputs():
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"
    headers = [
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
        "ml_home_win_prob",
        "ml_away_win_prob",
        "spread_margin_mean",
        "spread_margin_sd",
        "total_mean",
        "total_sd",
    ]
    ws.append(headers)
    letters = build_header_letter_map(ws)

    assert letters["market_type"] == "A"
    assert letters["model_prob"] == "D"
    assert letters["total_sd"] == "L"

    inputs = resolve_market_inputs(letters)
    assert inputs["ML"]["home_prob"] == letters["ml_home_win_prob"]
    assert inputs["ML"]["away_prob"] == letters["ml_away_win_prob"]
    assert inputs["spread"]["mean"] == letters["spread_margin_mean"]
    assert inputs["spread"]["sd"] == letters["spread_margin_sd"]
    assert inputs["total"]["mean"] == letters["total_mean"]
    assert inputs["total"]["sd"] == letters["total_sd"]


def test_resolve_market_inputs_fallbacks():
    cols_map = {
        "home_win_prob": "G",
        "away_win_prob": "H",
        "margin_mean": "I",
        "margin_sd": "J",
        "total": "K",
        "total_sd": "L",
        "line": "C",
    }
    inputs = resolve_market_inputs(cols_map)
    assert inputs["ML"]["home_prob"] == "G"
    assert inputs["ML"]["away_prob"] == "H"
    assert inputs["spread"]["mean"] == "I"
    assert inputs["spread"]["sd"] == "J"
    assert inputs["total"]["mean"] == "K"
    assert inputs["total"]["sd"] == "L"


def test_validate_bets_formulas_on_generated_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"

    headers = [
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
        "home_win_prob",
        "away_win_prob",
        "margin_mean",
        "margin_sd",
        "total",
        "total_sd",
    ]
    ws.append(headers)
    ws.append(["ML", "Home", 0.0, "", "Home", "Away", 0.55, 0.45, 4.0, 12.0, 210.5, 15.0])
    ws.append(["spread", "Home", -3.5, "", "Home", "Away", 0.55, 0.45, 4.0, 12.0, 210.5, 15.0])
    ws.append(["total", "Over", 210.5, "", "Home", "Away", 0.55, 0.45, 4.0, 12.0, 210.5, 15.0])

    apply_model_prob_formulas_for_bets_sheet(ws)
    out = tmp_path / "bets.xlsx"
    wb.save(out)

    reloaded = load_workbook(out)
    ws_loaded = reloaded["BETS"]
    validate_bets_formulas(ws_loaded)


def test_no_formula_contains_ellipsis_in_bets():
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"
    headers = [
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
        "home_win_prob",
        "away_win_prob",
        "margin_mean",
        "margin_sd",
        "total",
        "total_sd",
    ]
    ws.append(headers)
    ws.append(["spread", "Home", -3.5, "", "Home", "Away", 0.6, 0.4, 4.0, 12.0, 210.5, 15.0])

    apply_model_prob_formulas_for_bets_sheet(ws)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(("=", "=@")):
                assert "..." not in cell.value


@pytest.mark.parametrize(
    "selection,line,expected",
    [
        ("AwayTeam", -1.5, _norm_cdf(-1.5, 0.0, 1.0)),
        ("AwayTeam", 1.5, _norm_cdf(1.5, 0.0, 1.0)),
        ("HomeTeam", -1.5, 1 - _norm_cdf(1.5, 0.0, 1.0)),
        ("HomeTeam", 1.5, 1 - _norm_cdf(-1.5, 0.0, 1.0)),
    ],
    ids=[
        "away_favorite_signed_line",
        "away_underdog_signed_line",
        "home_favorite_signed_line",
        "home_underdog_signed_line",
    ],
)
def test_spread_model_prob_signed_lines(selection, line, expected):
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"

    headers = [
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
        "home_win_prob",
        "away_win_prob",
        "margin_mean",
        "margin_sd",
        "total",
        "total_sd",
    ]
    ws.append(headers)
    ws.append(
        [
            "spread",
            selection,
            line,
            "",
            "HomeTeam",
            "AwayTeam",
            0.55,
            0.45,
            0.0,
            1.0,
            210.5,
            15.0,
        ]
    )

    apply_model_prob_formulas_for_bets_sheet(ws)

    model_prob = _evaluate_model_prob(ws, 2)

    assert model_prob == pytest.approx(expected, rel=1e-6)
