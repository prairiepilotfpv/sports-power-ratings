from pathlib import Path
import tempfile
from datetime import date
import sqlite3

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.pipelines.review_runs import build_review_workbook_with_formulas
from src.data import betting_repository as br
from src.data import repository as repo


def test_review_workbook_formulas_exist():
    td = Path(tempfile.mkdtemp())
    try:
        db_path = Path(td) / "test.db"
        repo.init_db(db_path)
        repo.save_games(
            db_path,
            [
                repo.GameResult(
                    date=date(2025, 11, 10),
                    home_team="Los Angeles Lakers",
                    away_team="LA Clippers",
                    home_score=None,
                    away_score=None,
                    neutral=False,
                    overtime=False,
                    decision_type=None,
                    game_id="2025-11-10-lakers-clippers",
                    sport="nba",
                    season="2025-26",
                    division=None,
                    conference=None,
                    notes=None,
                )
            ],
        )
        br.init_db(db_path)
        rid = br.create_review_run(db_path, sport="nba", season="2025-26", model="elo", notes="test")
        with sqlite3.connect(db_path) as conn:
            conn.execute(
                "INSERT INTO opportunities (review_run_id, game_id, market_type, selection, line, odds, implied_prob, model_prob, edge, ev, source_market_snapshot_id, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))",
                (rid, "2025-11-10-lakers-clippers", "ML", "Los Angeles Lakers", 0.0, 110, 0.48, 0.55, 0.07, 0.08, None),
            )
            conn.commit()

        out = build_review_workbook_with_formulas(db_path, review_run_id=rid, sport="nba", season="2025-26")
        wb = load_workbook(out)
        try:
            for sheet_name in ("EV", "BETS"):
                ws = wb[sheet_name]
                header = {cell.value: cell.column for cell in ws[1] if cell.value}
                odds_cell = f"{get_column_letter(header['odds'])}2"
                implied_cell = f"{get_column_letter(header['implied_prob'])}2"
                model_cell = f"{get_column_letter(header['model_prob'])}2"
                edge_cell = f"{get_column_letter(header['edge'])}2"
                ev_cell = f"{get_column_letter(header['ev'])}2"

                expected_implied = (
                    f"=IF(OR({odds_cell}=\"\",{odds_cell}=0),\"\",IF({odds_cell}>0,100/({odds_cell}+100),-{odds_cell}/(-{odds_cell}+100)))"
                )
                expected_edge = f"=IF(OR({model_cell}=\"\",{implied_cell}=\"\"),\"\",{model_cell}-{implied_cell})"
                expected_ev = (
                    f"=IF(OR({model_cell}=\"\",{odds_cell}=\"\",{odds_cell}=0),\"\",({model_cell}*IF({odds_cell}>0,{odds_cell}/100,100/ABS({odds_cell})))-(1-{model_cell}))"
                )

                assert ws[implied_cell].value == expected_implied
                assert ws[edge_cell].value == expected_edge
                assert ws[ev_cell].value == expected_ev
        finally:
            try:
                wb.close()
            except Exception:
                pass
        try:
            sqlite3.connect(db_path).close()
        except Exception:
            pass
    finally:
        import shutil, errno

        try:
            shutil.rmtree(td)
        except OSError as exc:
            if exc.errno != errno.EACCES:
                raise
