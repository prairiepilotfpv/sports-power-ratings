from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from data.repository import load_games, save_games
from ingest.schema import GameResult
from models.bradley_terry import BradleyTerry
from pipelines.metadata import prediction_hash
from pipelines.projection_engines import get_projection_engine
from pipelines.schedule import (
    DASHBOARD_COLUMNS,
    MODEL_METADATA_DATA_START_ROW,
    SCHEDULE_EXPORT_COLUMNS,
    _project_row,
    _order_schedule_export,
    build_schedule_excel_report,
    build_schedule_with_projections,
)


def test_build_schedule_with_projections(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Team C",
            away_team="Team A",
            home_score=95,
            away_score=110,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team B",
            away_team="Team C",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    output_path = build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.csv",
    )

    df = pd.read_csv(output_path)
    assert len(df) == 3
    assert set(df["status"]) == {"final", "scheduled"}
    assert "notes" not in df.columns
    assert "model_error" not in df.columns
    assert "margin_mean" in df.columns
    assert "total_mean" in df.columns
    assert "projection_status" in df.columns

    upcoming = df[df["status"] == "scheduled"].iloc[0]
    assert upcoming["home_team"] == "Team B"
    assert upcoming["away_team"] == "Team C"
    assert upcoming["projection_status"] == "ok"
    assert pd.notna(upcoming["projected_winner"])
    assert pd.notna(upcoming["projected_spread"])
    assert upcoming["projected_total"] > 0
    assert pd.notna(upcoming["margin_mean"])
    assert upcoming["margin_sd"] > 0
    assert pd.notna(upcoming["total_mean"])
    assert upcoming["total_sd"] > 0
    assert upcoming["home_win_prob"] == pytest.approx(upcoming["projected_win_prob"])
    assert upcoming["model_win_prob"] == pytest.approx(upcoming["projected_win_prob"])


def test_build_schedule_with_elo_projections(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Team C",
            away_team="Team A",
            home_score=95,
            away_score=110,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team B",
            away_team="Team C",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    output_path = build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="elo",
        output_path=tmp_path / "schedule.csv",
    )

    df = pd.read_csv(output_path)
    upcoming = df[df["status"] == "scheduled"].iloc[0]
    assert pd.notna(upcoming["projected_winner"])
    assert pd.notna(upcoming["projected_spread"])
    assert upcoming["projected_total"] > 0


def test_poisson_schedule_exports_total_sd(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=3,
            away_score=2,
            sport="nhl",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Team B",
            away_team="Team A",
            home_score=4,
            away_score=1,
            sport="nhl",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team A",
            away_team="Team B",
            home_score=None,
            away_score=None,
            sport="nhl",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    output_path = build_schedule_with_projections(
        db_path,
        sport="nhl",
        season="2024-25",
        model="poisson",
        model_params={"n_simulations": 500, "random_seed": 7},
        output_path=tmp_path / "schedule_poisson.csv",
    )

    df = pd.read_csv(output_path)
    upcoming = df[df["status"] == "scheduled"].iloc[0]
    assert pd.notna(upcoming["projected_total"])
    assert pd.notna(upcoming["total_sd"])


def test_schedule_win_probs_follow_margin_sign() -> None:
    base_row = pd.Series(
        {
            "date": "2024-01-01",
            "home_team": "Home",
            "away_team": "Away",
            "home_score": None,
            "away_score": None,
            "neutral": False,
            "overtime": False,
            "game_id": "gid-1",
        }
    )
    ratings = {"Home": 5.0, "Away": 0.0}
    model_instance = BradleyTerry(max_iter=50)
    model_instance.fit(
        [
            {"home_team": "Home", "away_team": "Away", "home_score": 100, "away_score": 90},
            {"home_team": "Away", "away_team": "Home", "home_score": 95, "away_score": 105},
        ]
    )
    projection_engine = get_projection_engine(model_instance)
    projection_context = {
        "ratings": ratings,
        "base_total": 0.0,
        "scoring_averages": {},
        "total_intercept": None,
        "total_slope": None,
        "margin_std": 8.0,
        "total_std": 15.0,
        "conditional_sd_intercept": None,
        "conditional_sd_slope": None,
        "win_prob_k": 10.0,
    }
    positive = _project_row(
        base_row,
        ratings=ratings,
        status="scheduled",
        home_advantage=0.0,
        params_source="default",
        tuned_metric_used=None,
        model_instance=model_instance,
        projection_engine=projection_engine,
        projection_context=projection_context,
    )
    assert positive["margin_mean"] > 0
    assert positive["home_win_prob"] > 0.5
    assert positive["away_win_prob"] == pytest.approx(1.0 - positive["home_win_prob"])
    assert positive["winner_win_prob"] == pytest.approx(positive["home_win_prob"])

    negative_model = BradleyTerry(max_iter=50)
    negative_model.fit(
        [
            {"home_team": "Home", "away_team": "Away", "home_score": 90, "away_score": 100},
            {"home_team": "Away", "away_team": "Home", "home_score": 105, "away_score": 95},
        ]
    )
    negative = _project_row(
        base_row,
        ratings={"Home": 0.0, "Away": 5.0},
        status="scheduled",
        home_advantage=0.0,
        params_source="default",
        tuned_metric_used=None,
        model_instance=negative_model,
        projection_engine=get_projection_engine(negative_model),
        projection_context={
            **projection_context,
            "ratings": {"Home": 0.0, "Away": 5.0},
        },
    )
    assert negative["margin_mean"] < 0
    assert negative["home_win_prob"] < 0.5
    assert negative["away_win_prob"] == pytest.approx(1.0 - negative["home_win_prob"])
    assert negative["winner_win_prob"] == pytest.approx(negative["away_win_prob"])


def test_schedule_uses_latest_scores(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    game_id = "2024-01-03|Team A|Team B"
    initial_game = GameResult(
        date=date(2024, 1, 3),
        home_team="Team B",
        away_team="Team A",
        home_score=None,
        away_score=None,
        game_id=game_id,
        sport="nba",
        season="2024-25",
    )
    save_games(db_path, [initial_game])

    output_path = tmp_path / "schedule.csv"
    build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=output_path,
    )
    initial_df = pd.read_csv(output_path)
    initial_row = initial_df.loc[initial_df["game_id"] == game_id].iloc[0]
    assert pd.isna(initial_row["home_score"])
    assert pd.isna(initial_row["away_score"])
    assert initial_row["status"] == "scheduled"

    updated_game = GameResult(
        date=date(2024, 1, 3),
        home_team="Team B",
        away_team="Team A",
        home_score=101,
        away_score=99,
        game_id=game_id,
        sport="nba",
        season="2024-25",
    )
    save_games(db_path, [updated_game])

    build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=output_path,
    )
    updated_df = pd.read_csv(output_path)
    updated_row = updated_df.loc[updated_df["game_id"] == game_id].iloc[0]
    assert updated_row["home_score"] == 101
    assert updated_row["away_score"] == 99
    assert updated_row["status"] == "final"

    db_games = {
        game.game_id: game
        for game in load_games(db_path, sport="nba", season="2024-25")
    }
    assert db_games[game_id].home_score == updated_row["home_score"]
    assert db_games[game_id].away_score == updated_row["away_score"]


def test_schedule_export_column_ordering() -> None:
    row = {
        "date": "2024-01-01",
        "game_id": "gid-1",
        "status": "final",
        "projection_status": "ok",
        "params_source": "default",
        "tuned_metric_used": None,
        "home_team": "Home",
        "away_team": "Away",
        "neutral": False,
        "overtime": False,
        "home_score": 100,
        "away_score": 90,
        "result_margin": 10,
        "result_total": 190,
        "home_rating": 1.5,
        "away_rating": -0.2,
        "home_advantage": 3.0,
        "projected_winner": "Home",
        "projected_spread": -4.5,
        "projected_home_spread": 4.5,
        "projected_win_prob": 0.65,
        "home_win_prob": 0.65,
        "away_win_prob": 0.35,
        "winner_win_prob": 0.65,
        "logistic_home_win_prob": 0.62,
        "projected_win_prob_dist": '[{"p_home_win": 0.65, "weight": 1.0}]',
        "projected_home_score": 102.5,
        "projected_away_score": 95.5,
        "projected_total": 198.0,
        "margin_mean": 7.0,
        "margin_sd": 12.0,
        "total_mean": 198.0,
        "total_sd": 20.0,
        "margin_dist_params": '{"mean": 7.0, "sd": 12.0}',
        "total_dist_params": '{"mean": 198.0, "sd": 20.0}',
        "model_win_prob_samples": '[{"p_home_win": 0.65, "weight": 1.0}]',
        "model_win_prob": 0.65,
        "margin_std": 12.1,
        "total_std": 18.4,
    }
    # Shuffle columns to simulate unordered input
    shuffled_columns = sorted(row.keys())
    df = pd.DataFrame([row], columns=shuffled_columns)

    ordered = _order_schedule_export(df)
    assert list(ordered.columns) == SCHEDULE_EXPORT_COLUMNS


def test_schedule_export_column_ordering_missing_column() -> None:
    base_df = pd.DataFrame(
        [
            {
                "date": "2024-01-01",
                "status": "final",
                "projection_status": "ok",
                "params_source": "default",
                "tuned_metric_used": None,
                "home_team": "Home",
                "away_team": "Away",
                "neutral": False,
                "overtime": False,
                "home_score": 100,
                "away_score": 90,
                "result_margin": 10,
                "result_total": 190,
                "home_rating": 1.5,
                "away_rating": -0.2,
                "home_advantage": 3.0,
                "projected_winner": "Home",
                "projected_spread": -4.5,
                "projected_home_spread": 4.5,
                "projected_win_prob": 0.65,
                "home_win_prob": 0.65,
                "away_win_prob": 0.35,
                "winner_win_prob": 0.65,
                "logistic_home_win_prob": 0.62,
                "projected_win_prob_dist": None,
                "projected_home_score": 102.5,
                "projected_away_score": 95.5,
                "projected_total": 198.0,
                "margin_mean": None,
                "margin_sd": None,
                "total_mean": None,
                "total_sd": None,
                "margin_dist_params": None,
                "total_dist_params": None,
                "model_win_prob_samples": None,
                "model_win_prob": None,
            }
        ]
    )

    with pytest.raises(ValueError, match="game_id"):
        _order_schedule_export(base_df)


def test_schedule_upcoming_only_filters_completed(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team C",
            away_team="Team D",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    output_path = build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.csv",
        upcoming_only=True,
    )

    df = pd.read_csv(output_path)
    assert len(df) == 1
    assert df.iloc[0]["status"] == "scheduled"


def test_schedule_neutral_games_use_zero_home_advantage(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            neutral=True,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Team A",
            away_team="Team B",
            home_score=None,
            away_score=None,
            neutral=True,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    output_path = build_schedule_with_projections(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.csv",
    )

    df = pd.read_csv(output_path)
    assert (df["home_advantage"] == 0.0).all()


def test_schedule_excel_report_matches_csv_outputs(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team B",
            away_team="Team C",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    workbook_path = build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        output_path=tmp_path / "schedule.xlsx",
    )

    for model in ["bradley-terry", "elo", "toor"]:
        csv_path = build_schedule_with_projections(
            db_path,
            sport="nba",
            season="2024-25",
            model=model,
            output_path=tmp_path / f"{model}.csv",
        )
        expected = pd.read_csv(csv_path)
        actual = pd.read_excel(
            workbook_path,
            sheet_name=model,
            skiprows=MODEL_METADATA_DATA_START_ROW - 1,
        )
        assert list(actual.columns) == SCHEDULE_EXPORT_COLUMNS
        pd.testing.assert_frame_equal(actual, expected, check_dtype=False)

    dashboard = pd.read_excel(workbook_path, sheet_name="dashboard")
    assert list(dashboard.columns) == DASHBOARD_COLUMNS
    assert dashboard.empty


def test_schedule_excel_dashboard_includes_today_games(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    today = date.today()
    games = [
        GameResult(
            date=today - timedelta(days=1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=today - timedelta(days=1),
            home_team="Team C",
            away_team="Team D",
            home_score=102,
            away_score=98,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=today,
            home_team="Team A",
            away_team="Team B",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=today + timedelta(days=1),
            home_team="Team C",
            away_team="Team D",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    workbook_path = build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.xlsx",
    )

    dashboard = pd.read_excel(workbook_path, sheet_name="dashboard")
    assert list(dashboard.columns) == DASHBOARD_COLUMNS
    assert not dashboard.empty
    assert set(dashboard["model"]) == {"bradley-terry"}
    assert dashboard["model_version"].notna().all()
    assert set(dashboard["game"]) == {"Team B @ Team A"}
    assert (
        dashboard.loc[dashboard["game"] == "Team B @ Team A", "projected_winner"]
        .notna()
        .all()
    )
    assert (
        dashboard.loc[dashboard["game"] == "Team B @ Team A", "projected_spread"]
        .notna()
        .all()
    )
    assert (
        dashboard.loc[dashboard["game"] == "Team B @ Team A", "home_win_prob"]
        .notna()
        .all()
    )
    assert (
        dashboard.loc[dashboard["game"] == "Team B @ Team A", "winner_win_prob"]
        .notna()
        .all()
    )
    total_value = dashboard.loc[dashboard["game"] == "Team B @ Team A", "total"].iloc[0]
    home_score = dashboard.loc[
        dashboard["game"] == "Team B @ Team A", "projected_home_score"
    ].iloc[0]
    away_score = dashboard.loc[
        dashboard["game"] == "Team B @ Team A", "projected_away_score"
    ].iloc[0]
    assert total_value == pytest.approx(home_score + away_score)


def test_schedule_excel_report_includes_elo_dashboard_rows(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    today = date.today()
    games = [
        GameResult(
            date=today - timedelta(days=1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=today,
            home_team="Team A",
            away_team="Team B",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    workbook_path = build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        output_path=tmp_path / "schedule.xlsx",
    )

    workbook = openpyxl.load_workbook(workbook_path)
    assert "elo" in workbook.sheetnames

    dashboard = pd.read_excel(workbook_path, sheet_name="dashboard")
    assert "elo" in set(dashboard["model"])
    elo_rows = dashboard[dashboard["model"] == "elo"]
    assert elo_rows["projected_winner"].notna().all()
    assert elo_rows["projected_spread"].notna().all()


def test_schedule_excel_report_includes_model_metadata(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 5),
            home_team="Team B",
            away_team="Team C",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)

    workbook_path = build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        output_path=tmp_path / "schedule.xlsx",
    )

    workbook = openpyxl.load_workbook(workbook_path)
    for model in ["bradley-terry", "elo", "toor"]:
        ws = workbook[model]
        metadata: dict[str, str] = {}
        for row in ws.iter_rows(
            min_row=2, max_row=10, min_col=1, max_col=2, values_only=True
        ):
            key, value = row
            if key:
                metadata[str(key)] = str(value) if value is not None else ""

        assert metadata.get("model_id") == model

        schedule_df = pd.read_excel(
            workbook_path,
            sheet_name=model,
            skiprows=MODEL_METADATA_DATA_START_ROW - 1,
        )
        expected_hash = prediction_hash(schedule_df, SCHEDULE_EXPORT_COLUMNS)
        assert metadata.get("prediction_hash") == expected_hash
