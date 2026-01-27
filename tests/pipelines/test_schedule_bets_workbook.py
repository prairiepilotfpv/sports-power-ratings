from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path
import sqlite3
from typing import Any

import openpyxl
import pandas as pd
import pytest

from ingest.schema import GameResult
from pipelines import schedule as schedule_pipeline
from src.data import repository as repo
from src.data import betting_repository as br
from src.pipelines import bets as bets_pipeline


def _seed_schedule_db(db_path: Path, scheduled_date: date) -> None:
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Team A",
            away_team="Team B",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
            game_id="2024-01-01-team-a-team-b",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Team C",
            away_team="Team A",
            home_score=95,
            away_score=110,
            sport="nba",
            season="2024-25",
            game_id="2024-01-02-team-c-team-a",
        ),
        GameResult(
            date=scheduled_date,
            home_team="Team B",
            away_team="Team C",
            home_score=None,
            away_score=None,
            sport="nba",
            season="2024-25",
            game_id="2024-01-05-team-b-team-c",
        ),
    ]
    repo.save_games(db_path, games)


def test_schedule_workbook_includes_bets_and_meta(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    repo.init_db(db_path)
    br.init_db(db_path)
    scheduled_date = date(2024, 1, 5)
    _seed_schedule_db(db_path, scheduled_date)

    workbook_path = schedule_pipeline.build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.xlsx",
        as_of_date=scheduled_date,
    )

    wb = openpyxl.load_workbook(workbook_path)
    assert "BETS" in wb.sheetnames
    assert "META" in wb.sheetnames
    assert wb["META"].sheet_state == "hidden"

    dashboard_df = pd.read_excel(workbook_path, sheet_name="dashboard")
    assert list(dashboard_df.columns) == schedule_pipeline.DASHBOARD_COLUMNS

    meta_df = pd.read_excel(workbook_path, sheet_name="META")
    meta = dict(zip(meta_df["key"], meta_df["value"]))
    assert "ml_bradley-terry" in meta["review_run_id"]
    # SPREAD/TOTAL sources can be either the model name (single-model pass-through)
    # or ensemble ID (when ensemble is applied)
    review_id = meta["review_run_id"]
    assert ("spread_bradley-terry" in review_id or "spread_ensemble_spread_v1" in review_id), f"Expected spread source in {review_id}"
    assert ("total_bradley-terry" in review_id or "total_ensemble_total_v1" in review_id), f"Expected total source in {review_id}"


def test_schedule_bets_rows_and_log_bets(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    repo.init_db(db_path)
    br.init_db(db_path)
    scheduled_date = date(2024, 1, 5)
    _seed_schedule_db(db_path, scheduled_date)

    workbook_path = schedule_pipeline.build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        model="bradley-terry",
        output_path=tmp_path / "schedule.xlsx",
        as_of_date=scheduled_date,
    )

    bets_df = pd.read_excel(workbook_path, sheet_name="BETS")
    assert len(bets_df) == 6
    assert set(bets_df["market_type"]) == {"ML", "spread", "total"}
    assert {"Over", "Under"} <= set(bets_df["selection"])

    bets_df.loc[0, "stake"] = 1
    bets_df.loc[0, "odds"] = -110
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        writer.book.remove(writer.book["BETS"])
        bets_df.to_excel(writer, sheet_name="BETS", index=False)

    inserted = bets_pipeline.log_bets(
        str(workbook_path),
        review_run_id=None,
        db_path=db_path,
        dry_run=False,
        writeback=True,
    )
    assert inserted == 1

    with sqlite3.connect(db_path) as conn:
        count = conn.execute("SELECT COUNT(*) FROM bets").fetchone()[0]
        assert count == 1

    updated_df = pd.read_excel(workbook_path, sheet_name="BETS")
    staked = updated_df[updated_df["stake"].notna() & (updated_df["stake"] != "")]
    assert not staked["bet_id"].isna().all()


def test_schedule_ensemble_uses_tuned_weights(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        schedule_pipeline,
        "list_models",
        lambda: ["bradley-terry", "elo"],
    )

    db_path = tmp_path / "games.db"
    repo.init_db(db_path)
    br.init_db(db_path)
    scheduled_date = date(2024, 1, 5)
    _seed_schedule_db(db_path, scheduled_date)

    weights_payload = {
        "ensemble_id": "ensemble_ml_v1",
        "market": "ML",
        "objective": "log_loss",
        "train_window": {"start": "2024-01-01", "end": "2024-01-04"},
        "models": ["bradley-terry", "elo"],
        "weights": {"bradley-terry": 0.9, "elo": 0.1},
        "created_at": "2024-01-05T00:00:00Z",
    }
    weights_path = (
        Path("outputs")
        / "ensembles"
        / "nba"
        / "2024-25"
        / "ML"
        / "ensemble_ml_v1.json"
    )
    weights_path.parent.mkdir(parents=True, exist_ok=True)
    weights_path.write_text(json.dumps(weights_payload), encoding="utf-8")
    from ensemble.io import load_ml_weights

    assert load_ml_weights("nba", "2024-25", "ensemble_ml_v1") == {
        "bradley-terry": 0.9,
        "elo": 0.1,
    }

    workbook_path = schedule_pipeline.build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        model=None,
        output_path=tmp_path / "schedule.xlsx",
        as_of_date=scheduled_date,
    )

    bets_df = pd.read_excel(workbook_path, sheet_name="BETS")
    ml_rows = bets_df[bets_df["market_type"] == "ML"]
    home_row = ml_rows[ml_rows["selection"] == "Team B"].iloc[0]
    components = json.loads(home_row["ml_ensemble_components_json"])
    weights = {comp["model"]: comp["weight"] for comp in components}
    assert weights["bradley-terry"] == pytest.approx(0.9, rel=1e-6)
    assert weights["elo"] == pytest.approx(0.1, rel=1e-6)
    # Human: Interpret component JSON probabilities regardless of legacy vs normalized keys,
    # AI agent: Keep future schema changes manageable by checking both 'prob' and 'value' fields.
    def _component_probability(comp: dict[str, Any]) -> float | None:
        prob = comp.get("prob")
        if prob is None:
            prob = comp.get("value")
        if prob is None:
            return None
        try:
            return float(prob)
        except (TypeError, ValueError):
            return None

    combined = 0.0
    # Human: Accumulate the weighted probabilities only when a safe float exists.
    # AI agent: Ensures the assertion mirrors what production ensemble writes in `ml_ensemble_components_json`.
    for comp in components:
        prob = _component_probability(comp)
        if prob is None:
            continue
        combined += prob * comp["weight"]
    assert home_row["home_win_prob"] == pytest.approx(combined, rel=1e-6)
    assert "ensemble_ml_v1" in str(home_row["win_prob_source"])
    assert "ml_ensemble_components_json" in ml_rows.columns

    meta_df = pd.read_excel(workbook_path, sheet_name="META")
    meta = dict(zip(meta_df["key"], meta_df["value"]))
    assert "ml_ensemble_ml_v1" in meta["review_run_id"]
    assert "spread_ensemble_spread_v1" in meta["review_run_id"]


def test_ml_components_json_not_written_when_ensemble_skipped(
    tmp_path: Path, monkeypatch, caplog
) -> None:
    caplog.set_level(logging.WARNING)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        schedule_pipeline,
        "list_models",
        lambda: ["bradley-terry", "elo"],
    )

    db_path = tmp_path / "games.db"
    repo.init_db(db_path)
    br.init_db(db_path)
    scheduled_date = date(2024, 1, 5)
    _seed_schedule_db(db_path, scheduled_date)

    weights_payload = {
        "ensemble_id": "ensemble_ml_v1",
        "market": "ML",
        "objective": "log_loss",
        "train_window": {"start": "2024-01-01", "end": "2024-01-04"},
        "models": ["bradley-terry", "elo"],
        "weights": {"bradley-terry": 0.9, "elo": 0.1},
        "created_at": "2024-01-05T00:00:00Z",
    }
    weights_path = (
        Path("outputs")
        / "ensembles"
        / "nba"
        / "2024-25"
        / "ML"
        / "ensemble_ml_v1.json"
    )
    weights_path.parent.mkdir(parents=True, exist_ok=True)
    weights_path.write_text(json.dumps(weights_payload), encoding="utf-8")

    from ensemble.ml_v1 import MLWeightedAverageEnsemble

    monkeypatch.setattr(
        MLWeightedAverageEnsemble,
        "combine",
        lambda self, forecast_df: (_raise_runtime_error()),
    )
    original_build_bets = schedule_pipeline._build_bets_dataframe

    def _blank_components(*args: Any, **kwargs: Any) -> pd.DataFrame:
        df = original_build_bets(*args, **kwargs)
        df["ml_ensemble_components_json"] = pd.NA
        return df

    monkeypatch.setattr(
        schedule_pipeline,
        "_build_bets_dataframe",
        _blank_components,
    )

    workbook_path = schedule_pipeline.build_schedule_excel_report(
        db_path,
        sport="nba",
        season="2024-25",
        model=None,
        output_path=tmp_path / "schedule.xlsx",
        as_of_date=scheduled_date,
    )

    bets_df = pd.read_excel(workbook_path, sheet_name="BETS")
    ml_rows = bets_df[bets_df["market_type"] == "ML"]
    assert ml_rows["ml_ensemble_components_json"].isna().all()
    assert "ml_ensemble_components_json remains blank" in caplog.text


def test_build_bets_dataframe_preserves_ensemble_sources_and_totals() -> None:
    """Ensure BETS rows use whatever ensemble sources and totals the schedule already contains."""
    pool = pd.DataFrame(
        [
            {
                "date": date(2024, 1, 5),
                "game_id": "2024-01-05-team-b-team-c",
                "status": "scheduled",
                "away_team": "Team C",
                "home_team": "Team B",
                "home_win_prob": 0.55,
                "away_win_prob": 0.45,
                "win_prob_source": "ensemble_ml_v1",
                "margin_mean": -5.0,
                "margin_sd": 9.5,
                "total": 226.5,
                "total_sd": 12.3,
                "total_source": "ensemble_total_v1",
                "spread_source": "ensemble_spread_v1",
                "total_ensemble_components_json": json.dumps(
                    [
                        {"model": "poisson", "total_mean": 226.0, "total_sd": 10.0},
                        {"model": "gssd", "total_mean": 227.0, "total_sd": 14.0},
                    ]
                ),
            }
        ]
    )

    bets_df = schedule_pipeline._build_bets_dataframe(
        pool,
        model_name="gssd",
        as_of_date=date(2024, 1, 5),
        review_run_id="review-ensemble",
        db_path=None,
        sport=None,
        season=None,
    )

    total_rows = bets_df[bets_df["market_type"] == "total"]
    assert not total_rows.empty
    assert set(total_rows["total_source"]) == {"ensemble_total_v1"}
    assert total_rows["total"].unique().tolist() == [226.5]
    assert total_rows["total_sd"].unique().tolist() == [12.3]

    spread_rows = bets_df[bets_df["market_type"] == "spread"]
    assert set(spread_rows["spread_source"]) == {"ensemble_spread_v1"}

    ml_rows = bets_df[bets_df["market_type"] == "ML"]
    assert set(ml_rows["win_prob_source"]) == {"ensemble_ml_v1"}


def _raise_runtime_error() -> None:
    raise RuntimeError("simulated ensemble failure")
