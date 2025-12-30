from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from data.repository import save_games
from ingest.schema import GameResult
from pipelines.excel_report import build_excel_report as build_multi_report
from pipelines.report import build_excel_report as build_daily_report


def _seed_games(db_path: Path) -> None:
    games = [
        GameResult(
            date=date(2024, 1, 1),
            home_team="Alpha",
            away_team="Beta",
            home_score=100,
            away_score=90,
            sport="nba",
            season="2024-25",
        ),
        GameResult(
            date=date(2024, 1, 2),
            home_team="Beta",
            away_team="Alpha",
            home_score=95,
            away_score=102,
            sport="nba",
            season="2024-25",
        ),
    ]
    save_games(db_path, games)


def test_multi_sheet_excel_report(tmp_path: Path) -> None:
    db_path = tmp_path / "games.db"
    _seed_games(db_path)

    report_path = build_multi_report(
        db_path,
        sport="nba",
        season="2024-25",
        output_path=tmp_path / "report.xlsx",
    )

    assert report_path.exists()
    workbook = openpyxl.load_workbook(report_path)
    assert "bradley-terry" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames


def test_daily_excel_report(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    db_path = tmp_path / "games.db"
    _seed_games(db_path)

    monkeypatch.chdir(tmp_path)

    report_path = build_daily_report(
        db_path,
        sport="nba",
        season="2024-25",
        output_path=tmp_path / "daily_report.xlsx",
    )

    assert report_path.exists()
    workbook = openpyxl.load_workbook(report_path)
    assert "Summary" in workbook.sheetnames
