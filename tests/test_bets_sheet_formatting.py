from openpyxl import Workbook, load_workbook

from pipelines.excel_bets_format import apply_bets_sheet_formatting


def test_bets_sheet_formatting_preserves_headers_and_formulas(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BETS"
    headers = [
        "review_run_id",
        "game_id",
        "date",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "odds",
        "price",
        "implied_prob",
        "model_prob",
        "edge",
        "ev",
        "stake",
        "book",
        "bet_id",
        "logged_at",
        "log_status",
        "notes",
    ]
    ws.append(headers)
    # Put a formula in implied_prob to ensure it's preserved
    row2 = ["r1", "g1", "2025-11-01", "Away", "Home", "ML", "Home", "",  -110, "-110", "=1-0.5", "=0.6", "=0.1", "=0.02", 100, "Book", "", "", "planned", ""]
    ws.append(row2)

    # Apply formatting in-memory before saving
    apply_bets_sheet_formatting(ws)
    top = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert top[: len(headers)] == headers
    # implied_prob formula should be unchanged text-wise
    # find implied_prob column letter and ensure formula preserved
    implied_col = None
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value == "implied_prob":
            implied_col = cell.column_letter
            break
    assert implied_col is not None
    assert ws[f"{implied_col}2"].value == "=1-0.5"

    # helper columns appended
    last_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "american_odds_norm" in last_headers
    assert "decimal_odds" in last_headers
    assert "payout_mult" in last_headers
