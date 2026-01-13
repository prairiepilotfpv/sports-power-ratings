from __future__ import annotations

from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00, FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation


def _header_map(ws: Worksheet) -> Dict[str, str]:
    """Return a mapping of header -> column_letter for the first row."""
    header = next(ws.iter_rows(min_row=1, max_row=1))
    return {str(cell.value).strip(): cell.column_letter for cell in header if cell.value is not None}


def apply_bets_sheet_formatting(ws: Worksheet) -> None:
    """Apply formatting, validation, conditional formatting and helper columns to BETS sheet.

    This function is additive: it does not change existing headers, formulas, or column order.
    Helper columns are appended at the far right.
    """
    headers = _header_map(ws)

    # Section header style for row 1
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28

    # Number formats
    for name in ("implied_prob", "model_prob", "edge", "ev"):
        col = headers.get(name)
        if col:
            for r in range(2, ws.max_row + 1):
                try:
                    ws[f"{col}{r}"].number_format = FORMAT_PERCENTAGE_00
                except Exception:
                    continue

    # odds/price formatting as integer-like
    for name in ("odds", "price"):
        col = headers.get(name)
        if col:
            for r in range(2, ws.max_row + 1):
                try:
                    ws[f"{col}{r}"].number_format = FORMAT_NUMBER_00
                except Exception:
                    continue

    # Column widths (wider for names/notes, narrower for numeric)
    width_map = {
        "away_team": 22,
        "home_team": 22,
        "selection": 18,
        "notes": 36,
        "line": 10,
        "stake": 12,
        "odds": 8,
        "price": 8,
        "implied_prob": 10,
    }
    for name, w in width_map.items():
        col = headers.get(name)
        if col:
            ws.column_dimensions[col].width = w

    # Freeze panes so headers and identity columns remain visible. Freeze at H2 (A-G frozen).
    try:
        ws.freeze_panes = "H2"
    except Exception:
        pass

    # Visual divider border after inputs (after 'selection' column)
    sel_col = headers.get("selection")
    if sel_col:
        # find next column letter
        from openpyxl.utils import column_index_from_string, get_column_letter

        idx = column_index_from_string(sel_col) + 1
        sep_col = get_column_letter(idx)
        thin = Side(border_style="thin", color="444444")
        border = Border(left=thin)
        # apply left border to separator column for header and rows
        for r in range(1, ws.max_row + 1):
            try:
                ws[f"{sep_col}{r}"].border = border
            except Exception:
                continue

    # Conditional formatting for EDGE and EV
    for name in ("edge", "ev"):
        col = headers.get(name)
        if not col:
            continue
        rng = f"{col}2:{col}{ws.max_row}"
        # Green >= 2%
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.02"], stopIfTrue=True, fill=green_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0","0.02"], stopIfTrue=True, fill=yellow_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=True, fill=red_fill))

    # Data validation lists (only if those headers exist)
    def _add_list_validation(header_name: str, choices: list[str]):
        col = headers.get(header_name)
        if not col:
            return
            dv = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True)
            ws.add_data_validation(dv)
            # Use DataValidation.add() which works across openpyxl versions
            try:
                dv.add(f"{col}2:{col}{ws.max_row}")
            except Exception:
                # fallback: assign via ranges attribute if available
                try:
                    dv.ranges = f"{col}2:{col}{ws.max_row}"
                except Exception:
                    pass

    _add_list_validation("market_type", ["ML", "spread", "total"])
    _add_list_validation("log_status", ["planned", "placed", "settled", "ignored"])
    _add_list_validation("bet_result", ["W", "L", "P", "VOID"])

    # Helper columns appended at far right (ADD-ONLY)
    helper_headers = [
        "american_odds_norm",
        "decimal_odds",
        "payout_mult",
        "ev_per_unit",
        "exp_profit_stake",
    ]
    # compute start column index
    start_col_idx = ws.max_column + 1
    from openpyxl.utils import get_column_letter

    for i, h in enumerate(helper_headers, start=start_col_idx):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}1"].value = h
        ws[f"{col_letter}1"].font = header_font
        ws.column_dimensions[col_letter].width = 14

    # locate source columns for american odds (price then odds)
    price_col = headers.get("price")
    odds_col = headers.get("odds")
    ev_col = headers.get("ev")
    stake_col = headers.get("stake")

    for r in range(2, ws.max_row + 1):
        # american_odds_norm = IF(price<>'', price, odds)
        target = get_column_letter(start_col_idx)
        cell = ws[f"{target}{r}"]
        if price_col and odds_col:
            cell.value = f"=IF(TRIM(IFERROR({price_col}{r},\"\"))<>\"\", {price_col}{r}, {odds_col}{r})"
        elif price_col:
            cell.value = f"={price_col}{r}"
        elif odds_col:
            cell.value = f"={odds_col}{r}"
        else:
            cell.value = ""

        # decimal_odds
        target = get_column_letter(start_col_idx + 1)
        cell = ws[f"{target}{r}"]
        amer = get_column_letter(start_col_idx) + str(r)
        # handle blank and numeric conversion
        cell.value = (
            f"=IF({amer}=\"\" , \"\", IF(VALUE({amer})>0, 1+VALUE({amer})/100, 1+100/ABS(VALUE({amer}))))"
        )
        cell.number_format = FORMAT_NUMBER_00

        # payout_mult = decimal_odds - 1
        target = get_column_letter(start_col_idx + 2)
        cell = ws[f"{target}{r}"]
        dec = get_column_letter(start_col_idx + 1) + str(r)
        cell.value = f"=IF({dec}=\"\" , \"\", {dec}-1)"
        cell.number_format = FORMAT_NUMBER_00

        # ev_per_unit: copy ev (formatted as percent)
        target = get_column_letter(start_col_idx + 3)
        cell = ws[f"{target}{r}"]
        if ev_col:
            cell.value = f"=IF({ev_col}{r}=\"\" , \"\", {ev_col}{r})"
            cell.number_format = FORMAT_PERCENTAGE_00
        else:
            cell.value = ""

        # exp_profit_stake = IF(stake blank, blank, stake * ev)
        target = get_column_letter(start_col_idx + 4)
        cell = ws[f"{target}{r}"]
        if stake_col and ev_col:
            cell.value = f"=IF(OR({stake_col}{r}=\"\", {stake_col}{r}=0), \"\", {stake_col}{r}*{ev_col}{r})"
            cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        else:
            cell.value = ""
