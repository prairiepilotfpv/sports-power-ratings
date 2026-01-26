"""System validation report for tuning, ensembles, and EV calculations."""

from __future__ import annotations

import json
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

from models.registry import list_models, list_backtest_models, get_backtest_model
from pipelines.market_config import (
    get_all_market_specs,
    resolve_market_params,
    validate_market_isolation,
)
from pipelines.model_params import resolve_model_market_params_with_metadata
from pipelines.market_utils import _metric_name_for_market
from utils.odds import american_to_implied, expected_value
from backtest.runner import load_games_df_from_db, run_backtest


VALID_MARKETS = ("ML", "SPREAD", "TOTAL")


@dataclass
class ValidationReportOutputs:
    report_path: Path
    summary_path: Path | None
    workbook_path: Path
    frame_paths: dict[str, Path]


def _utcnow_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _safe_json_loads(raw: str | None, *, default: Any) -> Any:
    if not raw:
        return default
    try:
        return json.loads(raw)
    except Exception:
        return default


def _key_value_frame(data: dict[str, Any] | None) -> pd.DataFrame:
    rows = []
    for key, value in (data or {}).items():
        rows.append({"key": key, "value": value})
    return pd.DataFrame(rows, columns=["key", "value"])


def _list_frame(values: list[Any] | None, *, column: str) -> pd.DataFrame:
    rows = [{column: value} for value in (values or [])]
    return pd.DataFrame(rows, columns=[column])


def _issues_frame(issues: list[str] | None) -> pd.DataFrame:
    return pd.DataFrame([{"issue": issue} for issue in (issues or [])], columns=["issue"])


def _market_issue_frame(issues: dict[str, list[str]] | None) -> pd.DataFrame:
    rows = []
    for market, items in (issues or {}).items():
        for item in items:
            rows.append({"market": market, "issue": item})
    return pd.DataFrame(rows, columns=["market", "issue"])


def _table_counts_frame(counts: dict[str, int] | None) -> pd.DataFrame:
    rows = [{"table": key, "rows": value} for key, value in (counts or {}).items()]
    return pd.DataFrame(rows, columns=["table", "rows"])


def _history_frame(
    *,
    db_path: str | Path | None,
    sport: str,
    season: str,
    limit: int | None = None,
) -> pd.DataFrame:
    if db_path is None:
        return pd.DataFrame(
            columns=[
                "run_id",
                "created_at",
                "issues_count",
                "issues_sample",
                "workbook_path",
                "report_path",
                "summary_path",
                "artifacts_dir",
            ]
        )
    db_path = Path(db_path)
    if not db_path.exists():
        return pd.DataFrame(
            columns=[
                "run_id",
                "created_at",
                "issues_count",
                "issues_sample",
                "workbook_path",
                "report_path",
                "summary_path",
                "artifacts_dir",
            ]
        )
    rows: list[dict[str, Any]] = []
    try:
        with sqlite3.connect(db_path) as conn:
            tables = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
            if "validation_runs" not in tables:
                return pd.DataFrame(
                    columns=[
                        "run_id",
                        "created_at",
                        "issues_count",
                        "issues_sample",
                        "workbook_path",
                        "report_path",
                        "summary_path",
                        "artifacts_dir",
                    ]
                )
            cols = [row[1] for row in conn.execute("PRAGMA table_info(validation_runs)")]
            has_workbook = "workbook_path" in cols
            q = f"""
                SELECT run_id, created_at, summary_json, issues_json,
                       {"workbook_path" if has_workbook else "NULL AS workbook_path"},
                       report_path, summary_path, artifacts_dir
                FROM validation_runs
                WHERE sport = ? AND season = ?
                ORDER BY datetime(created_at) DESC
            """
            params: tuple[Any, ...] = (sport, season)
            if limit is not None:
                q += " LIMIT ?"
                params = (sport, season, int(limit))
            for row in conn.execute(q, params).fetchall():
                (
                    run_id,
                    created_at,
                    summary_json,
                    issues_json,
                    workbook_path,
                    report_path,
                    summary_path,
                    artifacts_dir,
                ) = row
                summary = _safe_json_loads(summary_json, default={})
                issues = _safe_json_loads(issues_json, default=[])
                entry: dict[str, Any] = {
                    "run_id": run_id,
                    "created_at": created_at,
                    "issues_count": len(issues) if isinstance(issues, list) else None,
                    "issues_sample": "; ".join(issues[:3]) if isinstance(issues, list) and issues else None,
                    "workbook_path": workbook_path,
                    "report_path": report_path,
                    "summary_path": summary_path,
                    "artifacts_dir": artifacts_dir,
                }
                if isinstance(summary, dict):
                    for key, value in summary.items():
                        entry[f"summary_{key}"] = value
                rows.append(entry)
    except Exception:
        return pd.DataFrame(
            columns=[
                "run_id",
                "created_at",
                "issues_count",
                "issues_sample",
                "workbook_path",
                "report_path",
                "summary_path",
                "artifacts_dir",
            ]
        )
    return pd.DataFrame(rows)


def _table_counts(conn: sqlite3.Connection) -> dict[str, int]:
    tables = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()
    counts: dict[str, int] = {}
    for (name,) in tables:
        try:
            cnt = conn.execute(f"SELECT COUNT(*) FROM \"{name}\"").fetchone()[0]
        except sqlite3.Error:
            cnt = -1
        counts[name] = int(cnt)
    return counts


def _resolve_models(conn: sqlite3.Connection, *, sport: str, season: str) -> list[str]:
    try:
        rows = conn.execute(
            "SELECT DISTINCT model FROM model_metrics WHERE sport = ? AND season = ?",
            (sport, season),
        ).fetchall()
        models = sorted({row[0] for row in rows if row and row[0]})
        if models:
            return models
    except sqlite3.Error:
        pass
    return list_models()


def _normalize_metric(metric_optimized: str | None, market: str) -> str | None:
    if metric_optimized:
        return metric_optimized
    return f"backtest_{_metric_name_for_market(market)}"


def _group_best_runs(rows: list[dict], *, key_fields: tuple[str, ...], top_n: int) -> list[dict]:
    grouped: dict[tuple, list[dict]] = {}
    for row in rows:
        key = tuple(row.get(field) for field in key_fields)
        grouped.setdefault(key, []).append(row)
    best_rows: list[dict] = []
    for key, items in grouped.items():
        items_sorted = sorted(
            items,
            key=lambda item: (
                float("inf") if item.get("best_score") is None else float(item["best_score"]),
                item.get("finished_at") or "",
            ),
        )
        best_rows.extend(items_sorted[:top_n])
    return best_rows


def _derive_run_id(
    *,
    games_df: pd.DataFrame,
    window: str,
    start_date: str | None,
    end_date: str | None,
    rolling_days: int | None,
    rolling_games: int | None,
) -> str | None:
    if games_df.empty or "date" not in games_df.columns:
        return None
    dates = pd.to_datetime(games_df["date"], errors="coerce").dropna()
    if dates.empty:
        return None
    start_dt = pd.to_datetime(start_date).normalize() if start_date else dates.min().normalize()
    end_dt = pd.to_datetime(end_date).normalize() if end_date else dates.max().normalize()
    if pd.isna(start_dt) or pd.isna(end_dt):
        return None
    start_label = start_dt.date().isoformat()
    end_label = end_dt.date().isoformat()
    if window == "expanding":
        return f"{start_label}_to_{end_label}_expanding"
    details: list[str] = []
    if rolling_days is not None:
        details.append(f"{rolling_days}d")
    if rolling_games is not None:
        details.append(f"{rolling_games}g")
    detail_label = "_".join(details) if details else "rolling"
    return f"{start_label}_to_{end_label}_rolling_{detail_label}"


def _load_scored_games(
    *, db_path: Path, sport: str, season: str
) -> pd.DataFrame:
    games = load_games_df_from_db(db_path, sport=sport, season=season)
    if games.empty:
        return games
    scored = games.copy()
    scored["home_score"] = pd.to_numeric(scored.get("home_score"), errors="coerce")
    scored["away_score"] = pd.to_numeric(scored.get("away_score"), errors="coerce")
    scored = scored[scored["home_score"].notna() & scored["away_score"].notna()].copy()
    return scored


def _run_backtests(
    *,
    games_df: pd.DataFrame,
    db_path: Path,
    sport: str,
    season: str,
    models: list[str],
    markets: tuple[str, ...],
    output_dir: Path | None,
    window: str,
    start_date: str | None,
    end_date: str | None,
    rolling_days: int | None,
    rolling_games: int | None,
) -> tuple[list[dict[str, Any]], list[pd.DataFrame], list[str]]:
    results: list[dict[str, Any]] = []
    calibration_frames: list[pd.DataFrame] = []
    issues: list[str] = []

    if games_df.empty:
        issues.append("backtest_no_scored_games")
        return results, calibration_frames, issues

    try:
        from data.validation import validate_dataset

        games_df = validate_dataset(games_df)
    except Exception as exc:
        issues.append(f"backtest_dataset_invalid={exc}")
        return results, calibration_frames, issues

    for model in models:
        try:
            model_cls = get_backtest_model(model)
        except Exception as exc:
            issues.append(f"backtest_model_unavailable={model}: {exc}")
            continue

        for market in markets:
            try:
                resolution = resolve_model_market_params_with_metadata(
                    model,
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    market=market,
                )
                params = resolution.params or {}
                model_factory = lambda: model_cls(**params)
            except Exception as exc:
                issues.append(f"backtest_params_invalid={model}/{market}: {exc}")
                continue

            model_out_dir = None
            if output_dir is not None:
                model_out_dir = output_dir / market.lower() / model

            try:
                backtest_result = run_backtest(
                    model_factory=model_factory,
                    games_df=games_df,
                    start_date=start_date,
                    end_date=end_date,
                    window=window,
                    rolling_days=rolling_days,
                    rolling_games=rolling_games,
                    output_dir=model_out_dir,
                    model_name=model,
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    calibrate=False,
                )
            except Exception as exc:
                issues.append(f"backtest_failed={model}/{market}: {exc}")
                results.append(
                    {
                        "model": model,
                        "market": market,
                        "run_id": None,
                        "params_source": resolution.params_source_label,
                        "metric_optimized": resolution.metric_optimized,
                        "error": str(exc),
                    }
                )
                continue

            if isinstance(backtest_result, tuple):
                outputs, run_id, _, _ = backtest_result
            else:
                outputs = backtest_result
                run_id = _derive_run_id(
                    games_df=games_df,
                    window=window,
                    start_date=start_date,
                    end_date=end_date,
                    rolling_days=rolling_days,
                    rolling_games=rolling_games,
                )

            metrics = (
                outputs.metrics_overall.iloc[0].to_dict()
                if not outputs.metrics_overall.empty
                else {}
            )
            metrics.update(
                {
                    "model": model,
                    "market": market,
                    "run_id": run_id,
                    "params_source": resolution.params_source_label,
                    "metric_optimized": resolution.metric_optimized,
                    "params_fingerprint": resolution.params_fingerprint,
                    "params_nonempty": resolution.params_nonempty,
                    "window": window,
                    "rolling_days": rolling_days,
                    "rolling_games": rolling_games,
                    "output_dir": str(outputs.output_dir),
                }
            )
            results.append(metrics)

            if not outputs.calibration.empty:
                calib_df = outputs.calibration.copy()
                calib_df["model"] = model
                calib_df["market"] = market
                calib_df["run_id"] = run_id
                calibration_frames.append(calib_df)

    return results, calibration_frames, issues


def _validate_ev_rows(
    rows: list[tuple],
    *,
    source: str,
    tolerance: float = 1e-6,
) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, Any]] = {}

    def ensure_market(market: str) -> dict[str, Any]:
        entry = stats.get(market)
        if entry is None:
            entry = {
                "source": source,
                "market_type": market,
                "rows_total": 0,
                "rows_checked": 0,
                "missing_model_prob": 0,
                "missing_implied_prob": 0,
                "edge_mismatch_count": 0,
                "ev_mismatch_count": 0,
                "edge_diff_max_abs": None,
                "ev_diff_max_abs": None,
                "edge_diff_mean": None,
                "ev_diff_mean": None,
                "_edge_diff_sum": 0.0,
                "_ev_diff_sum": 0.0,
                "_edge_diff_count": 0,
                "_ev_diff_count": 0,
            }
            stats[market] = entry
        return entry

    for market_type, odds, implied_prob, model_prob, edge, ev in rows:
        market = str(market_type or "unknown").lower()
        entry = ensure_market(market)
        entry["rows_total"] += 1

        try:
            model_prob = float(model_prob) if model_prob is not None else None
        except Exception:
            model_prob = None
        if model_prob is None:
            entry["missing_model_prob"] += 1
            continue

        implied = None
        if implied_prob is not None:
            try:
                implied = float(implied_prob)
            except Exception:
                implied = None
        if implied is None and odds is not None:
            try:
                implied = american_to_implied(int(odds))
            except Exception:
                implied = None
        if implied is None:
            entry["missing_implied_prob"] += 1
            continue

        entry["rows_checked"] += 1
        expected_edge = model_prob - implied
        try:
            expected_ev = expected_value(implied, model_prob, odds=int(odds) if odds is not None else None)
        except Exception:
            expected_ev = None

        if edge is not None:
            try:
                edge_val = float(edge)
                diff = edge_val - expected_edge
                entry["_edge_diff_sum"] += diff
                entry["_edge_diff_count"] += 1
                max_abs = entry["edge_diff_max_abs"]
                diff_abs = abs(diff)
                entry["edge_diff_max_abs"] = diff_abs if max_abs is None else max(max_abs, diff_abs)
                if diff_abs > tolerance:
                    entry["edge_mismatch_count"] += 1
            except Exception:
                pass

        if ev is not None and expected_ev is not None:
            try:
                ev_val = float(ev)
                diff = ev_val - expected_ev
                entry["_ev_diff_sum"] += diff
                entry["_ev_diff_count"] += 1
                max_abs = entry["ev_diff_max_abs"]
                diff_abs = abs(diff)
                entry["ev_diff_max_abs"] = diff_abs if max_abs is None else max(max_abs, diff_abs)
                if diff_abs > tolerance:
                    entry["ev_mismatch_count"] += 1
            except Exception:
                pass

    for entry in stats.values():
        if entry["_edge_diff_count"]:
            entry["edge_diff_mean"] = entry["_edge_diff_sum"] / entry["_edge_diff_count"]
        if entry["_ev_diff_count"]:
            entry["ev_diff_mean"] = entry["_ev_diff_sum"] / entry["_ev_diff_count"]
        entry.pop("_edge_diff_sum", None)
        entry.pop("_edge_diff_count", None)
        entry.pop("_ev_diff_sum", None)
        entry.pop("_ev_diff_count", None)

    return sorted(stats.values(), key=lambda item: item["market_type"])


def build_validation_report(
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    days_back: int = 7,
    top_n: int = 5,
    run_backtests: bool = True,
    backtest_models: list[str] | None = None,
    backtest_output_dir: Path | None = None,
    backtest_window: str = "expanding",
    backtest_start: str | None = None,
    backtest_end: str | None = None,
    backtest_rolling_days: int | None = None,
    backtest_rolling_games: int | None = None,
) -> tuple[dict[str, Any], dict[str, pd.DataFrame]]:
    report: dict[str, Any] = {
        "sport": sport,
        "season": season,
        "db_path": str(db_path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "issues": [],
    }
    frames: dict[str, pd.DataFrame] = {}
    if run_backtests:
        report["backtest"] = {
            "window": backtest_window,
            "start_date": backtest_start,
            "end_date": backtest_end,
            "rolling_days": backtest_rolling_days,
            "rolling_games": backtest_rolling_games,
            "models": backtest_models or list_backtest_models(),
        }

    db_path = Path(db_path)
    if not db_path.exists():
        report["issues"].append(f"db_not_found={db_path}")
        return report, frames

    with sqlite3.connect(db_path) as conn:
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        table_counts = _table_counts(conn)
        report["table_counts"] = table_counts

        models = _resolve_models(conn, sport=sport, season=season)
        report["models"] = models

        # Market isolation issues
        market_issues = validate_market_isolation(
            db_path=db_path, sport=sport, season=season, models=models
        )
        report["market_isolation_issues"] = market_issues

        # Model params resolutions
        param_rows: list[dict[str, Any]] = []
        for market in VALID_MARKETS:
            for model in models:
                resolution = resolve_market_params(
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    model=model,
                    market=market,
                )
                param_rows.append(
                    {
                        "market": market,
                        "model": resolution.model,
                        "params_source": resolution.params_source_label,
                        "metric_optimized": resolution.metric_optimized,
                        "source_run_id": resolution.source_run_id,
                        "best_score": resolution.best_score,
                        "params_nonempty": resolution.params_nonempty,
                    }
                )
        model_params_df = pd.DataFrame(param_rows)
        frames["model_params"] = model_params_df

        # Ensemble specs and weight checks
        ensemble_specs = get_all_market_specs(
            db_path=db_path, sport=sport, season=season
        )
        ensemble_rows: list[dict[str, Any]] = []
        ensemble_issues: list[str] = []
        for market, spec in ensemble_specs.items():
            weights = spec.weights or {}
            weight_sum = sum(weights.values()) if weights else None
            weights_valid = None
            if weights:
                weights_valid = abs(weight_sum - 1.0) <= 1e-6
                if not weights_valid:
                    ensemble_issues.append(
                        f"{market} ensemble weights sum to {weight_sum:.6f}"
                    )
            weights_missing = (
                [m for m in spec.models if m not in weights] if weights else []
            )
            if weights_missing:
                ensemble_issues.append(
                    f"{market} ensemble missing weights for models: {weights_missing}"
                )
            ensemble_rows.append(
                {
                    "market": spec.market,
                    "ensemble_id": spec.ensemble_id,
                    "metric_slot": spec.metric_slot,
                    "weights_source": spec.weights_source,
                    "source_run_id": spec.source_run_id,
                    "models": ",".join(spec.models),
                    "weights_json": json.dumps(weights) if weights else None,
                    "weight_sum": weight_sum,
                    "weights_valid": weights_valid,
                    "weights_missing_models": ",".join(weights_missing) if weights_missing else None,
                    "config_path": spec.config_path,
                }
            )
        frames["ensemble_specs"] = pd.DataFrame(ensemble_rows)
        if ensemble_issues:
            report["issues"].extend(ensemble_issues)

        # Model tuning runs leaderboard
        model_tuning_rows: list[dict[str, Any]] = []
        if "model_market_tuning_runs" in tables:
            rows = conn.execute(
                """
                SELECT model, market, metric_optimized, run_id, best_score,
                       best_params_json, summary_metrics_json, finished_at
                FROM model_market_tuning_runs
                WHERE sport = ? AND season = ?
                """,
                (sport, season),
            ).fetchall()
            for model, market, metric_optimized, run_id, best_score, params_json, summary_json, finished_at in rows:
                model_tuning_rows.append(
                    {
                        "model": model,
                        "market": market,
                        "metric_optimized": _normalize_metric(metric_optimized, market),
                        "run_id": run_id,
                        "best_score": best_score,
                        "best_params_json": params_json,
                        "summary_metrics_json": summary_json,
                        "finished_at": finished_at,
                    }
                )
        if model_tuning_rows:
            top_rows = _group_best_runs(
                model_tuning_rows,
                key_fields=("model", "market", "metric_optimized"),
                top_n=top_n,
            )
            frames["model_tuning_runs"] = pd.DataFrame(top_rows)
        else:
            report["issues"].append("no_model_tuning_runs")

        # Ensemble tuning runs leaderboard
        ensemble_tuning_rows: list[dict[str, Any]] = []
        if "ensemble_market_tuning_runs" in tables:
            rows = conn.execute(
                """
                SELECT market, ensemble_id, metric_optimized, run_id, best_score,
                       models_json, weights_json, finished_at
                FROM ensemble_market_tuning_runs
                WHERE sport = ? AND season = ?
                """,
                (sport, season),
            ).fetchall()
            for market, ensemble_id, metric_optimized, run_id, best_score, models_json, weights_json, finished_at in rows:
                ensemble_tuning_rows.append(
                    {
                        "market": market,
                        "ensemble_id": ensemble_id,
                        "metric_optimized": _normalize_metric(metric_optimized, market),
                        "run_id": run_id,
                        "best_score": best_score,
                        "models_json": models_json,
                        "weights_json": weights_json,
                        "finished_at": finished_at,
                    }
                )
        if ensemble_tuning_rows:
            top_rows = _group_best_runs(
                ensemble_tuning_rows,
                key_fields=("market", "metric_optimized"),
                top_n=top_n,
            )
            frames["ensemble_tuning_runs"] = pd.DataFrame(top_rows)
        else:
            report["issues"].append("no_ensemble_tuning_runs")

        # EV consistency checks
        ev_rows: list[dict[str, Any]] = []
        if "opportunities" in tables:
            opp_rows = conn.execute(
                """
                SELECT o.market_type, o.odds, o.implied_prob, o.model_prob, o.edge, o.ev
                FROM opportunities o
                JOIN games g ON g.game_id = o.game_id
                WHERE g.sport = ? AND g.season = ?
                """,
                (sport, season),
            ).fetchall()
            ev_rows.extend(_validate_ev_rows(opp_rows, source="opportunities"))
        else:
            report["issues"].append("opportunities_table_missing")

        if "bets" in tables:
            bet_rows = conn.execute(
                """
                SELECT b.market_type, b.odds, NULL AS implied_prob, b.model_prob, b.edge, b.ev
                FROM bets b
                JOIN games g ON g.game_id = b.game_id
                WHERE g.sport = ? AND g.season = ?
                """,
                (sport, season),
            ).fetchall()
            ev_rows.extend(_validate_ev_rows(bet_rows, source="bets"))
        else:
            report["issues"].append("bets_table_missing")

        if ev_rows:
            frames["ev_validation"] = pd.DataFrame(ev_rows)

    # Optional: fresh backtests for accuracy + calibration metrics.
    if run_backtests:
        scored_games = _load_scored_games(db_path=db_path, sport=sport, season=season)
        bt_models = backtest_models or list_backtest_models()
        bt_results, bt_calibration_frames, bt_issues = _run_backtests(
            games_df=scored_games,
            db_path=db_path,
            sport=sport,
            season=season,
            models=bt_models,
            markets=VALID_MARKETS,
            output_dir=backtest_output_dir,
            window=backtest_window,
            start_date=backtest_start,
            end_date=backtest_end,
            rolling_days=backtest_rolling_days,
            rolling_games=backtest_rolling_games,
        )
        if bt_results:
            frames["backtest_metrics"] = pd.DataFrame(bt_results)
        if bt_calibration_frames:
            frames["backtest_calibration"] = pd.concat(bt_calibration_frames, ignore_index=True)
        if bt_issues:
            report["issues"].extend(bt_issues)

    # Optional: ML ensemble weight validation (uses bets_predictions)
    try:
        from pipelines.ensemble_weight_validation import validate_ensemble_ml_weights

        report["ensemble_weight_validation"] = validate_ensemble_ml_weights(
            db_path=db_path,
            sport=sport,
            season=season,
            market="ML",
            days_back=days_back,
        )
    except Exception as exc:
        report["issues"].append(f"ensemble_weight_validation_failed={exc}")

    # Summary
    summary: dict[str, Any] = {}
    summary["models_count"] = len(report.get("models", []))
    summary["market_issue_count"] = sum(len(v) for v in market_issues.values()) if market_issues else 0
    if "model_params" in frames:
        missing = frames["model_params"]["params_source"].eq("missing_active").sum()
        summary["missing_model_params"] = int(missing)
    if "ensemble_specs" in frames:
        weight_invalid = frames["ensemble_specs"]["weights_valid"].eq(False).sum()
        summary["ensemble_weights_invalid"] = int(weight_invalid)
    if "ev_validation" in frames:
        ev_df = frames["ev_validation"]
        summary["ev_mismatch_count"] = int(ev_df["ev_mismatch_count"].sum())
        summary["edge_mismatch_count"] = int(ev_df["edge_mismatch_count"].sum())
    if "backtest_metrics" in frames:
        bt_df = frames["backtest_metrics"]
        summary["backtest_runs"] = int(len(bt_df))
        if "error" in bt_df.columns:
            summary["backtest_failures"] = int(bt_df["error"].notna().sum())
    report["summary"] = summary

    return report, frames


def _autosize_worksheet(ws, *, min_width: int = 10, max_width: int = 80) -> None:
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    for idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column:
            val = cell.value
            if val is None:
                continue
            try:
                text = str(val)
            except Exception:
                text = ""
            if len(text) > max_len:
                max_len = len(text)
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_diagnostics_sheet(
    writer: pd.ExcelWriter,
    *,
    sections: list[tuple[str, pd.DataFrame]],
) -> None:
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        ws = writer.book.create_sheet("DIAGNOSTICS")
        writer.sheets["DIAGNOSTICS"] = ws
        row_cursor = 0
        for title, frame in sections:
            ws.cell(row=row_cursor + 1, column=1, value=title)
            row_cursor += 1
            df = frame if frame is not None else pd.DataFrame()
            if df.empty:
                ws.cell(row=row_cursor + 1, column=1, value="(no data)")
                row_cursor += 2
                continue
            df.to_excel(writer, sheet_name="DIAGNOSTICS", startrow=row_cursor, startcol=0, index=False)
            row_cursor += len(df.index) + 2
        return

    ws = writer.book.create_sheet("DIAGNOSTICS")
    writer.sheets["DIAGNOSTICS"] = ws

    heading_font = Font(bold=True, color="FFFFFF")
    heading_fill = PatternFill("solid", fgColor="4F81BD")
    heading_alignment = Alignment(vertical="center")

    table_header_font = Font(bold=True)
    table_header_fill = PatternFill("solid", fgColor="D9D9D9")
    table_header_alignment = Alignment(vertical="center")

    row_cursor = 0
    for title, frame in sections:
        heading_cell = ws.cell(row=row_cursor + 1, column=1, value=title)
        heading_cell.font = heading_font
        heading_cell.fill = heading_fill
        heading_cell.alignment = heading_alignment
        row_cursor += 1

        df = frame if frame is not None else pd.DataFrame()
        if df.empty:
            ws.cell(row=row_cursor + 1, column=1, value="(no data)")
            row_cursor += 2
            continue

        df.to_excel(writer, sheet_name="DIAGNOSTICS", startrow=row_cursor, startcol=0, index=False)
        header_row = row_cursor + 1
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = table_header_alignment

        row_cursor += len(df.index) + 2


def _write_validation_workbook(
    *,
    workbook_path: Path,
    report: dict[str, Any],
    frames: dict[str, pd.DataFrame],
    db_path: str | Path | None,
    run_id: str,
    sport: str,
    season: str,
    report_path: Path,
    summary_path: Path | None,
) -> None:
    meta = _key_value_frame(
        {
            "run_id": run_id,
            "sport": sport,
            "season": season,
            "generated_at": report.get("generated_at"),
            "db_path": report.get("db_path"),
            "report_path": str(report_path),
            "summary_path": str(summary_path) if summary_path else None,
            "workbook_path": str(workbook_path),
        }
    )
    summary_df = _key_value_frame(report.get("summary"))
    issues_df = _issues_frame(report.get("issues"))
    models_df = _list_frame(report.get("models"), column="model")
    table_counts_df = _table_counts_frame(report.get("table_counts"))
    market_issue_df = _market_issue_frame(report.get("market_isolation_issues"))
    backtest_config_df = _key_value_frame(report.get("backtest"))

    ew_validation = report.get("ensemble_weight_validation")
    ew_summary_df = pd.DataFrame()
    ew_by_date_df = pd.DataFrame()
    if isinstance(ew_validation, dict):
        ew_summary = {k: v for k, v in ew_validation.items() if k != "by_date"}
        ew_summary_df = pd.DataFrame([ew_summary]) if ew_summary else pd.DataFrame()
        by_date = ew_validation.get("by_date")
        if isinstance(by_date, list):
            ew_by_date_df = pd.DataFrame(by_date)

    history_df = _history_frame(db_path=db_path, sport=sport, season=season)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        _write_diagnostics_sheet(
            writer,
            sections=[
                ("META", meta),
                ("SUMMARY", summary_df),
                ("ISSUES", issues_df),
                ("MODELS", models_df),
                ("TABLE_COUNTS", table_counts_df),
                ("MARKET_ISSUES", market_issue_df),
                ("BACKTEST_CONFIG", backtest_config_df),
            ],
        )

        frame_map = {
            "MODEL_PARAMS": frames.get("model_params"),
            "ENSEMBLE_SPECS": frames.get("ensemble_specs"),
            "MODEL_TUNING_RUNS": frames.get("model_tuning_runs"),
            "ENSEMBLE_TUNING_RUNS": frames.get("ensemble_tuning_runs"),
            "EV_VALIDATION": frames.get("ev_validation"),
            "BACKTEST_METRICS": frames.get("backtest_metrics"),
            "BACKTEST_CALIBRATION": frames.get("backtest_calibration"),
        }
        for sheet, frame in frame_map.items():
            (frame if frame is not None else pd.DataFrame()).to_excel(
                writer, sheet_name=sheet, index=False
            )

        ew_summary_df.to_excel(writer, sheet_name="ENSEMBLE_WEIGHT_SUM", index=False)
        ew_by_date_df.to_excel(writer, sheet_name="ENSEMBLE_WEIGHT_DATE", index=False)
        history_df.to_excel(writer, sheet_name="HISTORY", index=False)

        for ws in writer.sheets.values():
            _autosize_worksheet(ws)


def write_validation_report(
    *,
    report: dict[str, Any],
    frames: dict[str, pd.DataFrame],
    output_dir: str | Path,
    filename_prefix: str,
    db_path: str | Path | None = None,
) -> ValidationReportOutputs:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook_path = output_dir / f"{filename_prefix}_validation.xlsx"
    report["workbook_path"] = str(workbook_path)

    report_path = output_dir / f"{filename_prefix}_validation.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    summary_path = None
    if report.get("summary"):
        summary_path = output_dir / f"{filename_prefix}_summary.csv"
        summary_rows = [{"key": k, "value": v} for k, v in report["summary"].items()]
        pd.DataFrame(summary_rows).to_csv(summary_path, index=False)

    frame_paths: dict[str, Path] = {}
    for name, frame in frames.items():
        if frame is None:
            continue
        out_path = output_dir / f"{filename_prefix}_{name}.csv"
        frame.to_csv(out_path, index=False)
        frame_paths[name] = out_path

    try:
        _write_validation_workbook(
            workbook_path=workbook_path,
            report=report,
            frames=frames,
            db_path=db_path,
            run_id=filename_prefix,
            sport=report.get("sport", ""),
            season=report.get("season", ""),
            report_path=report_path,
            summary_path=summary_path,
        )
    except Exception as exc:
        report.setdefault("issues", []).append(f"workbook_write_failed={exc}")
        report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    return ValidationReportOutputs(
        report_path=report_path,
        summary_path=summary_path,
        workbook_path=workbook_path,
        frame_paths=frame_paths,
    )


def run_validation_report(
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    output_dir: str | Path | None = None,
    days_back: int = 7,
    top_n: int = 5,
    run_backtests: bool = True,
    backtest_models: list[str] | None = None,
    backtest_window: str = "expanding",
    backtest_start: str | None = None,
    backtest_end: str | None = None,
    backtest_rolling_days: int | None = None,
    backtest_rolling_games: int | None = None,
    keep_backtest_artifacts: bool = False,
) -> ValidationReportOutputs:
    prefix = f"{sport}_{season}_{_utcnow_stamp()}"
    out_dir = Path(output_dir) if output_dir is not None else (Path("outputs") / "validation" / sport / season)
    backtest_output_dir = out_dir / "backtests" / prefix if run_backtests else None

    report, frames = build_validation_report(
        db_path=db_path,
        sport=sport,
        season=season,
        days_back=days_back,
        top_n=top_n,
        run_backtests=run_backtests,
        backtest_models=backtest_models,
        backtest_output_dir=backtest_output_dir,
        backtest_window=backtest_window,
        backtest_start=backtest_start,
        backtest_end=backtest_end,
        backtest_rolling_days=backtest_rolling_days,
        backtest_rolling_games=backtest_rolling_games,
    )
    outputs = write_validation_report(
        report=report,
        frames=frames,
        output_dir=out_dir,
        filename_prefix=prefix,
        db_path=db_path,
    )
    try:
        config = {
            "days_back": days_back,
            "top_n": top_n,
            "run_backtests": run_backtests,
            "backtest_models": backtest_models or list_backtest_models(),
            "backtest_window": backtest_window,
            "backtest_start": backtest_start,
            "backtest_end": backtest_end,
            "backtest_rolling_days": backtest_rolling_days,
            "backtest_rolling_games": backtest_rolling_games,
            "keep_backtest_artifacts": keep_backtest_artifacts,
        }
        save_validation_run(
            db_path,
            run_id=prefix,
            sport=sport,
            season=season,
            config=config,
            summary=report.get("summary"),
            issues=report.get("issues"),
            workbook_path=str(outputs.workbook_path),
            report_path=str(outputs.report_path),
            summary_path=str(outputs.summary_path) if outputs.summary_path else None,
            artifacts_dir=str(out_dir),
        )
    except Exception as exc:
        report.setdefault("issues", []).append(f"validation_run_persist_failed={exc}")
        outputs.report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    if run_backtests and backtest_output_dir and not keep_backtest_artifacts:
        shutil.rmtree(backtest_output_dir, ignore_errors=True)
    return outputs


__all__ = [
    "build_validation_report",
    "run_validation_report",
    "write_validation_report",
    "ValidationReportOutputs",
]
from data.repository import save_validation_run
