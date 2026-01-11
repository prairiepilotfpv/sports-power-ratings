"""Review run orchestration (skeleton).

Responsibilities:
- Create a ReviewRun record
- Generate forecast_snapshots from model projections
- Generate opportunities by comparing model probabilities to market odds
- Emit an Excel review workbook with EV and BETS sheets
"""

from __future__ import annotations


import pandas as pd
from datetime import datetime, timezone
from pathlib import Path
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from data.paths import processed_path_for
from data.repository import load_games
from data.betting_repository import (
    get_opportunities_with_game_info,
    get_prediction_exclusions,
    create_review_run,
    init_db,
)


def _resolve_output_path(sport: str, season: str, review_run_id: str) -> Path:
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    filename = f"review_{review_run_id}_{ts}.xlsx"
    return processed_path_for(sport, season, f"review/{filename}")


def _load_ocr_raw_rows(db_path: str | Path, review_run_id: str) -> list[dict]:
    init_db(db_path)
    with sqlite3.connect(Path(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        q = """
            SELECT DISTINCT
                ms.id AS source_market_snapshot_id,
                st.image_path AS image_path,
                st.raw_text AS raw_text,
                st.team_home_raw AS team_home_raw,
                st.team_away_raw AS team_away_raw,
                st.match_status AS match_status,
                st.match_confidence AS match_confidence,
                st.hold_reason AS hold_reason,
                COALESCE(st.captured_at, ms.captured_at) AS captured_at,
                ms.book AS book,
                ms.market_type AS market_type,
                ms.selection AS selection,
                ms.line AS line,
                ms.odds AS odds
            FROM opportunities o
            JOIN market_snapshots ms ON o.source_market_snapshot_id = ms.id
            LEFT JOIN market_snapshot_staging st ON ms.source_staging_id = st.id
            WHERE o.review_run_id = ?
            ORDER BY COALESCE(st.captured_at, ms.captured_at), ms.market_type, ms.selection
        """
        rows = conn.execute(q, (review_run_id,)).fetchall()
        return [dict(r) for r in rows]


def build_review_workbook(
    db_path: str | Path,
    *,
    review_run_id: str,
    sport: str,
    season: str,
    output_path: str | Path | None = None,
    include_ocr_raw: bool = True,
) -> Path:
    """Build a review workbook for a review_run_id.

    - EV sheet is read-only/protected
    - BETS sheet is editable and contains base columns + stake/book/price/bet_id/logged_at/notes
    - META sheet contains key/value pairs (review_run_id, sport, season, created_at)
    """
    out = Path(output_path) if output_path else _resolve_output_path(sport, season, review_run_id)
    out.parent.mkdir(parents=True, exist_ok=True)

    rows = get_opportunities_with_game_info(db_path, review_run_id=review_run_id)

    # Build EV dataframe
    ev_cols = [
        "review_run_id",
        "game_id",
        "date",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "odds",
        "implied_prob",
        "model_prob",
        "edge",
        "ev",
        "book",
        "source_market_snapshot_id",
        "opportunity_id",
    ]
    ev_rows = []
    for r in rows:
        ev_rows.append(
            {
                "review_run_id": r["review_run_id"],
                "game_id": r["game_id"],
                "date": r["date"],
                "away_team": r["away_team"],
                "home_team": r["home_team"],
                "market_type": r["market_type"],
                "selection": r["selection"],
                "line": r["line"],
                "odds": r["odds"],
                "implied_prob": r["implied_prob"],
                "model_prob": r["model_prob"],
                "edge": r["edge"],
                "ev": r["ev"],
                "book": None,
                "source_market_snapshot_id": r["source_market_snapshot_id"],
                "opportunity_id": r["opportunity_id"],
            }
        )

    ev_df = pd.DataFrame(ev_rows, columns=ev_cols)

    # BETS sheet: copy EV base columns and add editable columns
    bets_cols = ev_cols + ["stake", "book", "price", "bet_id", "logged_at", "notes"]
    bets_df = ev_df.copy(deep=True)
    for c in ["stake", "book", "price", "bet_id", "logged_at", "notes"]:
        bets_df[c] = ""

    # META sheet
    meta_rows = [
        {"key": "review_run_id", "value": review_run_id},
        {"key": "sport", "value": sport},
        {"key": "season", "value": season},
        {"key": "created_at", "value": datetime.now(timezone.utc).isoformat()},
    ]
    meta_df = pd.DataFrame(meta_rows)

    ocr_cols = [
        "source_market_snapshot_id",
        "image_path",
        "raw_text",
        "team_home_raw",
        "team_away_raw",
        "match_status",
        "match_confidence",
        "hold_reason",
        "captured_at",
        "book",
        "market_type",
        "selection",
        "line",
        "odds",
    ]
    ocr_df = None
    if include_ocr_raw:
        ocr_rows = _load_ocr_raw_rows(db_path, review_run_id)
        ocr_df = pd.DataFrame(ocr_rows, columns=ocr_cols)

    exclusions = get_prediction_exclusions(db_path, review_run_id=review_run_id)
    exclusions_df = None
    if exclusions:
        exclusions_df = pd.DataFrame(
            exclusions, columns=["game_id", "model", "excluded_reason"]
        )

    # Write via pandas then post-process with openpyxl
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        ev_df.to_excel(writer, sheet_name="EV", index=False)
        bets_df.to_excel(writer, sheet_name="BETS", index=False)
        if ocr_df is not None:
            ocr_df.to_excel(writer, sheet_name="OCR_RAW", index=False)
        if exclusions_df is not None:
            exclusions_df.to_excel(writer, sheet_name="EXCLUSIONS", index=False)
        meta_df.to_excel(writer, sheet_name="META", index=False)

    # Post-process protection and hidden META
    wb = load_workbook(out)
    if "EV" in wb.sheetnames:
        ws = wb["EV"]
        ws.protection.sheet = True
        ws.protection.enable = True
    if "BETS" in wb.sheetnames:
        ws = wb["BETS"]
        ws.protection.sheet = False
    if "META" in wb.sheetnames:
        ws = wb["META"]
        ws.sheet_state = "hidden"

    wb.save(out)
    return out


def create_and_build_review(db_path: str | Path, *, sport: str, season: str, model: str, notes: str | None = None) -> Path:
    """Create a review_run and immediately build the workbook. Returns path."""
    rid = create_review_run(db_path, sport=sport, season=season, model=model, notes=notes)
    return build_review_workbook(db_path, review_run_id=rid, sport=sport, season=season)


def _apply_formula_sheet(ws) -> None:
    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    required = {"odds", "implied_prob", "model_prob", "edge", "ev"}
    if not required.issubset(header):
        return

    def cell_ref(col: int, row: int) -> str:
        return f"{get_column_letter(col)}{row}"

    odds_col = header["odds"]
    implied_col = header["implied_prob"]
    model_col = header["model_prob"]
    edge_col = header["edge"]
    ev_col = header["ev"]

    for row in range(2, ws.max_row + 1):
        odds_cell = cell_ref(odds_col, row)
        implied_cell = cell_ref(implied_col, row)
        model_cell = cell_ref(model_col, row)
        edge_cell = cell_ref(edge_col, row)
        ev_cell = cell_ref(ev_col, row)

        ws[implied_cell].value = (
            f"=IF(OR({odds_cell}=\"\",{odds_cell}=0),\"\",IF({odds_cell}>0,100/({odds_cell}+100),-{odds_cell}/(-{odds_cell}+100)))"
        )
        ws[edge_cell].value = f"=IF(OR({model_cell}=\"\",{implied_cell}=\"\"),\"\",{model_cell}-{implied_cell})"
        ws[ev_cell].value = (
            f"=IF(OR({model_cell}=\"\",{odds_cell}=\"\",{odds_cell}=0),\"\",({model_cell}*IF({odds_cell}>0,{odds_cell}/100,100/ABS({odds_cell})))-(1-{model_cell}))"
        )


def build_review_workbook_with_formulas(
    db_path: str | Path,
    *,
    review_run_id: str,
    sport: str,
    season: str,
    output_path: str | Path | None = None,
    include_ocr_raw: bool = True,
) -> Path:
    """Build a review workbook that includes formulas for implied_prob, edge, and ev."""
    out = Path(output_path) if output_path else _resolve_output_path(sport, season, review_run_id)
    out.parent.mkdir(parents=True, exist_ok=True)

    rows = get_opportunities_with_game_info(db_path, review_run_id=review_run_id)

    ev_cols = [
        "review_run_id",
        "game_id",
        "date",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "odds",
        "implied_prob",
        "model_prob",
        "edge",
        "ev",
        "book",
        "source_market_snapshot_id",
        "opportunity_id",
    ]
    ev_rows = []
    for r in rows:
        ev_rows.append(
            {
                "review_run_id": r["review_run_id"],
                "game_id": r["game_id"],
                "date": r["date"],
                "away_team": r["away_team"],
                "home_team": r["home_team"],
                "market_type": r["market_type"],
                "selection": r["selection"],
                "line": r["line"],
                "odds": r["odds"],
                "implied_prob": r["implied_prob"],
                "model_prob": r["model_prob"],
                "edge": r["edge"],
                "ev": r["ev"],
                "book": None,
                "source_market_snapshot_id": r["source_market_snapshot_id"],
                "opportunity_id": r["opportunity_id"],
            }
        )

    ev_df = pd.DataFrame(ev_rows, columns=ev_cols)

    bets_cols = ev_cols + ["stake", "book", "price", "bet_id", "logged_at", "notes"]
    bets_df = ev_df.copy(deep=True)
    for c in ["stake", "book", "price", "bet_id", "logged_at", "notes"]:
        bets_df[c] = ""

    meta_rows = [
        {"key": "review_run_id", "value": review_run_id},
        {"key": "sport", "value": sport},
        {"key": "season", "value": season},
        {"key": "created_at", "value": datetime.now(timezone.utc).isoformat()},
    ]
    meta_df = pd.DataFrame(meta_rows)

    ocr_cols = [
        "source_market_snapshot_id",
        "image_path",
        "raw_text",
        "team_home_raw",
        "team_away_raw",
        "match_status",
        "match_confidence",
        "hold_reason",
        "captured_at",
        "book",
        "market_type",
        "selection",
        "line",
        "odds",
    ]
    ocr_df = None
    if include_ocr_raw:
        ocr_rows = _load_ocr_raw_rows(db_path, review_run_id)
        ocr_df = pd.DataFrame(ocr_rows, columns=ocr_cols)

    exclusions = get_prediction_exclusions(db_path, review_run_id=review_run_id)
    exclusions_df = None
    if exclusions:
        exclusions_df = pd.DataFrame(
            exclusions, columns=["game_id", "model", "excluded_reason"]
        )

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        ev_df.to_excel(writer, sheet_name="EV", index=False)
        bets_df.to_excel(writer, sheet_name="BETS", index=False)
        if ocr_df is not None:
            ocr_df.to_excel(writer, sheet_name="OCR_RAW", index=False)
        if exclusions_df is not None:
            exclusions_df.to_excel(writer, sheet_name="EXCLUSIONS", index=False)
        meta_df.to_excel(writer, sheet_name="META", index=False)

    wb = load_workbook(out)
    if "EV" in wb.sheetnames:
        ws = wb["EV"]
        _apply_formula_sheet(ws)
        ws.protection.sheet = True
        ws.protection.enable = True
    if "BETS" in wb.sheetnames:
        ws = wb["BETS"]
        _apply_formula_sheet(ws)
        ws.protection.sheet = False
    if "META" in wb.sheetnames:
        ws = wb["META"]
        ws.sheet_state = "hidden"

    wb.save(out)
    return out
