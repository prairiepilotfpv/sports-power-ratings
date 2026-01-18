"""Pipeline to build a unified daily betting workbook."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from data.paths import processed_path_for
from data.repository import load_games
from data import betting_repository as br
from markets.base import Market
from pipelines import schedule as schedule_pipeline
from pipelines.common import normalize_games
from pipelines.model_params import resolve_model_market_params_with_metadata
from pipelines import opportunities as opportunities_pipeline
from pipelines.review_runs import _apply_formula_sheet, _load_ocr_raw_rows
from data.betting_repository import get_opportunities_with_game_info
from contracts import SCHEDULE_EXPORT_COLUMNS


PROJECTIONS_SHEET = "PROJECTIONS"
MARKET_SHEET = "MARKET_SNAPSHOTS"
OCR_SHEET = "OCR_RAW"
EV_SHEET = "EV"
BETS_SHEET = "BETS"
META_SHEET = "META"


def _normalize_date(value: str) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Invalid date: {value!r}")
    return parsed.date().isoformat()


def _resolve_review_run(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    as_of_date: str,
    review_run_id: str | None,
) -> tuple[str, str | None]:
    br.init_db(db_path)
    with sqlite3.connect(Path(db_path)) as conn:
        if review_run_id:
            row = conn.execute(
                "SELECT id, model FROM review_runs WHERE id = ?",
                (review_run_id,),
            ).fetchone()
            if row is None:
                raise ValueError(f"review_run_id {review_run_id!r} not found")
            return str(row[0]), row[1]

        row = conn.execute(
            """
            SELECT id, model
            FROM review_runs
            WHERE sport = ? AND season = ? AND date(created_at) = date(?)
            ORDER BY datetime(created_at) DESC
            LIMIT 1
            """,
            (sport, season, as_of_date),
        ).fetchone()
        if row is None:
            raise ValueError(
                "No review_run found for "
                f"sport={sport!r}, season={season!r}, date={as_of_date!r}"
            )
        return str(row[0]), row[1]


def _build_projections_frame(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    model: str,
    as_of_date: str,
) -> pd.DataFrame:
    rows = load_games(db_path, sport=sport, season=season)
    games_df = normalize_games(rows)
    if games_df.empty:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)

    resolution = resolve_model_market_params_with_metadata(
        model,
        db_path=db_path,
        sport=sport,
        season=season,
        market=Market.ML,
    )
    schedule_df = schedule_pipeline._build_schedule_dataframe(
        games_df,
        db_path=db_path,
        sport=sport,
        season=season,
        model=model,
        upcoming_only=False,
        model_params=resolution.params,
        params_source=resolution.params_source,
        tuned_metric_used=resolution.tuned_metric_used,
        params_run_id=resolution.source_run_id,
        params_market=resolution.market,
    )
    if schedule_df.empty:
        return schedule_df

    target_date = pd.to_datetime(as_of_date, errors="coerce").date()
    schedule_df = schedule_df.assign(
        _date=pd.to_datetime(schedule_df["date"], errors="coerce").dt.date
    )
    filtered = schedule_df[schedule_df["_date"] == target_date].drop(
        columns=["_date"], errors="ignore"
    )
    if filtered.empty:
        return filtered.reindex(columns=SCHEDULE_EXPORT_COLUMNS)
    return filtered.reindex(columns=SCHEDULE_EXPORT_COLUMNS)


def _build_market_snapshots_frame(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    as_of_date: str,
    snapshot_run_id: str | None,
) -> pd.DataFrame:
    if snapshot_run_id:
        market_df = opportunities_pipeline.load_market_snapshots(
            db_path,
            sport=sport,
            season=season,
            snapshot_run_id=snapshot_run_id,
        )
    else:
        market_df = opportunities_pipeline.load_market_snapshots(
            db_path,
            sport=sport,
            season=season,
            snapshot_date=as_of_date,
        )
    if market_df.empty:
        columns = [
            "source_market_snapshot_id",
            "snapshot_run_id",
            "captured_at",
            "book",
            "market_type",
            "selection",
            "line",
            "odds",
            "game_id",
            "home_team",
            "away_team",
            "sport",
            "season",
        ]
        return pd.DataFrame(columns=columns)
    return market_df


def _build_ocr_frame(
    db_path: str | Path,
    *,
    review_run_id: str,
) -> pd.DataFrame:
    rows = _load_ocr_raw_rows(db_path, review_run_id)
    columns = [
        "source_market_snapshot_id",
        "snapshot_run_id",
        "game_id",
        "source_staging_id",
        "staging_source",
        "image_path",
        "raw_text",
        "team_home_raw",
        "team_away_raw",
        "game_date",
        "match_status",
        "match_confidence",
        "hold_reason",
        "captured_at",
        "book",
        "market_type",
        "selection",
        "line",
        "odds",
    ]
    return pd.DataFrame(rows, columns=columns)


def _build_ev_bets_frames(
    db_path: str | Path,
    *,
    review_run_id: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = get_opportunities_with_game_info(db_path, review_run_id=review_run_id)
    ev_cols = [
        "review_run_id",
        "game_id",
        "date",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "odds",
        "implied_prob",
        "model_prob",
        "edge",
        "ev",
        "book",
        "source_market_snapshot_id",
        "opportunity_id",
    ]
    ev_rows: list[dict] = []
    for r in rows:
        ev_rows.append(
            {
                "review_run_id": r["review_run_id"],
                "game_id": r["game_id"],
                "date": r["date"],
                "away_team": r["away_team"],
                "home_team": r["home_team"],
                "market_type": r["market_type"],
                "selection": r["selection"],
                "line": r["line"],
                "odds": r["odds"],
                "implied_prob": r["implied_prob"],
                "model_prob": r["model_prob"],
                "edge": r["edge"],
                "ev": r["ev"],
                "book": None,
                "source_market_snapshot_id": r["source_market_snapshot_id"],
                "opportunity_id": r["opportunity_id"],
            }
        )
    ev_df = pd.DataFrame(ev_rows, columns=ev_cols)

    bets_cols = ev_cols + ["stake", "book", "price", "bet_id", "logged_at", "notes"]
    bets_df = ev_df.copy(deep=True)
    for col in ["stake", "book", "price", "bet_id", "logged_at", "notes"]:
        bets_df[col] = ""
    return ev_df, bets_df


def _add_cross_sheet_links(wb, *, sheet_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        return
    if PROJECTIONS_SHEET not in wb.sheetnames or MARKET_SHEET not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    if ws.max_row < 2:
        return

    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    if "game_id" not in header or "source_market_snapshot_id" not in header:
        return

    proj_ws = wb[PROJECTIONS_SHEET]
    proj_header = {cell.value: cell.column for cell in proj_ws[1] if cell.value}
    snap_ws = wb[MARKET_SHEET]
    snap_header = {cell.value: cell.column for cell in snap_ws[1] if cell.value}

    proj_lookup_col = proj_header.get("game_id")
    snap_lookup_col = snap_header.get("source_market_snapshot_id")
    if proj_lookup_col is None or snap_lookup_col is None:
        return

    proj_columns = [
        ("proj_home_team", "home_team"),
        ("proj_away_team", "away_team"),
        ("proj_projected_spread", "projected_spread"),
        ("proj_model_p_home_win", "model_p_home_win"),
        ("proj_projected_total", "projected_total"),
    ]
    snap_columns = [
        ("snapshot_run_id", "snapshot_run_id"),
        ("snapshot_captured_at", "captured_at"),
        ("snapshot_book", "book"),
        ("snapshot_market_type", "market_type"),
        ("snapshot_selection", "selection"),
        ("snapshot_line", "line"),
        ("snapshot_odds", "odds"),
    ]

    start_col = ws.max_column + 1
    game_id_col = header["game_id"]
    snapshot_id_col = header["source_market_snapshot_id"]

    def _xlookup_formula(
        lookup_cell: str,
        *,
        sheet: str,
        lookup_col: int,
        return_col: int,
    ) -> str:
        lookup_letter = get_column_letter(lookup_col)
        return_letter = get_column_letter(return_col)
        return (
            f"=IF({lookup_cell}=\"\",\"\","
            f"IFERROR(XLOOKUP({lookup_cell},'{sheet}'!${lookup_letter}:${lookup_letter},"
            f"'{sheet}'!${return_letter}:${return_letter},\"\"),\"\"))"
        )

    current_col = start_col
    for header_name, proj_field in proj_columns:
        return_col = proj_header.get(proj_field)
        if return_col is None:
            continue
        ws.cell(row=1, column=current_col, value=header_name)
        for row in range(2, ws.max_row + 1):
            lookup_cell = f"{get_column_letter(game_id_col)}{row}"
            ws.cell(row=row, column=current_col).value = _xlookup_formula(
                lookup_cell,
                sheet=PROJECTIONS_SHEET,
                lookup_col=proj_lookup_col,
                return_col=return_col,
            )
        current_col += 1

    for header_name, snap_field in snap_columns:
        return_col = snap_header.get(snap_field)
        if return_col is None:
            continue
        ws.cell(row=1, column=current_col, value=header_name)
        for row in range(2, ws.max_row + 1):
            lookup_cell = f"{get_column_letter(snapshot_id_col)}{row}"
            ws.cell(row=row, column=current_col).value = _xlookup_formula(
                lookup_cell,
                sheet=MARKET_SHEET,
                lookup_col=snap_lookup_col,
                return_col=return_col,
            )
        current_col += 1


def build_daily_workbook(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    date: str,
    model: str | None = None,
    review_run_id: str | None = None,
    snapshot_run_id: str | None = None,
    output_path: str | Path | None = None,
) -> Path:
    """Build a daily workbook with projections, market snapshots, OCR rows, EV, and BETS."""
    as_of_date = _normalize_date(date)
    resolved_review_run_id, review_model = _resolve_review_run(
        db_path,
        sport=sport,
        season=season,
        as_of_date=as_of_date,
        review_run_id=review_run_id,
    )
    model_name = model or review_model
    if not model_name:
        raise ValueError("Model could not be resolved for daily workbook generation.")

    projections_df = _build_projections_frame(
        db_path,
        sport=sport,
        season=season,
        model=model_name,
        as_of_date=as_of_date,
    )
    market_df = _build_market_snapshots_frame(
        db_path,
        sport=sport,
        season=season,
        as_of_date=as_of_date,
        snapshot_run_id=snapshot_run_id,
    )
    ocr_df = _build_ocr_frame(db_path, review_run_id=resolved_review_run_id)
    ev_df, bets_df = _build_ev_bets_frames(db_path, review_run_id=resolved_review_run_id)

    output = Path(output_path) if output_path else processed_path_for(
        sport, season, f"betting/daily_workbook_{as_of_date}.xlsx"
    )
    output.parent.mkdir(parents=True, exist_ok=True)

    meta_rows = [
        {"key": "review_run_id", "value": resolved_review_run_id},
        {"key": "sport", "value": sport},
        {"key": "season", "value": season},
        {"key": "date", "value": as_of_date},
        {"key": "model", "value": model_name},
        {"key": "snapshot_run_id", "value": snapshot_run_id or ""},
        {"key": "created_at", "value": datetime.now(timezone.utc).isoformat()},
    ]
    meta_df = pd.DataFrame(meta_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        projections_df.to_excel(writer, sheet_name=PROJECTIONS_SHEET, index=False)
        market_df.to_excel(writer, sheet_name=MARKET_SHEET, index=False)
        ocr_df.to_excel(writer, sheet_name=OCR_SHEET, index=False)
        ev_df.to_excel(writer, sheet_name=EV_SHEET, index=False)
        bets_df.to_excel(writer, sheet_name=BETS_SHEET, index=False)
        meta_df.to_excel(writer, sheet_name=META_SHEET, index=False)

    wb = load_workbook(output)
    if EV_SHEET in wb.sheetnames:
        ws = wb[EV_SHEET]
        _apply_formula_sheet(ws)
        _add_cross_sheet_links(wb, sheet_name=EV_SHEET)
        ws.protection.sheet = True
        ws.protection.enable = True
    if BETS_SHEET in wb.sheetnames:
        ws = wb[BETS_SHEET]
        _apply_formula_sheet(ws)
        _add_cross_sheet_links(wb, sheet_name=BETS_SHEET)
        ws.protection.sheet = False
    if META_SHEET in wb.sheetnames:
        wb[META_SHEET].sheet_state = "hidden"

    wb.save(output)
    return output


__all__ = ["build_daily_workbook"]
