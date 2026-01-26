"""Shared Excel formula helpers for betting workbooks."""

from __future__ import annotations

from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _header_index(ws: Worksheet) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def build_header_letter_map(ws: Worksheet) -> dict[str, str]:
    return {cell.value: cell.column_letter for cell in ws[1] if cell.value}


def _cell_ref(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _cell_ref_letter(letter: str, row: int) -> str:
    return f"{letter}{row}"


def _column_has_values(ws: Worksheet, col: int) -> bool:
    def _wrap_with_cal_cell(cell_ref: str | None, expr: str) -> str:
        if cell_ref:
            return f"IF({cell_ref}<>\"\",{cell_ref},{expr})"
        return expr

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        if str(value).strip() == "":
            continue
        return True
    return False


def resolve_market_inputs(cols_map: dict[str, str]) -> dict[str, dict[str, str | None]]:
    def _resolve(preferred: str, fallback: str | None = None) -> str | None:
        if preferred in cols_map:
            return cols_map[preferred]
        if fallback and fallback in cols_map:
            return cols_map[fallback]
        return None

    return {
        "ML": {
            "home_prob": _resolve("ml_home_win_prob", "home_win_prob"),
            "away_prob": _resolve("ml_away_win_prob", "away_win_prob"),
        },
        "spread": {
            "line": _resolve("line"),
            "mean": _resolve("spread_margin_mean", "margin_mean"),
            "sd": _resolve("spread_margin_sd", "margin_sd"),
        },
        "total": {
            "line": _resolve("line"),
            "mean": _resolve("total_mean", "total"),
            "sd": _resolve("total_sd"),
        },
    }


def _select_ml_prob_letters(
    ws: Worksheet,
    header_index: dict[str, int],
    header_letters: dict[str, str],
    inputs: dict[str, dict[str, str | None]],
) -> tuple[str | None, str | None]:
    if (
        "home_win_prob_calibrated" in header_index
        and "away_win_prob_calibrated" in header_index
    ):
        calibrated_has_values = _column_has_values(
            ws, header_index["home_win_prob_calibrated"]
        ) or _column_has_values(ws, header_index["away_win_prob_calibrated"])
        if calibrated_has_values:
            return (
                header_letters["home_win_prob_calibrated"],
                header_letters["away_win_prob_calibrated"],
            )
    return inputs["ML"]["home_prob"], inputs["ML"]["away_prob"]


def apply_ev_formulas(ws: Worksheet, *, use_price: bool = False) -> None:
    header = _header_index(ws)
    required = {"odds", "implied_prob", "model_prob", "edge", "ev"}
    if not required.issubset(header):
        return

    odds_col = header["odds"]
    price_col = header.get("price") if use_price else None
    implied_col = header["implied_prob"]
    model_col = header["model_prob"]
    edge_col = header["edge"]
    ev_col = header["ev"]

    for row in range(2, ws.max_row + 1):
        odds_cell = _cell_ref(odds_col, row)
        price_cell = _cell_ref(price_col, row) if price_col else None
        implied_cell = _cell_ref(implied_col, row)
        model_cell = _cell_ref(model_col, row)
        edge_cell = _cell_ref(edge_col, row)
        ev_cell = _cell_ref(ev_col, row)

        if price_cell:
            odds_expr = f"IF(OR({price_cell}=\"\",{price_cell}=0),{odds_cell},{price_cell})"
        else:
            odds_expr = odds_cell

        ws[implied_cell].value = (
            f"=IF(OR({odds_expr}=\"\",{odds_expr}=0),\"\","
            f"IF({odds_expr}>0,100/({odds_expr}+100),-{odds_expr}/(-{odds_expr}+100)))"
        )
        ws[edge_cell].value = f"=IF(OR({model_cell}=\"\",{implied_cell}=\"\"),\"\",{model_cell}-{implied_cell})"
        ws[ev_cell].value = (
            f"=IF(OR({model_cell}=\"\",{odds_expr}=\"\",{odds_expr}=0),\"\","
            f"({model_cell}*IF({odds_expr}>0,{odds_expr}/100,100/ABS({odds_expr})))-(1-{model_cell}))"
        )


def unlock_input_columns(ws: Worksheet, columns: list[str]) -> None:
    header = _header_index(ws)
    target_cols = [header[col] for col in columns if col in header]
    if not target_cols:
        return
    for row in range(2, ws.max_row + 1):
        for col in target_cols:
            ws.cell(row=row, column=col).protection = Protection(locked=False)


def apply_model_prob_formulas_for_bets_sheet(ws: Worksheet) -> None:
    header = _header_index(ws)
    required = {
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
    }
    if not required.issubset(header):
        return

    header_letters = {
        name: get_column_letter(col) for name, col in header.items() if name
    }
    inputs = resolve_market_inputs(header_letters)
    home_win_col, away_win_col = _select_ml_prob_letters(
        ws, header, header_letters, inputs
    )
    if not home_win_col or not away_win_col:
        return

    market_col = header_letters["market_type"]
    selection_col = header_letters["selection"]
    line_col = header_letters["line"]
    model_col = header_letters["model_prob"]
    home_team_col = header_letters["home_team"]
    away_team_col = header_letters["away_team"]
    spread_mean_col = inputs["spread"]["mean"]
    spread_sd_col = inputs["spread"]["sd"]
    total_mean_col = inputs["total"]["mean"]
    total_sd_col = inputs["total"]["sd"]
    if not spread_mean_col or not spread_sd_col or not total_mean_col or not total_sd_col:
        return
    spread_calibrated_col = header_letters.get("spread_prob_calibrated")
    total_calibrated_col = header_letters.get("total_prob_calibrated")
    def _wrap_with_cal_cell(cell_ref: str | None, expr: str) -> str:
        if cell_ref:
            return f"IF({cell_ref}<>\"\",{cell_ref},{expr})"
        return expr

    for row in range(2, ws.max_row + 1):
        market_cell = _cell_ref_letter(market_col, row)
        selection_cell = _cell_ref_letter(selection_col, row)
        line_cell = _cell_ref_letter(line_col, row)
        model_cell = _cell_ref_letter(model_col, row)
        home_team_cell = _cell_ref_letter(home_team_col, row)
        away_team_cell = _cell_ref_letter(away_team_col, row)
        home_win_cell = _cell_ref_letter(home_win_col, row)
        away_win_cell = _cell_ref_letter(away_win_col, row)
        spread_mean_cell = _cell_ref_letter(spread_mean_col, row)
        spread_sd_cell = _cell_ref_letter(spread_sd_col, row)
        total_mean_cell = _cell_ref_letter(total_mean_col, row)
        total_sd_cell = _cell_ref_letter(total_sd_col, row)

        spread_cdf_selection_away = (
            f"(0.5*(1+ERF((({line_cell})-{spread_mean_cell})/({spread_sd_cell}*SQRT(2)))))"
        )
        spread_cdf_selection_home = (
            f"(0.5*(1+ERF(((-{line_cell})-{spread_mean_cell})/({spread_sd_cell}*SQRT(2)))))"
        )
        total_cdf = f"(0.5*(1+ERF(({line_cell}-{total_mean_cell})/({total_sd_cell}*SQRT(2)))))"

        spread_base = (
            f"IF(OR({line_cell}=\"\",{spread_sd_cell}=\"\"),\"\","
            f"IF({selection_cell}={home_team_cell},"
            f"1-{spread_cdf_selection_home},"
            f"IF({selection_cell}={away_team_cell},"
            f"{spread_cdf_selection_away},\"\")))"
        )
        total_base = (
            f"IF(OR({line_cell}=\"\",{total_sd_cell}=\"\"),\"\","
            f"IF({selection_cell}=\"Over\","
            f"1-{total_cdf},"
            f"IF({selection_cell}=\"Under\","
            f"{total_cdf},\"\")))"
        )
        spread_cal_cell = (
            _cell_ref_letter(spread_calibrated_col, row)
            if spread_calibrated_col
            else None
        )
        total_cal_cell = (
            _cell_ref_letter(total_calibrated_col, row)
            if total_calibrated_col
            else None
        )
        spread_formula = _wrap_with_cal_cell(spread_cal_cell, spread_base)
        total_formula = _wrap_with_cal_cell(total_cal_cell, total_base)

        ws[model_cell].value = (
            f"=@IF({market_cell}=\"ML\","
            f"IF({selection_cell}={home_team_cell},{home_win_cell},"
            f"IF({selection_cell}={away_team_cell},{away_win_cell},\"\")),"
            f"IF({market_cell}=\"spread\",{spread_formula},"
            f"IF({market_cell}=\"total\",{total_formula},\"\")))"
        )


def validate_no_ellipsis_formulas(wb: Workbook) -> None:
    """Fail if any worksheet contains a formula with ellipsis text."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if not cell.value.startswith(("=", "=@")):
                    continue
                if "..." in cell.value:
                    raise ValueError(
                        f"Formula validation failed: {ws.title}!{cell.coordinate} contains '...'"
                    )


def validate_bets_formulas(ws_bets: Worksheet) -> None:
    header_index = _header_index(ws_bets)
    header_letters = build_header_letter_map(ws_bets)
    required_headers = {
        "market_type",
        "selection",
        "line",
        "model_prob",
        "home_team",
        "away_team",
    }
    missing_headers = sorted(name for name in required_headers if name not in header_index)
    if missing_headers:
        raise ValueError(f"BETS formula validation failed: missing headers {missing_headers}")

    inputs = resolve_market_inputs(header_letters)
    ml_home_col, ml_away_col = _select_ml_prob_letters(
        ws_bets, header_index, header_letters, inputs
    )
    missing_inputs: list[str] = []
    if not ml_home_col:
        missing_inputs.append("ml_home_win_prob/home_win_prob")
    if not ml_away_col:
        missing_inputs.append("ml_away_win_prob/away_win_prob")
    if not inputs["spread"]["mean"]:
        missing_inputs.append("spread_margin_mean/margin_mean")
    if not inputs["spread"]["sd"]:
        missing_inputs.append("spread_margin_sd/margin_sd")
    if not inputs["total"]["mean"]:
        missing_inputs.append("total_mean/total")
    if not inputs["total"]["sd"]:
        missing_inputs.append("total_sd")
    if missing_inputs:
        raise ValueError(
            "BETS formula validation failed: missing required input columns: "
            + ", ".join(missing_inputs)
        )

    market_col = header_index["market_type"]
    model_col = header_index["model_prob"]
    sample_rows: dict[str, list[int]] = {"ML": [], "spread": [], "total": []}
    for row in range(2, ws_bets.max_row + 1):
        value = ws_bets.cell(row=row, column=market_col).value
        market = str(value).strip() if value is not None else ""
        if market in sample_rows and len(sample_rows[market]) < 3:
            sample_rows[market].append(row)

    def _assert_formula(row: int, required_refs: list[str], market: str) -> None:
        cell = ws_bets.cell(row=row, column=model_col)
        if not isinstance(cell.value, str):
            raise ValueError(
                f"BETS formula validation failed: model_prob not a string for {market} row {row}"
            )
        formula = cell.value.strip()
        if not (formula.startswith("=") or formula.startswith("=@")):
            raise ValueError(
                f"BETS formula validation failed: model_prob missing formula for {market} row {row}"
            )
        if "..." in formula:
            raise ValueError(
                f"BETS formula validation failed: model_prob contains '...' for {market} row {row}"
            )
        missing_refs = [ref for ref in required_refs if ref not in formula]
        if missing_refs:
            raise ValueError(
                f"BETS formula validation failed: model_prob missing refs {missing_refs} "
                f"for {market} row {row}"
            )

    for row in sample_rows["ML"]:
        _assert_formula(
            row,
            [
                _cell_ref_letter(ml_home_col, row),
                _cell_ref_letter(ml_away_col, row),
            ],
            "ML",
        )
    for row in sample_rows["spread"]:
        _assert_formula(
            row,
            [
                _cell_ref_letter(inputs["spread"]["line"], row),
                _cell_ref_letter(inputs["spread"]["mean"], row),
                _cell_ref_letter(inputs["spread"]["sd"], row),
            ],
            "spread",
        )
    for row in sample_rows["total"]:
        _assert_formula(
            row,
            [
                _cell_ref_letter(inputs["total"]["line"], row),
                _cell_ref_letter(inputs["total"]["mean"], row),
                _cell_ref_letter(inputs["total"]["sd"], row),
            ],
            "total",
        )
