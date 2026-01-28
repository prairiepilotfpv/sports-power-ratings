"""Schedule export pipeline with projection fields."""

from __future__ import annotations

import json
import logging
import math
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, TypeAlias
import re
import sqlite3

from config import (
    MARGIN_SD_GUARDRAIL_MIN,
    TOTAL_SD_GUARDRAIL_MIN,
    CLIP_RATE_WARN_THRESHOLD,
    CLIP_RATE_ERROR_THRESHOLD,
)

import pandas as pd
from pandas._libs.missing import NAType
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

from contracts import SCHEDULE_EXPORT_COLUMNS, validate_schedule_export_frame
from data.paths import processed_path_for
from data.repository import (
    get_active_ensemble_market_weights,
    get_active_ensemble_market_weights_and_models,
    get_active_ensemble_market_weights_source,
    get_active_model_market_params,
    get_active_model_market_params_source,
    load_best_ensemble_market_tuning_weights_by_optimized_metric,
    get_active_ensemble_market_selection,
    get_active_ensemble_market_tuning_run,
    load_selection_run,
    load_tuning_run,
    load_games,
    load_model_metrics,
)
from data import teams as team_repo
from forecasting import build_forecasts_df
from forecasting.forecast_service import (
    _completed_games,
    _project_row as _forecast_project_row,
    _rating_lookup as _forecast_rating_lookup,
    _safe_date,
)
from pipelines.model_params import (
    resolve_active_model_market_params,
    resolve_active_ensemble_weights,
    resolve_effective_params,
    resolve_model_market_params_with_metadata,
)
resolve_model_params_with_metadata = resolve_model_market_params_with_metadata
from pipelines.common import normalize_games, resolve_output_path
from ensemble.config import load_ensemble_config
from ensemble.io import load_market_weights
from ensemble.ml_v1 import MLWeightedAverageEnsemble

# ============================================================================
# KNOWN MISSING FEATURES (TODO)
# ============================================================================
# 1. Total Recency Adjustment Feature
#    - Functions load_latest_total_recency_adjustment() and
#      _calculate_total_recency_adjustment() are referenced in tests but
#      not implemented in this module.
#    - Purpose: Adjust total predictions based on temporal patterns (e.g.,
#      recent trends in scoring)
#    - Tests that depend on this:
#      - tests/pipelines/test_schedule_refresh.py::test_total_recency_adjustment_only_from_artifact
#      - tests/pipelines/test_schedule_refresh.py::test_artifact_precedence_production
#    - Implementation needed in: src/pipelines/schedule.py or src/forecasting/
#
# 2. _build_market_forecasts_for_ensembles Signature Mismatch
#    - Tests pass 'forecast_params_by_model' parameter but function doesn't
#      accept it (test_schedule_ensemble_config_usage.py line 64)
#    - Need to determine: should function accept this param or should test
#      be updated?
#    - Likely purpose: pass pre-computed forecast params per model to avoid
#      redundant computation
#
# ============================================================================

from ensemble.spread_v1 import SpreadWeightedAverageEnsemble
from ensemble.total_v1 import TotalWeightedAverageEnsemble
from pipelines.excel_bets_format import apply_bets_sheet_formatting
from pipelines.excel_formulas import (
    apply_ev_formulas,
    apply_model_prob_formulas_for_bets_sheet,
    validate_bets_formulas,
    validate_no_ellipsis_formulas,
)
from pipelines.market_utils import _metric_name_for_market, _metric_optimized_label
from pipelines.matchups import team_home_advantages
from pipelines.projection_engines import get_projection_engine
from pipelines.projections import (
    average_total_points,
    cover_prob,
    fit_total_model,
    over_prob,
    team_scoring_averages,
)
from pipelines.metadata import prediction_hash
from calibration.io import load_latest_calibrator
from models.base import resolve_model_identity
from models.registry import get_model, list_models, normalize_model_name
from markets.base import Market


_project_row = _forecast_project_row
_rating_lookup = _forecast_rating_lookup

EXPORT_MARKETS: List[Market] = [Market.ML, Market.SPREAD, Market.TOTAL]
MARKET_ORDER: Dict[str, int] = {market.name: idx for idx, market in enumerate(EXPORT_MARKETS)}

CalibratedProbability: TypeAlias = float | NAType

_LOG = logging.getLogger(__name__)


class HealthCheckError(Exception):
    """Raised when health gate fails and fail_on_health_check is enabled."""

    def __init__(self, status: str, details: dict[str, Any]):
        self.status = status
        self.details = details
        super().__init__(
            f"Health check failed: {status}. "
            f"spread_clip_rate={details.get('spread_clip_rate')}, "
            f"total_clip_rate={details.get('total_clip_rate')}"
        )


def _compute_health_status(
    spread_clip_rate: float | None,
    total_clip_rate: float | None,
) -> tuple[str, dict[str, Any]]:
    """Compute health status from uncertainty clip rates.

    Evaluates whether the proportion of predictions hitting SD guardrail floors
    exceeds warning or error thresholds.

    Args:
        spread_clip_rate: Fraction of SPREAD rows where margin_sd was clipped to floor
        total_clip_rate: Fraction of TOTAL rows where total_sd was clipped to floor

    Returns:
        Tuple of (status_label, details_dict) where status_label is one of:
        - "OK": Clip rates are within acceptable bounds
        - "WARN_UNCERTAINTY_CLIP": Clip rates exceed warning threshold
        - "FAIL_UNCERTAINTY_CLIP": Clip rates exceed error threshold
    """
    details = {
        "spread_clip_rate": spread_clip_rate,
        "total_clip_rate": total_clip_rate,
        "warn_threshold": CLIP_RATE_WARN_THRESHOLD,
        "error_threshold": CLIP_RATE_ERROR_THRESHOLD,
    }

    max_rate = max(
        spread_clip_rate or 0.0,
        total_clip_rate or 0.0,
    )

    if max_rate >= CLIP_RATE_ERROR_THRESHOLD:
        return ("FAIL_UNCERTAINTY_CLIP", details)
    elif max_rate >= CLIP_RATE_WARN_THRESHOLD:
        return ("WARN_UNCERTAINTY_CLIP", details)
    else:
        return ("OK", details)


DASHBOARD_COLUMNS: List[str] = [
    "model",
    "model_version",
    "params_source",
    "params_source_label",
    "params_source_run_id",
    "tuned_metric_used",
    "params_metric_optimized",
    "params_best_score",
    "params_fingerprint",
    "params_nonempty",
    "tuning_run_id",
    "run_timestamp_utc",
    "prediction_hash",
    "date",
    "params_market",
    "game",
    "projected_home_score",
    "projected_away_score",
    "total",
    "projected_winner",
    "projected_spread",
    "margin_mean",
    "margin_sd",
    "model_p_home_win",
    "normal_p_home_win",
    "home_win_prob",
    "away_win_prob",
    "winner_win_prob",
    "logistic_home_win_prob",
    "win_prob_source",
    "margin_dist_assumption",
    "total_sd",
]

MODEL_METADATA_DATA_START_ROW = 60  # Allow multi-market metadata; bump if new metadata keys appear


def _finalize_schedule_export(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Ensure the export includes all expected columns before validation."""
    finalized = schedule_df.copy()
    
    # Empty DataFrames are valid (e.g., no upcoming games); return as-is
    if finalized.empty:
        return finalized
    
    required_columns = ["date", "game_id"]
    missing_required = [col for col in required_columns if col not in finalized.columns]
    if missing_required:
        raise ValueError(
            f"Missing required schedule export columns: {missing_required}"
        )
    for column in SCHEDULE_EXPORT_COLUMNS:
        if column not in finalized.columns:
            finalized[column] = pd.NA
    return finalized.loc[:, SCHEDULE_EXPORT_COLUMNS]


def _order_schedule_export(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Ensure a consistent column order for downstream reporting."""
    finalized = _finalize_schedule_export(schedule_df)
    return validate_schedule_export_frame(
        finalized,
        expected_columns=SCHEDULE_EXPORT_COLUMNS,
        context="Schedule export",
    )


def _resolve_models(model: str | Iterable[str] | None) -> list[str]:
    if model is None:
        return list_models()
    if isinstance(model, (list, tuple, set)):
        models = [normalize_model_name(m) for m in model]
        return list(dict.fromkeys(models))
    normalized = normalize_model_name(model)
    if normalized in {"all", "*"}:
        return list_models()
    return [normalized]


def _resolve_workbook_path(
    output_path: str | Path | None,
    *,
    sport: str,
    season: str,
    default_name: str,
) -> Path:
    default_path = processed_path_for(sport, season, default_name)
    resolved = resolve_output_path(
        output_path,
        default_path=default_path,
    )
    resolved.parent.mkdir(parents=True, exist_ok=True)
    if resolved.exists():
        next_path = _next_available_path(resolved)
        _LOG.warning("Output exists: %s. Writing to %s.", resolved, next_path)
        resolved = next_path
    return resolved


def _next_available_path(path: Path) -> Path:
    """Return the first unused filename by appending a numeric suffix."""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for index in range(1, 1000):
        candidate = parent / f"{stem}-{index}{suffix}"
        if not candidate.exists():
            return candidate
    raise FileExistsError(
        f"Output exists: {path}. Tried 999 alternate names."
    )


def _build_schedule_dataframe(
    df: pd.DataFrame,
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    model: str,
    upcoming_only: bool,
    model_params: dict[str, float] | None,
    params_source: str,
    params_source_label: str | None = None,
    params_source_run_id: str | None = None,
    tuned_metric_used: str | None = None,
    params_metric_optimized: str | None = None,
    params_best_score: float | None = None,
    params_fingerprint: str | None = None,
    params_nonempty: bool | None = None,
    params_run_id: str | None = None,
    params_market: str | None = None,
    fit_end_date: str | date | None = None,
    fail_on_health_check: bool = False,
    mode: str = "dev",
) -> pd.DataFrame:
    schedule_df = build_forecasts_df(
        db_path=db_path,
        sport=sport,
        season=season,
        model=model,
        games_df=df,
        include_upcoming=True,
        include_played=not upcoming_only,
        model_params=model_params,
        params_source=params_source,
        params_source_label=params_source_label,
        params_source_run_id=params_source_run_id,
        tuned_metric_used=tuned_metric_used,
        params_metric_optimized=params_metric_optimized,
        params_best_score=params_best_score,
        params_fingerprint=params_fingerprint,
        params_nonempty=params_nonempty,
        params_run_id=params_run_id,
        params_market=params_market,
        fit_end_date=fit_end_date,
        mode=mode,
    )
    schedule_df = _apply_calibration_to_schedule_df(
        schedule_df,
        sport=sport,
        season=season,
        model=model,
        fail_on_health_check=fail_on_health_check,
    )
    for col in (
        "home_win_prob_raw",
        "away_win_prob_raw",
        "home_win_prob_calibrated",
        "away_win_prob_calibrated",
    ):
        if col not in schedule_df.columns:
            schedule_df[col] = pd.NA
    return _order_schedule_export(schedule_df)


def _compute_total_recency_adjustment(
    db_path: str | Path,
    sport: str,
    as_of_date: str | date | None = None,
    lookback_games: int = 100,
) -> float | None:
    """Compute adjustment factor for total_mean based on recent game performance.
    
    Compares the average total from recent completed games to the season-long average.
    This accounts for mid-season trends (e.g., January slowdown in NBA).
    
    Args:
        db_path: Path to the sport/season database.
        sport: Sport identifier (e.g., 'nba').
        as_of_date: Optional cutoff date; if None, uses all completed games.
        lookback_games: Number of recent games to average for recency adjustment.
    
    Returns:
        Adjustment factor (e.g., -6.0 means subtract 6 points from predicted totals).
        Returns None if insufficient data.
    """
    try:
        conn = sqlite3.connect(db_path)
        query = """
            SELECT date, home_score, away_score,
                   home_score + away_score as total_score
            FROM games
            WHERE home_score IS NOT NULL AND away_score IS NOT NULL
        """
        params: list[object] = []
        as_of = None
        if as_of_date:
            as_of = pd.to_datetime(as_of_date)
            query += " AND date <= ?"
            params.append(as_of.strftime("%Y-%m-%d"))
        query += " ORDER BY date DESC LIMIT ?"
        params.append(lookback_games)

        recent_games = pd.read_sql(query, conn, params=params)
        conn.close()
        
        if recent_games.empty or len(recent_games) < 20:
            return None
        
        # Also get season-to-date average for context
        conn = sqlite3.connect(db_path)
        all_query = "SELECT home_score + away_score as total FROM games WHERE home_score IS NOT NULL"
        all_params: list[object] = []
        if as_of is not None:
            all_query += " AND date <= ?"
            all_params.append(as_of.strftime("%Y-%m-%d"))
        all_games = pd.read_sql(all_query, conn, params=all_params or None)
        conn.close()
        
        if all_games.empty:
            return None
        
        recent_avg = recent_games["total_score"].mean()
        season_avg = all_games["total"].mean()
        
        # Return adjustment as (recent - season), so positive means recent games are higher-scoring
        adjustment = recent_avg - season_avg
        
        _LOG.info(
            "Total recency adjustment: recent_avg=%.1f, season_avg=%.1f, adjustment=%+.1f",
            recent_avg,
            season_avg,
            adjustment,
        )
        
        return adjustment
    except Exception as e:
        _LOG.warning("Failed to compute total recency adjustment: %s", e)
        return None


def _apply_calibration_to_schedule_df(
    schedule_df: pd.DataFrame,
    *,
    sport: str,
    season: str,
    model: str,
    fail_on_health_check: bool = False,
) -> pd.DataFrame:
    """Apply calibrators for all markets (ML, SPREAD, TOTAL) to schedule predictions.

    Loads the latest calibrator for each market and applies transformations:
    - ML: Calibrates home_win_prob and away_win_prob; appends "+calibrated_ml" to win_prob_source
    - SPREAD: Calibrates margin_mean and margin_sd; appends "+calibrated_spread" to win_prob_source
    - TOTAL: Calibrates total_mean and total_sd; appends "+calibrated_total" to win_prob_source

    Provenance tags are appended to win_prob_source to track which calibrators were applied.
    Tags are idempotent: appending the same tag twice will not result in duplicates.

    Args:
        fail_on_health_check: If True, raises HealthCheckError when health gate fails.
    
    Returns modified DataFrame with calibrated predictions and updated win_prob_source.
    If calibrators are missing or calibration fails, original predictions are preserved.
    
    Args:
        schedule_df: Input DataFrame with predictions from models
        sport: Sport code (e.g., "nba", "nfl")
        season: Season identifier (e.g., "2025-26")
        model: Model name to look up calibrators for
        
    Returns:
        DataFrame with calibrated predictions and provenance tags appended to win_prob_source
    """
    if schedule_df.empty:
        return schedule_df

    _LOG.info(f"[Calibration] Starting calibration for {sport}/{season}/{model}")
    df = schedule_df.copy()

    def _win_prob_tag_present(tag: str) -> bool:
        if "win_prob_source" not in df.columns:
            return False
        tag_lower = tag.lower()
        return (
            df["win_prob_source"]
            .fillna("")
            .astype(str)
            .str.lower()
            .str.contains(tag_lower)
            .any()
        )

    # Track pre-guardrail SD values for debugging/logging
    if "margin_sd" in df.columns:
        df["margin_sd_pre_guardrail"] = pd.NA
    if "total_sd" in df.columns:
        df["total_sd_pre_guardrail"] = pd.NA
    
    # Track which markets had calibration successfully applied.
    # This set is used later to determine which provenance tags should be appended.
    # Only markets in this set will have tags added to win_prob_source.
    calibrated_markets = set()
    
    # ========== ML MARKET CALIBRATION ==========
    # Calibrate moneyline (win probability) predictions using Platt scaling or similar method
    # Only processes if home_win_prob column exists (indicates ML predictions are present)
    if "home_win_prob" in df.columns and not _win_prob_tag_present("calibrated_ml"):
        ml_calibrator = load_latest_calibrator(
            sport=sport,
            season=season,
            model=model,
            market="ML",
        )
        
        if ml_calibrator is not None:
            _LOG.info(f"[ML calibration] Loaded calibrator for {sport}/{season}/{model}")
            raw_probs = pd.to_numeric(df["home_win_prob"], errors="coerce")
            valid_mask = raw_probs.notna()
            
            if valid_mask.any():
                # Store raw values
                df["home_win_prob_raw"] = df["home_win_prob"]
                df["away_win_prob_raw"] = df["away_win_prob"]
                
                try:
                    # Use float('nan') (not pd.NA) for float-dtype Series to avoid NAType conversion errors
                    # pd.NA is for nullable dtypes (Float64), NaN is for numpy float64
                    calibrated = pd.Series(float("nan"), index=df.index, dtype="float")
                    calibrated_values = ml_calibrator.transform(raw_probs.loc[valid_mask].astype(float))
                    calibrated.loc[valid_mask] = pd.to_numeric(calibrated_values, errors="coerce")
                    df["home_win_prob_calibrated"] = calibrated
                    df["away_win_prob_calibrated"] = 1.0 - calibrated
                    
                    # Use calibrated values in place of raw
                    calibrated_mask = df["home_win_prob_calibrated"].notna()
                    df.loc[calibrated_mask, "home_win_prob"] = df.loc[
                        calibrated_mask, "home_win_prob_calibrated"
                    ]
                    df.loc[calibrated_mask, "away_win_prob"] = df.loc[
                        calibrated_mask, "away_win_prob_calibrated"
                    ]
                    
                    calibrated_markets.add("ML")
                    
                    _LOG.info(
                        f"[_apply_calibration_to_schedule_df] Applied ML calibrator to "
                        f"{calibrated_mask.sum()} rows"
                    )
                except Exception as e:
                    _LOG.warning(f"[_apply_calibration_to_schedule_df] ML calibration failed: {e}")
    
    # ========== SPREAD MARKET CALIBRATION ==========
    # Calibrates predicted spread mean and standard deviation using a distribution-aware calibrator.
    # Unlike ML calibration (which recalibrates point probabilities via Platt scaling),
    # SPREAD calibration operates on the predicted distribution parameters:
    # - Shifts the mean prediction toward observed center
    # - Adjusts variance/std.dev. to improve predictive coverage
    # These calibrations are applied BEFORE BETS are generated from probabilities.
    # Instrumentation below logs changes to verify calibration impact on predictions.
    if (
        "margin_mean" in df.columns
        and "margin_sd" in df.columns
        and not _win_prob_tag_present("calibrated_spread")
    ):
        spread_calibrator = load_latest_calibrator(
            sport=sport,
            season=season,
            model=model,
            market="spread",
        )
        
        if spread_calibrator is not None:
            _LOG.info(f"[SPREAD calibration] Loaded calibrator for {sport}/{season}/{model}") # Check if it's a distribution calibrator
            if hasattr(spread_calibrator, "metadata") and "variance" in str(spread_calibrator.metadata.get("method", "")):
                margin_mean = pd.to_numeric(df["margin_mean"], errors="coerce")
                margin_sd = pd.to_numeric(df["margin_sd"], errors="coerce")
                valid_mask = (margin_mean.notna()) & (margin_sd.notna()) & (margin_sd > 0)
                
                # ===== INSTRUMENTATION: CAPTURE PRE-CALIBRATION STATE =====
                # Store snapshots before calibration to compute deltas afterward.
                # This instrumentation exists to verify calibration impact, not to tune models.
                margin_mean_before = margin_mean.copy()
                margin_sd_before = margin_sd.copy()
                rows_with_mean_before = margin_mean.notna().sum()
                
                if valid_mask.any():
                    try:
                        # Build input DataFrame for calibrator
                        calib_input = pd.DataFrame({
                            "pred_mean": margin_mean.loc[valid_mask],
                            "pred_sd": margin_sd.loc[valid_mask],
                        })
                        
                        # Apply calibrator
                        calib_result = spread_calibrator.transform(calib_input)
                        pre_guardrail_sd = calib_result["calibrated_sd"]
                        df.loc[valid_mask, "margin_sd_pre_guardrail"] = pre_guardrail_sd.values
                        clipped_sd = pre_guardrail_sd.clip(lower=MARGIN_SD_GUARDRAIL_MIN)
                        calib_result["calibrated_sd"] = clipped_sd

                        # Update values
                        df.loc[valid_mask, "margin_mean"] = calib_result["calibrated_mean"].values
                        df.loc[valid_mask, "margin_sd"] = calib_result["calibrated_sd"].values

                        # ===== INSTRUMENTATION: COMPUTE AND LOG DELTAS =====
                        margin_mean_after = df["margin_mean"].copy()
                        margin_sd_after = df["margin_sd"].copy()
                        rows_with_mean_after = margin_mean_after.notna().sum()
                        
                        # Compute max absolute change in mean and sd across all rows
                        mean_delta = (margin_mean_after - margin_mean_before).abs().max()
                        sd_delta = (margin_sd_after - margin_sd_before).abs().max()
                        
                        # Log at DEBUG level to verify calibration behavior
                        _LOG.debug(
                            "[CALIBRATION][SPREAD] rows_before=%d, rows_after=%d, "
                            "max_mean_delta=%.6f, max_sd_delta=%.6f",
                            int(rows_with_mean_before),
                            int(rows_with_mean_after),
                            float(mean_delta) if pd.notna(mean_delta) else 0.0,
                            float(sd_delta) if pd.notna(sd_delta) else 0.0,
                        )

                        calibrated_markets.add("SPREAD")

                        _LOG.info(
                            f"[_apply_calibration_to_schedule_df] Applied SPREAD distribution calibrator "
                            f"to {valid_mask.sum()} rows"
                        )

                        guardrail_mask = pre_guardrail_sd.notna()
                        guardrail_count = int(guardrail_mask.sum())
                        if guardrail_count:
                            clipped_mask = guardrail_mask & (
                                pre_guardrail_sd < MARGIN_SD_GUARDRAIL_MIN
                            )
                            clipped_count = int(clipped_mask.sum())
                            pct_clipped = float(clipped_count / guardrail_count * 100)
                            raw_p50 = float(pre_guardrail_sd[guardrail_mask].median())
                            used_p50 = float(clipped_sd[guardrail_mask].median())
                            _LOG.info(
                                "[_apply_calibration_to_schedule_df] SPREAD guardrail summary: "
                                "threshold=%.2f, clipped=%.1f%% (%d/%d), "
                                "p50 raw=%.3f, p50 used=%.3f",
                                MARGIN_SD_GUARDRAIL_MIN,
                                pct_clipped,
                                clipped_count,
                                guardrail_count,
                                raw_p50,
                                used_p50,
                            )
                    except Exception as e:
                        _LOG.warning(f"[_apply_calibration_to_schedule_df] SPREAD calibration failed: {e}")
        else:
            _LOG.debug(f"[SPREAD calibration] No calibrator found for {sport}/{season}/{model} market=spread")
    
    # ========== TOTAL MARKET CALIBRATION ==========
    # Calibrates predicted total mean and standard deviation using a distribution-aware calibrator.
    # Unlike ML calibration (which recalibrates point probabilities via Platt scaling),
    # TOTAL calibration operates on the predicted distribution parameters:
    # - Shifts the mean prediction toward observed center
    # - Adjusts variance/std.dev. to improve predictive coverage
    # These calibrations are applied BEFORE BETS are generated from probabilities.
    # Instrumentation below logs changes to verify calibration impact on predictions.
    if (
        "total_mean" in df.columns
        and "total_sd" in df.columns
        and not _win_prob_tag_present("calibrated_total")
    ):
        total_calibrator = load_latest_calibrator(
            sport=sport,
            season=season,
            model=model,
            market="total",
        )
        
        if total_calibrator is not None:
            _LOG.info(f"[TOTAL calibration] Loaded calibrator for {sport}/{season}/{model}") # Check if it's a distribution calibrator
            if hasattr(total_calibrator, "metadata") and "variance" in str(total_calibrator.metadata.get("method", "")):
                total_mean = pd.to_numeric(df["total_mean"], errors="coerce")
                total_sd = pd.to_numeric(df["total_sd"], errors="coerce")
                valid_mask = (total_mean.notna()) & (total_sd.notna()) & (total_sd > 0)
                
                # ===== INSTRUMENTATION: CAPTURE PRE-CALIBRATION STATE =====
                # Store snapshots before calibration to compute deltas afterward.
                # This instrumentation exists to verify calibration impact, not to tune models.
                total_mean_before = total_mean.copy()
                total_sd_before = total_sd.copy()
                rows_with_mean_before = total_mean.notna().sum()
                
                if valid_mask.any():
                    try:
                        # Build input DataFrame for calibrator
                        calib_input = pd.DataFrame({
                            "pred_mean": total_mean.loc[valid_mask],
                            "pred_sd": total_sd.loc[valid_mask],
                        })
                        
                        # Apply calibrator
                        calib_result = total_calibrator.transform(calib_input)
                        pre_guardrail_sd = calib_result["calibrated_sd"]
                        df.loc[valid_mask, "total_sd_pre_guardrail"] = pre_guardrail_sd.values
                        clipped_sd = pre_guardrail_sd.clip(lower=TOTAL_SD_GUARDRAIL_MIN)
                        calib_result["calibrated_sd"] = clipped_sd

                        # Update values
                        df.loc[valid_mask, "total_mean"] = calib_result["calibrated_mean"].values
                        df.loc[valid_mask, "total_sd"] = calib_result["calibrated_sd"].values

                        # ===== INSTRUMENTATION: COMPUTE AND LOG DELTAS =====
                        total_mean_after = df["total_mean"].copy()
                        total_sd_after = df["total_sd"].copy()
                        rows_with_mean_after = total_mean_after.notna().sum()
                        
                        # Compute max absolute change in mean and sd across all rows
                        mean_delta = (total_mean_after - total_mean_before).abs().max()
                        sd_delta = (total_sd_after - total_sd_before).abs().max()
                        
                        # Log at DEBUG level to verify calibration behavior
                        _LOG.debug(
                            "[CALIBRATION][TOTAL] rows_before=%d, rows_after=%d, "
                            "max_mean_delta=%.6f, max_sd_delta=%.6f",
                            int(rows_with_mean_before),
                            int(rows_with_mean_after),
                            float(mean_delta) if pd.notna(mean_delta) else 0.0,
                            float(sd_delta) if pd.notna(sd_delta) else 0.0,
                        )

                        calibrated_markets.add("TOTAL")

                        _LOG.info(
                            f"[_apply_calibration_to_schedule_df] Applied TOTAL distribution calibrator "
                            f"to {valid_mask.sum()} rows"
                        )

                        guardrail_mask = pre_guardrail_sd.notna()
                        guardrail_count = int(guardrail_mask.sum())
                        if guardrail_count:
                            clipped_mask = guardrail_mask & (
                                pre_guardrail_sd < TOTAL_SD_GUARDRAIL_MIN
                            )
                            clipped_count = int(clipped_mask.sum())
                            pct_clipped = float(clipped_count / guardrail_count * 100)
                            raw_p50 = float(pre_guardrail_sd[guardrail_mask].median())
                            used_p50 = float(clipped_sd[guardrail_mask].median())
                            _LOG.info(
                                "[_apply_calibration_to_schedule_df] TOTAL guardrail summary: "
                                "threshold=%.2f, clipped=%.1f%% (%d/%d), "
                                "p50 raw=%.3f, p50 used=%.3f",
                                TOTAL_SD_GUARDRAIL_MIN,
                                pct_clipped,
                                clipped_count,
                                guardrail_count,
                                raw_p50,
                                used_p50,
                            )
                    except Exception as e:
                        _LOG.warning(f"[_apply_calibration_to_schedule_df] TOTAL calibration failed: {e}")
        else:
            _LOG.debug(f"[TOTAL calibration] No calibrator found for {sport}/{season}/{model} market=total")
    
    # ========== UPDATE WINNER WIN PROB WITH CALIBRATED VALUES ==========
    # After calibrating individual market probabilities, sync the aggregated winner_win_prob
    # to use calibrated values. This ensures consistency across all probability fields.
    if "projected_winner" in df.columns:
        home_wins = df["projected_winner"] == df["home_team"]
        away_wins = df["projected_winner"] == df["away_team"]
        df.loc[home_wins, "winner_win_prob"] = df.loc[home_wins, "home_win_prob"]
        df.loc[away_wins, "winner_win_prob"] = df.loc[away_wins, "away_win_prob"]

    # ========== APPEND MARKET-SPECIFIC CALIBRATION TAGS TO win_prob_source ==========
    # Append provenance tags to track which markets were calibrated.
    # This is critical for auditability and reproducibility:
    # - Allows downstream consumers to see exactly which calibrators were applied
    # - Enables filtering/grouping by calibration status
    # - Maintains audit trail for model performance analysis
    if "win_prob_source" in df.columns and calibrated_markets:
        # Map market names to their canonical provenance tags
        # These tags are standardized across all sports and models for consistency
        market_to_tag = {
            "ML": "calibrated_ml",
            "SPREAD": "calibrated_spread",
            "TOTAL": "calibrated_total",
        }
        
        def _append_calibration_tags(value: Any, tags: set[str]) -> str:
            """Append market-specific calibration tags idempotently to win_prob_source.
            
            For each tag in the input set:
            - Check if tag is already present (case-insensitive)
            - If not present, append it with '+' separator
            - Sort tags for deterministic output across runs
            
            Args:
                value: Current win_prob_source value (may be None or empty string)
                tags: Set of tag strings to append (e.g., {"calibrated_ml", "calibrated_spread"})
                
            Returns:
                Updated win_prob_source string with tags appended
                
            Examples:
                _append_calibration_tags("model_x", {"calibrated_ml"}) -> "model_x+calibrated_ml"
                _append_calibration_tags("model_x+calibrated_ml", {"calibrated_ml"}) -> "model_x+calibrated_ml"
                _append_calibration_tags(None, {"calibrated_spread"}) -> "calibrated_spread"
            """
            if value is None:
                text = ""
            else:
                text = str(value).strip()
            
            # Iterate through sorted tags for deterministic output regardless of set ordering
            for tag in sorted(tags):
                # Case-insensitive check: don't append if tag already exists
                if text and tag not in text.lower():
                    text = f"{text}+{tag}"
                elif not text:
                    text = tag
            
            return text
        
        # Build the set of tags from successfully applied calibrators.
        # Only markets that actually ran calibration (added to calibrated_markets set)
        # will have their tags included here.
        # Example: if calibrated_markets = {"ML", "SPREAD"}, then tags_to_append will be
        #          {"calibrated_ml", "calibrated_spread"}
        tags_to_append = {market_to_tag[m] for m in calibrated_markets}
        tags_list = ", ".join(sorted(tags_to_append))
        
        # Apply the tag-appending function to every row's win_prob_source value.
        # The _append_calibration_tags function handles idempotency: if a tag is already
        # present (case-insensitive), it won't be appended again. This ensures the function
        # is safe to call multiple times on the same DataFrame without creating duplicates.
        df["win_prob_source"] = df["win_prob_source"].apply(
            lambda val: _append_calibration_tags(val, tags_to_append)
        )
        
        # Log the tag appending operation for audit trail and debugging.
        # This enables:
        # 1. Monitoring which markets are being calibrated in production
        # 2. Debugging issues with missing calibrations
        # 3. Tracking calibration coverage over time
        _LOG.info(
            f"[_apply_calibration_to_schedule_df] Appended calibration provenance tags to "
            f"win_prob_source: {tags_list}"
        )

    # ========== COMPUTE HEALTH STATUS FROM CLIP RATES ==========
    # Compute the proportion of predictions that hit SD guardrail floors.
    # High clip rates indicate potential uncertainty collapse issues.
    spread_clip_rate = None
    total_clip_rate = None
    spread_clip_count = 0
    spread_total_count = 0
    total_clip_count = 0
    total_total_count = 0

    if "margin_sd_pre_guardrail" in df.columns:
        pre = df["margin_sd_pre_guardrail"].dropna()
        spread_total_count = len(pre)
        if spread_total_count > 0:
            spread_clip_count = int((pre < MARGIN_SD_GUARDRAIL_MIN).sum())
            spread_clip_rate = spread_clip_count / spread_total_count

    if "total_sd_pre_guardrail" in df.columns:
        pre = df["total_sd_pre_guardrail"].dropna()
        total_total_count = len(pre)
        if total_total_count > 0:
            total_clip_count = int((pre < TOTAL_SD_GUARDRAIL_MIN).sum())
            total_clip_rate = total_clip_count / total_total_count

    # Compute health status and log warnings/errors based on clip rates
    health_status, health_details = _compute_health_status(spread_clip_rate, total_clip_rate)

    if health_status == "FAIL_UNCERTAINTY_CLIP":
        _LOG.error(
            "[_apply_calibration_to_schedule_df] Health gate FAIL: uncertainty collapse detected. "
            "spread_clip_rate=%.1f%% (%d/%d), total_clip_rate=%.1f%% (%d/%d)",
            (spread_clip_rate or 0) * 100,
            spread_clip_count,
            spread_total_count,
            (total_clip_rate or 0) * 100,
            total_clip_count,
            total_total_count,
        )
        if fail_on_health_check:
            raise HealthCheckError(health_status, health_details)
    elif health_status == "WARN_UNCERTAINTY_CLIP":
        _LOG.warning(
            "[_apply_calibration_to_schedule_df] Health gate WARN: elevated uncertainty clipping. "
            "spread_clip_rate=%.1f%% (%d/%d), total_clip_rate=%.1f%% (%d/%d)",
            (spread_clip_rate or 0) * 100,
            spread_clip_count,
            spread_total_count,
            (total_clip_rate or 0) * 100,
            total_clip_count,
            total_total_count,
        )

    # Store health metadata on DataFrame for later retrieval by metadata builder
    df.attrs["_health_status"] = health_status
    df.attrs["_health_details"] = health_details

    return df


def _build_market_forecasts_for_ensembles(
    schedule_df: pd.DataFrame,
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    models: list[str],
    as_of_date: date,
    allowed_models_by_market: dict[str, list[str] | None] | None = None,
    market_metrics: dict[str, str] | None = None,
    fit_end_date: date | None = None,
    mode: str = "dev",
) -> dict[str, list[dict[str, Any]]]:
    rows_by_market: dict[str, list[dict[str, Any]]] = {
        Market.ML.name: [],
        Market.SPREAD.name: [],
        Market.TOTAL.name: [],
    }
    if schedule_df.empty:
        return rows_by_market

    for market in (Market.ML, Market.SPREAD, Market.TOTAL):
        market_allowed = None
        if allowed_models_by_market and market.name in allowed_models_by_market:
            market_allowed = allowed_models_by_market.get(market.name)
        model_iter = market_allowed if market_allowed else models
        metric_for_market = None
        if market_metrics and market.name in market_metrics:
            metric_for_market = market_metrics.get(market.name)
        for model_name in model_iter:
            # Resolve active market params (active -> best run -> defaults)
            resolved = resolve_effective_params(
                db_path=db_path,
                sport=sport,
                season=season,
                model=model_name,
                market=market.name,
            )
            params = resolved.params
            params_source = resolved.params_source_label
            tuned_metric_used = (
                resolved.metric_optimized.replace("backtest_", "", 1)
                if resolved.metric_optimized and str(resolved.metric_optimized).startswith("backtest_")
                else resolved.metric_optimized
            )

            model_df = schedule_df.copy(deep=True)
            market_schedule = _build_schedule_dataframe(
                model_df,
                db_path=db_path,
                sport=sport,
                season=season,
                model=model_name,
                upcoming_only=True,
                model_params=params,
                params_source=params_source,
                params_source_label=resolved.params_source_label,
                params_source_run_id=resolved.source_run_id,
                tuned_metric_used=tuned_metric_used,
                params_metric_optimized=resolved.metric_optimized,
                params_best_score=resolved.best_score,
                params_fingerprint=resolved.params_fingerprint,
                params_nonempty=resolved.params_nonempty,
                params_run_id=resolved.source_run_id,
                params_market=market.name,
                fit_end_date=fit_end_date,
                mode="dev",
            )
            if market_schedule.empty:
                continue
            if "date" in market_schedule.columns and "status" in market_schedule.columns:
                _dt = pd.to_datetime(market_schedule["date"], errors="coerce").dt.date
                mask = market_schedule["status"] == "scheduled"
                if as_of_date is not None:
                    mask &= _dt >= as_of_date
                subset = market_schedule.loc[mask]
            else:
                subset = market_schedule
            if subset.empty:
                continue
            for _, r in subset.iterrows():
                if market == Market.ML:
                    p_raw = None
                    if (
                        "home_win_prob_raw" in subset.columns
                        and pd.notna(r.get("home_win_prob_raw"))
                    ):
                        p_raw = r.get("home_win_prob_raw")
                    elif pd.notna(r.get("model_p_home_win")):
                        p_raw = r.get("model_p_home_win")
                    else:
                        p_raw = r.get("home_win_prob")
                    rows_by_market[market.name].append(
                        {
                            "game_id": r.get("game_id"),
                            "model_name": model_name,
                            "p_home_win": p_raw,
                            "margin_mean": r.get("margin_mean"),
                            "margin_sd": r.get("margin_sd"),
                            "total_mean": (
                                r.get("total")
                                if not _is_missing(r.get("total"))
                                else r.get("total_mean")
                            ),
                            "total_sd": r.get("total_sd"),
                        }
                    )
                else:
                    rows_by_market[market.name].append(
                        {
                            "game_id": r.get("game_id"),
                            "model_name": model_name,
                            "p_home_win": r.get("home_win_prob"),
                            "margin_mean": r.get("margin_mean"),
                            "margin_sd": r.get("margin_sd"),
                            "total_mean": (
                                r.get("total")
                                if not _is_missing(r.get("total"))
                                else r.get("total_mean")
                            ),
                            "total_sd": r.get("total_sd"),
                        }
                    )
    return rows_by_market


def _validate_market_tuning_inputs(
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    models: list[str],
    ensemble_ids: dict[str, str],
    strict: bool,
) -> None:
    missing_model_market: list[tuple[str, str]] = []
    missing_ensemble_market: list[tuple[str, str]] = []
    for market in (Market.ML, Market.SPREAD, Market.TOTAL):
        for model_name in models:
            resolved = resolve_effective_params(
                db_path=db_path,
                sport=sport,
                season=season,
                model=model_name,
                market=market.name,
            )
            if resolved.params is None:
                missing_model_market.append((model_name, market.name))
        ensemble_id = ensemble_ids.get(market.name)
        if ensemble_id:
            weights = get_active_ensemble_market_weights(
                db_path,
                sport=sport,
                season=season,
                market=market.name,
                ensemble_id=ensemble_id,
            )
            if weights is None:
                missing_ensemble_market.append((market.name, ensemble_id))
    if missing_model_market or missing_ensemble_market:
        print(f"[tuning] DB: {db_path}")
        for model_name, market_name in missing_model_market:
            print(
                "[tuning] Missing active params for "
                f"model={model_name} market={market_name}; using defaults."
            )
        for market_name, ensemble_id in missing_ensemble_market:
            print(
                "[tuning] Missing active ensemble weights for "
                f"market={market_name} ensemble_id={ensemble_id}; using file or equal weights."
            )
        if missing_model_market:
            model_arg = (
                models[0] if len(models) == 1 else "all"
            )
            suggestion = (
                "python -m src.cli.pipeline bootstrap-market-actives "
                f"--sport {sport} --season {season} --model {model_arg}"
            )
            print(f"[tuning] To bootstrap missing model market actives, run: {suggestion}")
        if strict:
            missing_count = len(missing_model_market) + len(missing_ensemble_market)
            raise ValueError(
                "Missing active market tuning inputs "
                f"({missing_count}). Run bootstrap-market-actives or disable --strict."
            )


def _collect_market_param_sources(
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    models: list[str],
) -> dict[str, dict[str, str | None]]:
    sources: dict[str, dict[str, str | None]] = {}
    for market in (Market.ML, Market.SPREAD, Market.TOTAL):
        market_sources: dict[str, str | None] = {}
        for model_name in models:
            resolved = resolve_effective_params(
                db_path=db_path,
                sport=sport,
                season=season,
                model=model_name,
                market=market.name,
            )
            if resolved.params is not None:
                metric_display = (
                    resolved.metric_optimized.replace("backtest_", "", 1)
                    if resolved.metric_optimized and str(resolved.metric_optimized).startswith("backtest_")
                    else resolved.metric_optimized
                )
                # Include tuned metric in reported source when available
                market_sources[model_name] = (
                    f"{resolved.params_source_label}/{metric_display}"
                    if metric_display
                    else resolved.params_source_label
                )
            else:
                market_sources[model_name] = None
        sources[market.name] = market_sources
    return sources


def _collect_ensemble_weight_sources(
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    ensemble_ids: dict[str, str],
) -> dict[str, str | None]:
    sources: dict[str, str | None] = {}
    for market in (Market.ML, Market.SPREAD, Market.TOTAL):
        ensemble_id = ensemble_ids.get(market.name)
        if not ensemble_id:
            continue
        sources[market.name] = get_active_ensemble_market_weights_source(
            db_path,
            sport=sport,
            season=season,
            market=market.name,
            ensemble_id=ensemble_id,
        )
    return sources


def _format_game_name(away_team: Any, home_team: Any) -> str:
    """Render a simple matchup label."""
    away = str(away_team or "").strip()
    home = str(home_team or "").strip()
    if away and home:
        return f"{away} @ {home}"
    return away or home


def _dashboard_rows_for_today(
    schedule_df: pd.DataFrame,
    model_name: str,
    as_of_date: date | None = None,
    *,
    model_metadata: dict[str, Any] | None = None,
) -> list[Dict[str, Any]]:
    """Collect scheduled games for the current day for the dashboard sheet."""
    if schedule_df.empty:
        return []

    today = as_of_date or pd.Timestamp.today().date()
    df = schedule_df.assign(
        _date=pd.to_datetime(schedule_df["date"], errors="coerce").dt.date
    )
    df = df[df["status"] == "scheduled"]
    if df.empty:
        return []

    if as_of_date is not None:
        df = df[df["_date"] == today]
    else:
        df = df[df["_date"] == today]

    if df.empty:
        return []

    rows: list[Dict[str, Any]] = []
    metadata = model_metadata or {}
    for _, row in df.iterrows():
        projected_home_score = row.get("projected_home_score")
        projected_away_score = row.get("projected_away_score")
        projected_total = row.get("projected_total")
        total_mean = row.get("total_mean")
        if not _is_missing(total_mean):
            total = total_mean
        elif projected_home_score is not None and projected_away_score is not None:
            total = float(projected_home_score) + float(projected_away_score)
        else:
            total = projected_total
        rows.append(
            {
                "_sort_dt": pd.to_datetime(row.get("date"), errors="coerce"),
                "_sort_game_id": row.get("game_id"),
                "model": model_name,
                "model_version": metadata.get("model_version"),
                "params_source": metadata.get("params_source"),
                "params_source_label": metadata.get("params_source_label"),
                "params_source_run_id": metadata.get("params_source_run_id"),
                "tuned_metric_used": metadata.get("tuned_metric_used"),
                "params_metric_optimized": metadata.get("params_metric_optimized"),
                "params_best_score": metadata.get("params_best_score"),
                "params_fingerprint": metadata.get("params_fingerprint"),
                "params_nonempty": metadata.get("params_nonempty"),
                "tuning_run_id": metadata.get("tuning_run_id"),
                "params_market": metadata.get("params_market"),
                "run_timestamp_utc": metadata.get("run_timestamp_utc"),
                "prediction_hash": metadata.get("prediction_hash"),
                "date": row.get("date"),
                "game": _format_game_name(row.get("away_team"), row.get("home_team")),
                "projected_home_score": projected_home_score,
                "projected_away_score": projected_away_score,
                "total": total,
                "projected_winner": row.get("projected_winner"),
                "projected_spread": row.get("projected_spread"),
                "margin_mean": row.get("margin_mean"),
                "margin_sd": row.get("margin_sd"),
                "model_p_home_win": row.get("model_p_home_win"),
                "normal_p_home_win": row.get("normal_p_home_win"),
                "home_win_prob": row.get("home_win_prob"),
                "away_win_prob": row.get("away_win_prob"),
                "winner_win_prob": row.get("winner_win_prob"),
                "logistic_home_win_prob": row.get("logistic_home_win_prob"),
                "win_prob_source": row.get("win_prob_source"),
                "margin_dist_assumption": row.get("margin_dist_assumption"),
                "total_sd": row.get("total_sd"),
            }
        )
    return rows


def _filter_games_through(
    df: pd.DataFrame, as_of_date: date | None
) -> pd.DataFrame:
    """Filter games up to and including a cutoff date."""
    if as_of_date is None or df.empty or "date" not in df.columns:
        return df
    parsed = pd.to_datetime(as_of_date, errors="coerce")
    if pd.isna(parsed):
        return df
    cutoff = parsed.date()
    dates = pd.to_datetime(df["date"], errors="coerce").dt.date
    return df[dates <= cutoff]


def _resolve_as_of_date(value: str | date | None) -> date:
    if value is None:
        return pd.Timestamp.today().date()
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception as exc:
        raise ValueError(f"Invalid as_of_date: {value}") from exc


def _resolve_bets_model(models: list[str], bets_model: str | None) -> str:
    if bets_model:
        normalized = normalize_model_name(bets_model)
        if normalized not in models:
            raise ValueError(
                f"bets_model={bets_model!r} not in schedule models: {', '.join(models)}"
            )
        return normalized
    if len(models) == 1:
        return models[0]
    return models[0]


def _resolve_best_ensemble_weights(
    *,
    db_path: str | Path | None,
    sport: str,
    season: str,
    market: str,
    ensemble_id: str,
) -> tuple[dict[str, float] | None, list[str] | None, str | None]:
    if db_path is None:
        return None, None, None
    metric_optimized = _metric_optimized_label(market)
    weights, run_id = load_best_ensemble_market_tuning_weights_by_optimized_metric(
        db_path,
        sport=sport,
        season=season,
        market=market,
        ensemble_id=ensemble_id,
        metric_optimized=metric_optimized,
    )
    if not weights:
        return None, None, None
    return weights, list(weights.keys()), run_id


def _resolve_ensemble_weights(
    *,
    db_path: str | Path | None,
    sport: str,
    season: str,
    market: str,
    ensemble_id: str,
    config_weights: dict[str, float] | None,
    selection_context: dict[str, object] | None = None,
    tuning_context: dict[str, object] | None = None,
    config_warnings: list[str] | None = None,
) -> tuple[
    dict[str, float] | None,
    list[str] | None,
    str | None,
    str | None,
    str | None,
]:
    if selection_context:
        if tuning_context and _selection_matches_tuning(selection_context, tuning_context):
            return (
                tuning_context["weights"],
                tuning_context.get("models") or selection_context["models"],
                "db_tuned",
                tuning_context["run_id"],
                selection_context["run_id"],
            )
        if tuning_context:
            warning = (
                f"Tuning run {tuning_context['run_id']} selection mismatch "
                f"(tuning selection={tuning_context.get('selection_models')} "
                f"vs active selection={selection_context['models']}); using equal weights."
            )
            _LOG.warning(warning)
            if config_warnings is not None:
                config_warnings.append(warning)
        models = selection_context["models"]
        equal_weights = {model: 1.0 / len(models) for model in models}
        return equal_weights, models, "selection_equal", None, selection_context["run_id"]

    weights, models, run_id = _resolve_best_ensemble_weights(
        db_path=db_path,
        sport=sport,
        season=season,
        market=market,
        ensemble_id=ensemble_id,
    )
    if weights is not None:
        return weights, models, "db_best_run", run_id, None
    if config_weights is not None:
        return config_weights, None, "config", None, None
    if db_path is not None:
        weights, models = get_active_ensemble_market_weights_and_models(
            db_path,
            sport=sport,
            season=season,
            market=market,
            ensemble_id=ensemble_id,
        )
        if weights is not None:
            source_run_id = get_active_ensemble_market_weights_source(
                db_path,
                sport=sport,
                season=season,
                market=market,
                ensemble_id=ensemble_id,
            )
            return weights, models, "db_active", source_run_id, None
    weights = load_market_weights(sport, season, market, ensemble_id)
    if weights is not None:
        return weights, None, "file", None, None
    return None, None, None, None, None


def _market_required_columns(market: str | Market) -> tuple[str, ...]:
    market_name = market if isinstance(market, str) else market.name
    market_key = market_name.upper()
    if market_key == Market.ML.name:
        return ("p_home_win",)
    if market_key == Market.SPREAD.name:
        return ("margin_mean", "margin_sd")
    if market_key == Market.TOTAL.name:
        return ("total_mean", "total_sd")
    return ()


def _filter_market_weights_for_forecast(
    *,
    weights: dict[str, float] | None,
    forecast_df: pd.DataFrame,
    market: str | Market,
) -> tuple[dict[str, float], set[str]]:
    market_name = market if isinstance(market, str) else market.name
    market_key = market_name.upper()
    required_columns = _market_required_columns(market_key)

    forecast_models: set[str] = set()
    model_validity: dict[str, bool] = {}
    invalid_reasons: dict[str, str] = {}

    for model_name, group in forecast_df.groupby("model_name"):
        if pd.isna(model_name):
            continue
        normalized = normalize_model_name(str(model_name))
        if not normalized:
            continue
        forecast_models.add(normalized)
        missing: list[str] = []
        for column in required_columns:
            if column not in group.columns or not group[column].notna().any():
                missing.append(column)
        if missing:
            model_validity[normalized] = False
            invalid_reasons[normalized] = f"missing {', '.join(missing)}"
        else:
            model_validity[normalized] = True

    candidate_weights: dict[str, float] = {}
    if weights is not None:
        for model_name, value in weights.items():
            normalized = normalize_model_name(model_name)
            if not normalized:
                continue
            try:
                weight_value = float(value)
            except Exception:
                weight_value = 0.0
            candidate_weights[normalized] = weight_value
    else:
        for model in forecast_models:
            candidate_weights[model] = 1.0

    for model in forecast_models:
        candidate_weights.setdefault(model, 0.0)

    filtered_weights: dict[str, float] = {}
    drop_reasons: dict[str, str] = {}
    for model, weight in candidate_weights.items():
        if weight <= 0:
            drop_reasons[model] = f"weight={weight}"
            continue
        if model not in forecast_models:
            drop_reasons[model] = "missing forecast rows"
            continue
        if required_columns and not model_validity.get(model, True):
            reason = invalid_reasons.get(model, "missing required fields")
            drop_reasons[model] = reason
            continue
        filtered_weights[model] = weight

    valid_models = {m for m, valid in model_validity.items() if valid}
    if len(filtered_weights) <= 1 and len(valid_models) > 1:
        fallback_models = sorted(valid_models)
        uniform_weights = {model: 1.0 / len(fallback_models) for model in fallback_models}
        _LOG.info(
            "[_filter_market_weights_for_forecast] Tuned weights for %s collapsed to %d model(s) "
            "after filtering (candidates=%s); falling back to uniform weights over %s.",
            market_key,
            len(filtered_weights),
            json.dumps(candidate_weights, sort_keys=True),
            fallback_models,
        )
        filtered_weights = uniform_weights

    def _format_drop_info(source: dict[str, str]) -> str:
        if not source:
            return "none"
        entries = "; ".join(f"{m}: {source[m]}" for m in sorted(source))
        return entries

    if not filtered_weights:
        if not valid_models:
            _LOG.error(
                "[_filter_market_weights_for_forecast] No valid %s ensemble members after filtering "
                "(candidates=%s, drops=%s). No fallback available.",
                market_key,
                json.dumps(candidate_weights, sort_keys=True),
                _format_drop_info(drop_reasons),
            )
            raise ValueError(
                f"No market-valid models available for {market_key}; cannot run ensemble."
            )
        fallback_weights = {
            model: 1.0 / len(valid_models) for model in sorted(valid_models)
        }
        _LOG.error(
            "[_filter_market_weights_for_forecast] No positive weights for %s after filtering "
            "(candidates=%s, drops=%s). Falling back to uniform weights over %s.",
            market_key,
            json.dumps(candidate_weights, sort_keys=True),
            _format_drop_info(drop_reasons),
            sorted(valid_models),
        )
        filtered_weights = fallback_weights
    total_weight = sum(filtered_weights.values())
    normalized_weights = {
        model: weight / total_weight for model, weight in filtered_weights.items()
    }

    if drop_reasons and normalized_weights:
        _LOG.info(
            "[_filter_market_weights_for_forecast] Dropped models for %s: %s",
            market_key,
            _format_drop_info(drop_reasons),
        )

    return normalized_weights, set(normalized_weights.keys())


def _load_active_selection_context(
    *,
    db_path: str | Path | None,
    sport: str,
    season: str,
    market: str,
    ensemble_id: str,
) -> dict[str, object] | None:
    if db_path is None:
        return None
    pointer = get_active_ensemble_market_selection(
        db_path=db_path,
        sport=sport,
        season=season,
        market=market,
        ensemble_id=ensemble_id,
    )
    if not pointer:
        return None
    selection = load_selection_run(db_path, run_id=pointer["active_run_id"])
    if not selection:
        return None
    models = [normalize_model_name(m) for m in selection.get("selected", []) if m]
    if not models:
        return None
    return {"run_id": selection["run_id"], "models": models}


def _load_active_tuning_context(
    *,
    db_path: str | Path | None,
    sport: str,
    season: str,
    market: str,
    ensemble_id: str,
) -> dict[str, object] | None:
    if db_path is None:
        return None
    pointer = get_active_ensemble_market_tuning_run(
        db_path=db_path,
        sport=sport,
        season=season,
        market=market,
        ensemble_id=ensemble_id,
    )
    if not pointer:
        return None
    tuning = load_tuning_run(db_path, run_id=pointer["active_run_id"])
    if not tuning:
        return None
    return {
        "run_id": tuning["run_id"],
        "weights": tuning.get("weights") or {},
        "models": tuning.get("models") or [],
        "selection_run_id": tuning.get("selection_run_id"),
        "selection_models": tuning.get("selection_models"),
    }


def _selection_matches_tuning(
    selection_context: dict[str, object], tuning_context: dict[str, object]
) -> bool:
    selection_run_id = selection_context.get("run_id")
    tuning_selection_run_id = tuning_context.get("selection_run_id")
    if tuning_selection_run_id and selection_run_id:
        return tuning_selection_run_id == selection_run_id
    tuning_selection_models = tuning_context.get("selection_models")
    if tuning_selection_models:
        return tuning_selection_models == selection_context["models"]
    return tuning_context.get("models") == selection_context["models"]

def _sanitize_source_id(value: Any, *, default: str = "direct") -> str:
    text = _normalize_source_label(value, default=default).lower()
    cleaned = re.sub(r"[^a-z0-9._-]+", "-", text).strip("-")
    return cleaned or default


def _normalize_source_label(value: Any, *, default: str = "direct") -> str:
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    text = str(value).strip()
    return text if text else default


def _is_missing(value: Any) -> bool:
    """Check if a value is None, NaN, or empty string."""
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return isinstance(value, str) and not value.strip()


def _first_nonempty_source(
    schedule_df: pd.DataFrame, column: str, *, as_of_date: date
) -> str | None:
    if schedule_df.empty or column not in schedule_df.columns:
        return None
    subset = schedule_df
    try:
        if "status" in schedule_df.columns and "date" in schedule_df.columns:
            _dt = pd.to_datetime(schedule_df["date"], errors="coerce").dt.date
            mask = (schedule_df["status"] == "scheduled") & (_dt == as_of_date)
            subset = schedule_df.loc[mask]
    except Exception:
        subset = schedule_df
    for value in subset[column]:
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        text = str(value).strip()
        if text:
            return text
    return None


def _deterministic_review_run_id(
    *,
    sport: str,
    season: str,
    as_of_date: date,
    ml_source_id: str,
    spread_source_id: str,
    total_source_id: str,
) -> str:
    ml_id = _sanitize_source_id(ml_source_id)
    spread_id = _sanitize_source_id(spread_source_id)
    total_id = _sanitize_source_id(total_source_id)
    return (
        f"schedule-{sport}-{season}-{as_of_date.isoformat()}"
        f"-ml_{ml_id}-spread_{spread_id}-total_{total_id}"
    )


def _build_bets_dataframe(
    schedule_df: pd.DataFrame,
    *,
    model_name: str,
    as_of_date: date,
    review_run_id: str,
    db_path: str | Path | None = None,
    sport: str | None = None,
    season: str | None = None,
    spread_ensemble_applied: bool = False,
    total_ensemble_applied: bool = False,
) -> pd.DataFrame:
    """Build the BETS dataframe from a schedule dataframe.

    Creates exactly 6 rows per game (2×ML, 2×SPREAD, 2×TOTAL).
    - ML rows: source_id from win_prob_source (direct or model ensemble)
    - SPREAD rows: always source_id="ensemble_spread_v1" (ensemble system is source of truth)
    - TOTAL rows: always source_id="ensemble_total_v1" (ensemble system is source of truth)
    - Enriches all rows with market line data via left-join lookups
    - No duplicate direct+ensemble variants; source selected up-front based on ensemble availability

    Note: spread_ensemble_applied and total_ensemble_applied parameters are accepted for
    backward compatibility but are no longer used for row gating. SPREAD/TOTAL always come
    from ensemble sources (no fallback to direct). These parameters may be removed in future.

    Args:
        schedule_df: Schedule DataFrame with forecasts
        model_name: Primary model name
        as_of_date: As-of date for this build
        review_run_id: Unique identifier for this review run
        db_path: Path to SQLite database for market line lookups
        sport: Sport identifier for team ID resolution
        season: Season identifier for team ID resolution
        spread_ensemble_applied: Deprecated; kept for backward compatibility
        total_ensemble_applied: Deprecated; kept for backward compatibility

    Logging coverage:
    - INFO: Entry/exit with row counts, games processed, market line stats, source labels
    - DEBUG: Per-game details, team ID resolution, market line lookups
    - WARNING: Anomalies (duplicates, invariant violations, missing data)
    """
    # -------------------------------------------------------------------------
    # ENTRY LOGGING: Log function entry with input statistics
    # -------------------------------------------------------------------------
    _LOG.info(
        "[_build_bets_dataframe] ENTRY: model=%s, as_of_date=%s, review_run_id=%s, "
        "input_rows=%d, sport=%s, season=%s",
        model_name,
        as_of_date,
        review_run_id,
        len(schedule_df),
        sport,
        season,
    )

    include_calibrated = (
        "home_win_prob_calibrated" in schedule_df.columns
        and schedule_df["home_win_prob_calibrated"].notna().any()
    )

    # Log whether calibrated columns are present (helps debug calibration pipeline issues)
    _LOG.debug(
        "[_build_bets_dataframe] Calibrated columns detected: %s",
        include_calibrated,
    )
    bets_columns = [
        "review_run_id",
        "game_id",
        "date",
        "away_team",
        "home_team",
        "market_type",
        "selection",
        "line",
        "odds",
        "price",
        "implied_prob",
        "model_prob",
        "edge",
        "ev",
        "stake",
        "book",
        "bet_id",
        "logged_at",
        "log_status",
        "notes",
        "source_market_snapshot_id",
        "opportunity_id",
        "home_win_prob",
        "away_win_prob",
        "win_prob_source",
    ]
    if include_calibrated:
        bets_columns.extend(
            [
                "home_win_prob_raw",
                "away_win_prob_raw",
                "home_win_prob_calibrated",
                "away_win_prob_calibrated",
            ]
        )
    bets_columns.extend(
        [
            "margin_mean",
            "margin_sd",
            "total",
            "total_sd",
            "model",
            "market_forecast_source",
            "ml_ensemble_components_json",
            "spread_source",
            "spread_ensemble_components_json",
            "total_source",
            "total_ensemble_components_json",
            "spread_prob_calibrated",
            "total_prob_calibrated",
        ]
    )

    if schedule_df.empty:
        _LOG.info("[_build_bets_dataframe] EXIT: Input schedule empty, returning empty BETS dataframe")
        return pd.DataFrame(columns=bets_columns)

    # -------------------------------------------------------------------------
    # INPUT VALIDATION: Log input statistics before any processing
    # -------------------------------------------------------------------------
    if not schedule_df.empty and "game_id" in schedule_df.columns:
        input_game_count = schedule_df["game_id"].nunique()
        input_total_rows = len(schedule_df)
        if input_total_rows > input_game_count:
            _LOG.warning(
                f"BETS input has {input_total_rows} rows but only {input_game_count} unique game_ids. "
                f"Ratio: {input_total_rows / input_game_count:.1f}x"
            )

    df = schedule_df.assign(
        _date=pd.to_datetime(schedule_df["date"], errors="coerce").dt.date
    )
    # Filter to games marked as scheduled (no scores yet)
    # NOTE: We do NOT filter by as_of_date here, because the schedule_df
    # passed in is already filtered by market CSV. Applying an additional
    # date filter would exclude multi-day betting data.
    # The as_of_date parameter is kept for backwards compatibility but not used here.
    if "status" in df.columns:
        df = df[df["status"] == "scheduled"]
    
    # CRITICAL DEDUPLICATION: Ensure each game_id appears exactly once in input.
    # If the input schedule contains duplicate game_ids (e.g., from multiple market runs),
    # keep only the first occurrence to prevent creating multiple 6-row blocks per game.
    if "game_id" in df.columns:
        initial_count = len(df)
        # Log details about duplicates BEFORE deduplicating
        if df.duplicated(subset=["game_id"], keep=False).any():
            dup_games = df[df.duplicated(subset=["game_id"], keep=False)]["game_id"].unique()
            _LOG.warning(
                f"Found {len(dup_games)} game_id(s) with duplicates in BETS input. "
                f"Sample: {list(dup_games[:3])}. Total input rows: {initial_count}."
            )
        
        df = df.drop_duplicates(subset=["game_id"], keep="first")
        if len(df) < initial_count:
            _LOG.warning(
                f"Deduplicated {initial_count - len(df)} duplicate game_id(s) in BETS input. "
                f"Final unique games: {len(df)}."
            )
    
    # Ensure deterministic ordering so selection rows list the away team above
    # the home team (matches schedule sheet presentation). Sort by date,
    # game_id, and away_team to guarantee away-first ordering per game.
    sort_cols: list[str] = []
    if "date" in df.columns:
        sort_cols.append("date")
    if "game_id" in df.columns:
        sort_cols.append("game_id")
    # final tie-breaker: away_team first
    if "away_team" in df.columns:
        sort_cols.append("away_team")
    if sort_cols:
        try:
            df = df.sort_values(by=sort_cols, kind="mergesort")
        except Exception:
            # best-effort: ignore sorting failures and continue
            pass
    if df.empty:
        _LOG.info("[_build_bets_dataframe] EXIT: No games after filtering, returning empty BETS dataframe")
        return pd.DataFrame(columns=bets_columns)

    # -------------------------------------------------------------------------
    # TEAM ID RESOLUTION: Cache team IDs for market line lookups
    # Tracks cache hits/misses for debugging lookup performance
    # -------------------------------------------------------------------------
    team_id_cache: dict[tuple[str | None, str | None, str], int | None] = {}
    # Counters for logging team ID resolution statistics
    _team_id_cache_hits = 0
    _team_id_cache_misses = 0
    _team_id_lookup_failures = 0

    def _resolve_team_id_for(name: str | None) -> int | None:
        nonlocal _team_id_cache_hits, _team_id_cache_misses, _team_id_lookup_failures
        if not name or db_path is None or not sport or not season:
            return None
        key = (sport, season, name)
        if key in team_id_cache:
            _team_id_cache_hits += 1
            return team_id_cache[key]
        _team_id_cache_misses += 1
        team_id = None
        try:
            with sqlite3.connect(Path(db_path)) as team_conn:
                team_id = team_repo.get_team_id(
                    team_conn,
                    sport=sport,
                    season=season,
                    canonical_name=name,
                )
        except Exception:
            team_id = None
            _team_id_lookup_failures += 1
        team_id_cache[key] = team_id
        return team_id

    # -------------------------------------------------------------------------
    # MARKET LINE LOOKUPS: Import betting_repository for enriching rows with odds
    # Tracks found/missing market lines per market type for diagnostics
    # -------------------------------------------------------------------------
    br = None
    if db_path is not None:
        try:
            from src.data import betting_repository as br
            _LOG.debug("[_build_bets_dataframe] betting_repository loaded for market line lookups")
        except Exception as e:
            _LOG.debug("[_build_bets_dataframe] betting_repository not available: %s", e)
            br = None

    # Counters for market line lookup statistics (logged at end)
    _market_line_stats: dict[str, dict[str, int]] = {
        "ML": {"found": 0, "missing": 0, "error": 0},
        "spread": {"found": 0, "missing": 0, "error": 0},
        "total": {"found": 0, "missing": 0, "error": 0},
    }

    # CANONICAL ROW ASSEMBLY: For each game, create exactly 6 rows (2×ML, 2×spread, 2×total),
    # enrich them in-place, then move to the next game. No additional rows appended after
    # a game's canonical 6 rows are complete.
    rows: list[dict[str, Any]] = []
    _games_processed = 0
    
    for _, game_row in df.iterrows():
        game_id = game_row.get("game_id")
        home_team = game_row.get("home_team")
        away_team = game_row.get("away_team")
        
        # Resolve team IDs for market line lookups
        home_team_id = _resolve_team_id_for(home_team)
        away_team_id = _resolve_team_id_for(away_team)
        
        # Resolve total/projected_total before row creation
        projected_home_score = game_row.get("projected_home_score")
        projected_away_score = game_row.get("projected_away_score")
        projected_total = game_row.get("projected_total")
        total = game_row.get("total")
        total_mean = game_row.get("total_mean")
        if _is_missing(total):
            if not _is_missing(total_mean):
                total = total_mean
            elif projected_home_score is not None and projected_away_score is not None:
                total = float(projected_home_score) + float(projected_away_score)
            else:
                total = projected_total

        # Normalize source labels for market-specific metadata
        ml_source_label = _normalize_source_label(
            game_row.get("win_prob_source") or game_row.get("ml_source")
        )
        spread_source_label = _normalize_source_label(game_row.get("spread_source"))
        total_source_label = _normalize_source_label(game_row.get("total_source"))
        
        win_prob_source = game_row.get("win_prob_source")
        if _is_missing(win_prob_source):
            win_prob_source = ml_source_label
        
        spread_source = game_row.get("spread_source")
        if _is_missing(spread_source):
            spread_source = spread_source_label
        
        total_source = game_row.get("total_source")
        if _is_missing(total_source):
            total_source = total_source_label

        # Build shared base row template for all 6 selections of this game.
        # Include all forecast fields (initialized to blanks; overridden per market type).
        base_row = {
            "review_run_id": review_run_id,
            "game_id": game_id,
            "date": _safe_date(game_row.get("date")),
            "away_team": away_team,
            "home_team": home_team,
            "line": "",
            "odds": "",
            "price": "",
            "implied_prob": "",
            "model_prob": "",
            "edge": "",
            "ev": "",
            "stake": "",
            "book": "",
            "bet_id": "",
            "logged_at": "",
            "log_status": "",
            "notes": "",
            "source_market_snapshot_id": "",
            "opportunity_id": "",
            "model": model_name,
            "market_forecast_source": "",
            # Initialize all forecast fields to blank; overridden per market type below
            "home_win_prob": "",
            "away_win_prob": "",
            "win_prob_source": "",
            "margin_mean": "",
            "margin_sd": "",
            "total": "",
            "total_sd": "",
            "ml_ensemble_components_json": "",
            "spread_source": "",
            "spread_ensemble_components_json": "",
            "total_source": "",
            "total_ensemble_components_json": "",
            "spread_prob_calibrated": "",
            "total_prob_calibrated": "",
        }
        
        # Add calibrated fields if present
        if include_calibrated:
            base_row.update({
                "home_win_prob_raw": "",
                "away_win_prob_raw": "",
                "home_win_prob_calibrated": "",
                "away_win_prob_calibrated": "",
            })


        # CANONICAL 6 ROWS: Construct exactly once per game (2×ML, 2×SPREAD, 2×TOTAL)
        # Source IDs are selected up-front:
        # - ML: uses ml_source_label (direct or model ensemble, depends on forecast pipeline)
        # - SPREAD: always "ensemble_spread_v1" (ensemble is the source of truth)
        # - TOTAL: always "ensemble_total_v1" (ensemble is the source of truth)
        # No duplicate direct+ensemble variants; each market type appears exactly once per game.
        canonical_specs = [
            # 2 × ML (source: ml_source_label, which may be direct or ensemble depending on pipeline)
            ("ML", away_team, ml_source_label, "home_win_prob", "away_win_prob", "win_prob_source", "ml_ensemble_components_json"),
            ("ML", home_team, ml_source_label, "home_win_prob", "away_win_prob", "win_prob_source", "ml_ensemble_components_json"),
            # 2 × SPREAD (always ensemble-sourced: ensemble_spread_v1)
            ("spread", away_team, "ensemble_spread_v1", "margin_mean", "margin_sd", "spread_source", "spread_ensemble_components_json"),
            ("spread", home_team, "ensemble_spread_v1", "margin_mean", "margin_sd", "spread_source", "spread_ensemble_components_json"),
            # 2 × TOTAL (always ensemble-sourced: ensemble_total_v1)
            ("total", "Over", "ensemble_total_v1", "total", "total_sd", "total_source", "total_ensemble_components_json"),
            ("total", "Under", "ensemble_total_v1", "total", "total_sd", "total_source", "total_ensemble_components_json"),
        ]
        
        for market_type, selection, source_label, prob_col1, prob_col2, source_col, ensemble_col in canonical_specs:
                
            # Create canonical row for this market/selection
            canonical_row = dict(base_row)
            canonical_row["market_type"] = market_type
            canonical_row["selection"] = selection
            canonical_row["market_forecast_source"] = source_label
            canonical_row["model"] = source_label
            
            # Enrich with market-specific forecast data (in-place, no additional rows)
            if market_type == "ML":
                canonical_row.update({
                    "home_win_prob": game_row.get("home_win_prob"),
                    "away_win_prob": game_row.get("away_win_prob"),
                    "win_prob_source": win_prob_source,
                    "ml_ensemble_components_json": game_row.get("ml_ensemble_components_json"),
                })
                if include_calibrated:
                    canonical_row.update({
                        "home_win_prob_raw": game_row.get("home_win_prob_raw"),
                        "away_win_prob_raw": game_row.get("away_win_prob_raw"),
                        "home_win_prob_calibrated": game_row.get("home_win_prob_calibrated"),
                        "away_win_prob_calibrated": game_row.get("away_win_prob_calibrated"),
                    })
            elif market_type == "spread":
                canonical_row.update({
                    "margin_mean": game_row.get("margin_mean"),
                    "margin_sd": game_row.get("margin_sd"),
                    "spread_source": spread_source,
                    "spread_ensemble_components_json": game_row.get("spread_ensemble_components_json"),
                })
            else:  # total
                canonical_row.update({
                    "total": total,
                    "total_sd": game_row.get("total_sd"),
                    "total_source": total_source,
                    "total_ensemble_components_json": game_row.get("total_ensemble_components_json"),
                })
            
            # LEFT-JOIN market lines: lookup only if market_type and selection match;
            # if missing, cells remain blank (no separate rows created for missing lines)
            snap = None
            if br is not None:
                try:
                    if market_type in ("ML", "spread"):
                        if selection == home_team:
                            snap = br.get_latest_market_line(
                                db_path,
                                sport=sport,
                                season=season,
                                game_id=str(game_id),
                                market_type=market_type,
                                selection_team_id=home_team_id,
                            )
                        elif selection == away_team:
                            snap = br.get_latest_market_line(
                                db_path,
                                sport=sport,
                                season=season,
                                game_id=str(game_id),
                                market_type=market_type,
                                selection_team_id=away_team_id,
                            )
                    elif market_type == "total":
                        snap = br.get_latest_market_line(
                            db_path,
                            sport=sport,
                            season=season,
                            game_id=str(game_id),
                            market_type=market_type,
                            selection=selection,
                        )
                except Exception as e:
                    snap = None
                    _market_line_stats[market_type]["error"] += 1
                    _LOG.debug(
                        "[_build_bets_dataframe] Market line lookup error: game_id=%s, market=%s, selection=%s, error=%s",
                        game_id, market_type, selection, e
                    )

            # Track market line found/missing statistics for summary logging
            if snap and snap.get("line") is not None:
                _market_line_stats[market_type]["found"] += 1
            elif br is not None:
                # Only count as missing if we actually tried to look up (br is available)
                _market_line_stats[market_type]["missing"] += 1

            # Enrich canonical row with market line data (left-join: blanks if missing)
            canonical_row["line"] = snap.get("line") if snap and snap.get("line") is not None else ""
            canonical_row["odds"] = int(snap.get("odds")) if snap and snap.get("odds") is not None else ""
            canonical_row["source_market_snapshot_id"] = snap.get("id") if snap and snap.get("id") is not None else ""

            # Append canonical row (exactly once per game/market/selection combination)
            rows.append(canonical_row)

        # Increment games processed counter after all 6 rows for this game are created
        _games_processed += 1

    # -------------------------------------------------------------------------
    # SUMMARY LOGGING: Log statistics about the BETS dataframe construction
    # -------------------------------------------------------------------------
    # Log source configuration for transparency
    _LOG.info(
        "[BETS] Source configuration: ML varies by forecast, SPREAD=ensemble_spread_v1, TOTAL=ensemble_total_v1 "
        "(no duplicate direct rows; ensemble is authoritative for SPREAD/TOTAL)"
    )
    
    _LOG.info(
        "[_build_bets_dataframe] Row assembly complete: games_processed=%d, total_rows=%d (expected=%d)",
        _games_processed,
        len(rows),
        _games_processed * 6,
    )

    # Log team ID resolution statistics
    _LOG.debug(
        "[_build_bets_dataframe] Team ID resolution: cache_hits=%d, cache_misses=%d, lookup_failures=%d, cache_size=%d",
        _team_id_cache_hits,
        _team_id_cache_misses,
        _team_id_lookup_failures,
        len(team_id_cache),
    )

    # Log market line lookup statistics (important for diagnosing missing odds)
    for mkt, stats in _market_line_stats.items():
        total_lookups = stats["found"] + stats["missing"] + stats["error"]
        if total_lookups > 0:
            found_pct = (stats["found"] / total_lookups * 100) if total_lookups > 0 else 0
            _LOG.info(
                "[_build_bets_dataframe] Market line stats [%s]: found=%d (%.1f%%), missing=%d, errors=%d",
                mkt, stats["found"], found_pct, stats["missing"], stats["error"],
            )

    # INVARIANT ENFORCEMENT: Assert each game_id has the expected number of rows
    # Rows per game = 2 (ML) + (2 if not spread_ensemble_applied else 0) + (2 if not total_ensemble_applied else 0)
    # = 2 ML + 2 SPREAD (unless ensemble) + 2 TOTAL (unless ensemble)
    expected_rows_per_game = 2  # ML rows always present
    if not spread_ensemble_applied:
        expected_rows_per_game += 2  # Direct SPREAD rows
    if not total_ensemble_applied:
        expected_rows_per_game += 2  # Direct TOTAL rows
    
    bets_df = pd.DataFrame(rows, columns=bets_columns)
    if not bets_df.empty and "game_id" in bets_df.columns:
        game_counts = bets_df["game_id"].value_counts()
        invalid_counts = game_counts[game_counts != expected_rows_per_game]
        if not invalid_counts.empty:
            _LOG.warning(
                f"BETS invariant violation: {len(invalid_counts)} game(s) do not have exactly {expected_rows_per_game} rows: {invalid_counts.to_dict()}"
            )
    
    # NOTE: _apply_spread_total_calibrators was removed because:
    # 1. The distribution parameters (margin_mean, margin_sd, total, total_sd) are already
    #    calibrated upstream by _apply_calibration_to_schedule_df before being passed here
      # 2. The function had an interface mismatch: it loaded VarianceCalibrator
    #    objects but called .transform(Series) expecting probability calibrators
    # 3. The probabilities computed in _calculate_model_prob use the calibrated distribution
    #    parameters, so additional calibration is redundant
    # See: https://github.com/prairiepilotfpv/sports-power-ratings/issues/XXX

    # Calculate model_prob for each market type
    bets_df = _calculate_model_prob(bets_df)

    # -------------------------------------------------------------------------
    # EXIT LOGGING: Final summary of BETS dataframe construction
    # -------------------------------------------------------------------------
    _LOG.info(
        "[_build_bets_dataframe] EXIT: final_rows=%d, unique_games=%d, model=%s",
        len(bets_df),
        bets_df["game_id"].nunique() if "game_id" in bets_df.columns else 0,
        model_name,
    )

    return bets_df


def _calculate_model_prob(bets_df: pd.DataFrame) -> pd.DataFrame:
    """Calculate model_prob for each market type row.

    For ML: uses home_win_prob or away_win_prob based on selection
    For SPREAD: calculates CDF probability for cover
    For TOTAL: calculates CDF probability for over/under

    This ensures model_prob is populated for database persistence and calibration.

    Logging coverage:
    - INFO: Entry/exit with row counts per market type
    - DEBUG: Detailed statistics about probability calculations
    """
    # -------------------------------------------------------------------------
    # ENTRY LOGGING: Log function entry with input size
    # -------------------------------------------------------------------------
    _LOG.debug(
        "[_calculate_model_prob] ENTRY: total_rows=%d",
        len(bets_df),
    )

    if bets_df.empty or "market_type" not in bets_df.columns:
        _LOG.debug("[_calculate_model_prob] EXIT: Empty input or missing market_type column")
        return bets_df

    out_df = bets_df.copy()

    # Track statistics for each market type
    _prob_stats: dict[str, dict[str, int]] = {
        "ML": {"rows": 0, "populated": 0, "null": 0},
        "spread": {"rows": 0, "populated": 0, "null": 0},
        "total": {"rows": 0, "populated": 0, "null": 0},
    }

    # -------------------------------------------------------------------------
    # ML MARKET: Assign home_win_prob or away_win_prob based on selection
    # -------------------------------------------------------------------------
    if "home_team" in out_df.columns and "selection" in out_df.columns:
        ml_mask = out_df["market_type"] == "ML"
        if ml_mask.any():
            _prob_stats["ML"]["rows"] = int(ml_mask.sum())

            def _ml_prob(row):
                sel = str(row.get("selection", "")).strip()
                home = str(row.get("home_team", "")).strip()
                if sel == home:
                    return row.get("home_win_prob")
                else:
                    return row.get("away_win_prob")

            out_df.loc[ml_mask, "model_prob"] = out_df[ml_mask].apply(_ml_prob, axis=1)

            # Count populated vs null probabilities
            ml_probs = out_df.loc[ml_mask, "model_prob"]
            _prob_stats["ML"]["populated"] = int(ml_probs.notna().sum())
            _prob_stats["ML"]["null"] = int(ml_probs.isna().sum())

    # -------------------------------------------------------------------------
    # SPREAD MARKET: Calculate CDF probability for cover
    # -------------------------------------------------------------------------
    if "spread_source" in out_df.columns:
        spread_mask = out_df["market_type"] == "spread"
        if spread_mask.any():
            _prob_stats["spread"]["rows"] = int(spread_mask.sum())

            out_df.loc[spread_mask, "model_prob"] = out_df[spread_mask].apply(
                _spread_raw_probability, axis=1
            )

            # Count populated vs null probabilities
            spread_probs = out_df.loc[spread_mask, "model_prob"]
            _prob_stats["spread"]["populated"] = int(spread_probs.notna().sum())
            _prob_stats["spread"]["null"] = int(spread_probs.isna().sum())

    # -------------------------------------------------------------------------
    # TOTAL MARKET: Calculate CDF probability for over/under
    # -------------------------------------------------------------------------
    if "total_source" in out_df.columns:
        total_mask = out_df["market_type"] == "total"
        if total_mask.any():
            _prob_stats["total"]["rows"] = int(total_mask.sum())

            out_df.loc[total_mask, "model_prob"] = out_df[total_mask].apply(
                _total_raw_probability, axis=1
            )

            # Count populated vs null probabilities
            total_probs = out_df.loc[total_mask, "model_prob"]
            _prob_stats["total"]["populated"] = int(total_probs.notna().sum())
            _prob_stats["total"]["null"] = int(total_probs.isna().sum())

    # -------------------------------------------------------------------------
    # EXIT LOGGING: Summary of probability calculations by market type
    # -------------------------------------------------------------------------
    for mkt, stats in _prob_stats.items():
        if stats["rows"] > 0:
            populated_pct = (stats["populated"] / stats["rows"] * 100) if stats["rows"] > 0 else 0
            _LOG.info(
                "[_calculate_model_prob] %s: rows=%d, populated=%d (%.1f%%), null=%d",
                mkt, stats["rows"], stats["populated"], populated_pct, stats["null"],
            )

    _LOG.debug("[_calculate_model_prob] EXIT: total_rows=%d", len(out_df))
    return out_df


def _apply_spread_total_calibrators(
    bets_df: pd.DataFrame,
    *,
    sport: str | None,
    season: str | None,
) -> pd.DataFrame:
    """
    DEPRECATED: This function is no longer called as of 2025-01-27.

    This function was designed to apply probability calibration to SPREAD/TOTAL markets
    in the BETS sheet. However, it had a fundamental interface mismatch:

     1. It called _load_market_calibrators() which loads VarianceCalibrator objects
    2. These calibrators expect a DataFrame with (pred_mean, pred_sd) columns
    3. But _apply_market_calibrator() called .transform(pd.Series([raw_prob])) expecting
       a probability calibrator interface

    The distribution parameters (margin_mean, margin_sd, total, total_sd) are now
    calibrated upstream by _apply_calibration_to_schedule_df() before being passed
    to _build_bets_dataframe(). The probabilities computed by _calculate_model_prob()
    already use these calibrated distribution parameters.

    This function is kept for reference but should not be called. If per-row probability
    calibration is needed in the future, implement using ProbabilityFromDistributionCalibrator
    (defined in src/calibration/distribution.py) with proper interface handling.

    See: test_calibration_bets_integration.py for examples of proper calibrator mocking.
    """
    if bets_df.empty:
        return bets_df
    out_df = bets_df.copy()
    for col in ("spread_prob_calibrated", "total_prob_calibrated"):
        if col not in out_df.columns:
            out_df[col] = pd.NA

    spread_sources = set(
        out_df["spread_source"].dropna().astype(str).tolist() if "spread_source" in out_df else []
    )
    total_sources = set(
        out_df["total_source"].dropna().astype(str).tolist() if "total_source" in out_df else []
    )

    spread_cals = _load_market_calibrators(spread_sources, sport, season, Market.SPREAD)
    total_cals = _load_market_calibrators(total_sources, sport, season, Market.TOTAL)

    if spread_sources and not spread_cals:
        _LOG.info(
            f"[_apply_spread_total_calibrators] No spread calibrators available for sources: {', '.join(sorted(spread_sources))}"
        )
    if total_sources and not total_cals:
        _LOG.info(
            f"[_apply_spread_total_calibrators] No total calibrators available for sources: {', '.join(sorted(total_sources))}"
        )

    if spread_cals and "spread_prob_calibrated" in out_df.columns:
        mask = out_df["market_type"] == "spread"
        if mask.any():
            out_df.loc[mask, "spread_prob_calibrated"] = out_df.loc[mask].apply(
                lambda row: _apply_market_calibrator(
                    spread_cals,
                    row.get("spread_source"),
                    _spread_raw_probability(row),
                ),
                axis=1,
            )
            spread_rows = int(mask.sum())
            spread_calibrated = int(out_df.loc[mask, "spread_prob_calibrated"].notna().sum())
            spread_calibrators_label = ", ".join(sorted(spread_cals.keys()))
            _LOG.info(
                f"[_apply_spread_total_calibrators] Applied spread calibrator(s) {spread_calibrators_label} to {spread_calibrated}/{spread_rows} rows"
            )

    if total_cals and "total_prob_calibrated" in out_df.columns:
        mask = out_df["market_type"] == "total"
        if mask.any():
            out_df.loc[mask, "total_prob_calibrated"] = out_df.loc[mask].apply(
                lambda row: _apply_market_calibrator(
                    total_cals,
                    row.get("total_source"),
                    _total_raw_probability(row),
                ),
                axis=1,
            )
            total_rows = int(mask.sum())
            total_calibrated = int(out_df.loc[mask, "total_prob_calibrated"].notna().sum())
            total_calibrators_label = ", ".join(sorted(total_cals.keys()))
            _LOG.info(
                f"[_apply_spread_total_calibrators] Applied total calibrator(s) {total_calibrators_label} to {total_calibrated}/{total_rows} rows"
            )

    return out_df


def _apply_market_calibrator(
    calibrators: dict[str, object],
    source: str | None,
    raw_prob: float | None,
) -> CalibratedProbability:
    """
    DEPRECATED: This function is no longer called as of 2025-01-27.

    This helper was designed to apply a probability calibrator to a single raw probability.
    It assumes the calibrator has a .transform(pd.Series) method that accepts a Series of
    probabilities and returns calibrated probabilities.

    Interface mismatch issue:
     - VarianceCalibrator.transform() expects DataFrame with (pred_mean, pred_sd)
    - This function calls .transform(pd.Series([raw_prob])) which is incompatible

    The silent try/except block at line 1733 catches the interface mismatch, causing
    all calibration attempts to return pd.NA instead of calibrated values.

    If probability calibration is needed in the future, use ProbabilityFromDistributionCalibrator
    or implement proper interface detection to handle different calibrator types.
    """
    if not source or raw_prob is None:
        return pd.NA
    calibrator = calibrators.get(source)
    if calibrator is None:
        return pd.NA
    try:
        transformed = calibrator.transform(pd.Series([raw_prob]))
    except Exception:
        return pd.NA
    if transformed.empty:
        return pd.NA
    val = transformed.iloc[0]
    if pd.isna(val):
        return pd.NA
    return float(val)


def _load_market_calibrators(
    sources: Iterable[str],
    sport: str | None,
    season: str | None,
    market: Market,
) -> dict[str, object]:
    """
    DEPRECATED: This function is only called by _apply_spread_total_calibrators,
    which is no longer used as of 2025-01-27.

    This helper loads calibrators for multiple source models for a given market.
    It was designed to support per-source calibration in the BETS sheet, where
    different rows might use different model sources.

     The loaded calibrators are VarianceCalibrator objects (for SPREAD/TOTAL),
    which have a different interface than what _apply_market_calibrator expects.

    For proper calibration, use _apply_calibration_to_schedule_df() which handles
    the calibrator interfaces correctly.
    """
    cals: dict[str, object] = {}
    if not sport or not season:
        return cals
    for source in sources:
        if not source or source in cals:
            continue
        try:
            calibrator = load_latest_calibrator(
                sport=sport,
                season=season,
                model=source,
                market=market,
            )
        except Exception:
            calibrator = None
        if calibrator is not None:
            cals[source] = calibrator
    return cals


def _log_bets_probability(
    *,
    market: str,
    selection: str | None,
    line: float | None,
    mean: float | None,
    sd_used: float | None,
    raw_sd: float | None,
    raw_p: float | None,
    clamped_p: float | None,
) -> None:
    if not _LOG.isEnabledFor(logging.DEBUG):
        return

    def _z_for(sd_value: float | None) -> float | None:
        if line is None or mean is None or sd_value is None or sd_value == 0.0:
            return None
        return (line - mean) / sd_value

    z_used = _z_for(sd_used)
    raw_z = _z_for(raw_sd)
    _LOG.debug(
        "[bets-prob] market=%s selection=%s line=%s mean=%s sd_used=%s z=%s raw_sd=%s raw_z=%s raw_p=%s clamped_p=%s",
        market,
        selection or "",
        line,
        mean,
        sd_used,
        z_used,
        raw_sd,
        raw_z,
        raw_p,
        clamped_p,
    )


def _spread_probability_from_sd(
    line: float | None,
    mean: float | None,
    sd_value: float | None,
    selection: str,
    home_team: str,
    away_team: str,
) -> float | None:
    if line is None or mean is None or sd_value is None or sd_value <= 0:
        return None
    sel = selection.strip()
    if sel and home_team and sel == home_team:
        return cover_prob(line, mean, sd_value, sign_convention="away_minus_home")
    if sel and away_team and sel == away_team:
        base = cover_prob(-line, mean, sd_value, sign_convention="away_minus_home")
        if base is None:
            return None
        return 1.0 - base
    return None


def _spread_raw_probability(row: Mapping[str, Any]) -> float | None:
    line = _coerce_float(row.get("line"))
    mean = _coerce_float(row.get("margin_mean"))
    sd = _coerce_float(row.get("margin_sd"))
    raw_sd = _coerce_float(row.get("margin_sd_pre_guardrail"))
    if line is None or mean is None or sd is None:
        return None
    selection = str(row.get("selection") or "").strip()
    home_team = str(row.get("home_team") or "").strip()
    away_team = str(row.get("away_team") or "").strip()
    clamped_p = _spread_probability_from_sd(line, mean, sd, selection, home_team, away_team)
    raw_p = _spread_probability_from_sd(line, mean, raw_sd, selection, home_team, away_team)
    _log_bets_probability(
        market="spread",
        selection=selection,
        line=line,
        mean=mean,
        sd_used=sd,
        raw_sd=raw_sd,
        raw_p=raw_p,
        clamped_p=clamped_p,
    )
    return clamped_p


def _total_probability_from_sd(
    line: float | None,
    mean: float | None,
    sd_value: float | None,
    selection: str | None,
) -> float | None:
    if line is None or mean is None or sd_value is None or sd_value <= 0:
        return None
    sel = str(selection or "").strip().lower()
    over_p = over_prob(line, mean, sd_value)
    if over_p is None:
        return None
    if sel == "over":
        return over_p
    if sel == "under":
        return 1.0 - over_p
    return None


def _total_raw_probability(row: Mapping[str, Any]) -> float | None:
    line = _coerce_float(row.get("line"))
    mean = _coerce_float(row.get("total"))
    sd = _coerce_float(row.get("total_sd"))
    raw_sd = _coerce_float(row.get("total_sd_pre_guardrail"))
    if line is None or mean is None or sd is None:
        return None
    selection = str(row.get("selection") or "").strip().lower()
    if selection not in {"over", "under"}:
        return None
    raw_p = _total_probability_from_sd(line, mean, raw_sd, selection)
    clamped_p = _total_probability_from_sd(line, mean, sd, selection)
    _log_bets_probability(
        market="total",
        selection=selection,
        line=line,
        mean=mean,
        sd_used=sd,
        raw_sd=raw_sd,
        raw_p=raw_p,
        clamped_p=clamped_p,
    )
    return clamped_p


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        val = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(val) or math.isinf(val):
        return None
    return val


def _training_date_range(df: pd.DataFrame) -> str:
    if df.empty or "date" not in df.columns:
        return ""
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if dates.empty:
        return ""
    return f"{dates.min().date().isoformat()} to {dates.max().date().isoformat()}"


def _serialize_params(params: Any) -> str:
    try:
        return json.dumps(params, sort_keys=True, default=str)
    except TypeError:
        return str(params)


def _build_model_metadata(
    *,
    model_name: str,
    played: pd.DataFrame,
    schedule_df: pd.DataFrame,
    params_source: str,
    tuned_metric_used: str | None,
    params_source_label: str | None,
    params_source_run_id: str | None,
    params_metric_optimized: str | None,
    params_best_score: float | None,
    params_fingerprint: str | None,
    params_nonempty: bool | None,
    params_run_id: str | None,
    params_market: str | None = None,
    health_status: str | None = None,
    health_details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    model_instance = get_model(model_name)()
    identity = resolve_model_identity(model_instance)
    metadata = {
        "model_id": identity["model_id"],
        "model_version": identity["model_version"],
        "prediction_hash": prediction_hash(schedule_df, SCHEDULE_EXPORT_COLUMNS),
        "params": _serialize_params(identity["params"]),
        "params_source": params_source,
        "params_source_label": params_source_label or params_source,
        "params_source_run_id": params_source_run_id or params_run_id,
        "tuned_metric_used": tuned_metric_used,
        "params_metric_optimized": params_metric_optimized,
        "params_best_score": params_best_score,
        "params_fingerprint": params_fingerprint,
        "params_nonempty": params_nonempty,
        "tuning_run_id": params_run_id,
        "params_market": params_market or Market.ML.name,
        "trained_on_date_range": _training_date_range(played),
        "n_games_train": int(len(played)),
        "run_timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "health_status": health_status or "UNKNOWN",
        "health_details_json": json.dumps(health_details) if health_details else None,
    }
    return metadata


def _write_metadata_section(
    writer: pd.ExcelWriter,
    sheet_name: str,
    metadata: dict[str, Any],
) -> int:
    """Write model metadata above the schedule sheet and return the data start row."""
    metadata_df = pd.DataFrame(
        [{"metadata_key": key, "metadata_value": value} for key, value in metadata.items()]
    )
    metadata_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    # Always start schedule data at a fixed row below the metadata block.
    return MODEL_METADATA_DATA_START_ROW - 1


def _build_schedule_for_model(
    df: pd.DataFrame,
    *,
    db_path: str | Path,
    sport: str,
    season: str,
    model: str,
    output_path: str | Path | None,
    upcoming_only: bool,
    add_prefix: bool,
    model_params: dict[str, float] | None,
    model_params_file: str | Path | None,
    tuned_metric: str | None,
) -> Path:
    resolution = resolve_model_market_params_with_metadata(
        model,
        params=model_params,
        params_file=model_params_file,
        db_path=db_path,
        sport=sport,
        season=season,
        tuned_metric=tuned_metric,
        market=Market.ML,
    )
    resolved_params = resolution.params
    params_run_id = resolution.source_run_id
    params_source_label = resolution.params_source_label or resolution.params_source
    params_source_run_id = resolution.source_run_id
    params_metric_optimized = resolution.metric_optimized
    params_best_score = resolution.best_score
    params_fingerprint = resolution.params_fingerprint
    params_nonempty = resolution.params_nonempty
    schedule_df = _build_schedule_dataframe(
        df,
        db_path=db_path,
        sport=sport,
        season=season,
        model=model,
        upcoming_only=upcoming_only,
        model_params=resolved_params,
        params_source=resolution.params_source,
        params_source_label=params_source_label,
        params_source_run_id=params_source_run_id,
        tuned_metric_used=resolution.tuned_metric_used,
        params_metric_optimized=params_metric_optimized,
        params_best_score=params_best_score,
        params_fingerprint=params_fingerprint,
        params_nonempty=params_nonempty,
        params_run_id=params_run_id,
        params_market=resolution.market,
    )
    schedule_df = _order_schedule_export(schedule_df)
    default_path = processed_path_for(sport, season, "schedule_with_projections.csv")
    resolved_output = resolve_output_path(
        output_path,
        default_path=default_path,
        model=model,
        add_prefix=add_prefix,
    )
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    schedule_df.to_csv(resolved_output, index=False)
    return resolved_output


def build_schedule_with_projections(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    division: str | None = None,
    conference: str | None = None,
    model: str | None = None,
    output_path: str | Path | None = None,
    upcoming_only: bool = False,
    model_params: dict[str, float] | None = None,
    model_params_file: str | Path | None = None,
    tuned_metric: str | None = None,
    mode: str = "dev",
) -> Path | list[Path]:
    """Build a schedule export containing projections for upcoming games."""
    rows = load_games(
        db_path,
        sport=sport,
        season=season,
        division=division,
        conference=conference,
    )
    df = normalize_games(rows)
    if df.empty:
        raise ValueError(f"No games found for sport={sport!r}, season={season!r}")

    models = _resolve_models(model)
    multiple = len(models) > 1
    results = [
        _build_schedule_for_model(
            df.copy(deep=True),
            db_path=db_path,
            sport=sport,
            season=season,
            model=model_name,
            output_path=output_path,
            upcoming_only=upcoming_only,
            add_prefix=multiple,
            model_params=model_params,
            model_params_file=model_params_file,
            tuned_metric=tuned_metric,
        )
        for model_name in models
    ]
    return results[0] if len(results) == 1 else results


def build_schedule_excel_report(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    division: str | None = None,
    conference: str | None = None,
    model: str | None = None,
    output_path: str | Path | None = None,
    upcoming_only: bool = False,
    model_params: dict[str, float] | None = None,
    model_params_file: str | Path | None = None,
    tuned_metric: str | None = None,
    as_of_date: str | date | None = None,
    bets_model: str | None = None,
    strict: bool = False,
    fail_on_health_check: bool = False,
    mode: str = "dev",
    allow_best_fallback: bool = False,
    apply_total_recency_adjustment: bool = False,
    calibration_mode: str = "off",
) -> Path:
    """Build an Excel workbook with schedule projections (one sheet per model)."""
    rows = load_games(
        db_path,
        sport=sport,
        season=season,
        division=division,
        conference=conference,
    )
    df = normalize_games(rows)
    if df.empty:
        raise ValueError(f"No games found for sport={sport!r}, season={season!r}")

    models = _resolve_models(model)
    bets_model_name = _resolve_bets_model(models, bets_model)
    resolved_as_of_date = _resolve_as_of_date(as_of_date)
    display_date = resolved_as_of_date if as_of_date is not None else None
    
    # Keep full schedule for model training - filtering happens later for display
    # _build_market_forecasts_for_ensembles needs all games to build ratings
    all_games_df = df.copy(deep=True)

    fit_end_date = resolved_as_of_date if as_of_date is not None else None
    report_path = _resolve_workbook_path(
        output_path,
        sport=sport,
        season=season,
        default_name="schedule_with_projections.xlsx",
    )
    dashboard_rows: list[Dict[str, Any]] = []
    bets_schedule_df: pd.DataFrame | None = None

    ensemble_config = load_ensemble_config(
        sport,
        season,
        available_models=list_models(),
    )
    ensemble_config_meta = (ensemble_config or {}).get("_meta", {})
    market_configs = (ensemble_config or {}).get("markets", {})
    market_config_meta = ensemble_config_meta.get("markets", {})
    config_warnings: list[str] = list(ensemble_config_meta.get("warnings", []) or [])

    def _market_config_value(market_name: str, field: str, default: Any = None) -> Any:
        cfg = market_configs.get(market_name) or {}
        return cfg.get(field, default)

    ensemble_ids = {
        Market.ML.name: _market_config_value(Market.ML.name, "ensemble_id", "ensemble_ml_v1"),
        Market.SPREAD.name: _market_config_value(
            Market.SPREAD.name, "ensemble_id", "ensemble_spread_v1"
        ),
        Market.TOTAL.name: _market_config_value(
            Market.TOTAL.name, "ensemble_id", "ensemble_total_v1"
        ),
    }
    selection_contexts = {
        Market.ML.name: _load_active_selection_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.ML.name, ensemble_id=ensemble_ids[Market.ML.name],
        ),
        Market.SPREAD.name: _load_active_selection_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.SPREAD.name, ensemble_id=ensemble_ids[Market.SPREAD.name],
        ),
        Market.TOTAL.name: _load_active_selection_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.TOTAL.name, ensemble_id=ensemble_ids[Market.TOTAL.name],
        ),
    }
    tuning_contexts = {
        Market.ML.name: _load_active_tuning_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.ML.name, ensemble_id=ensemble_ids[Market.ML.name],
        ),
        Market.SPREAD.name: _load_active_tuning_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.SPREAD.name, ensemble_id=ensemble_ids[Market.SPREAD.name],
        ),
        Market.TOTAL.name: _load_active_tuning_context(
            db_path=db_path, sport=sport, season=season,
            market=Market.TOTAL.name, ensemble_id=ensemble_ids[Market.TOTAL.name],
        ),
    }
    allowed_models_map = {}
    for market_name, selection in selection_contexts.items():
        if selection:
            allowed_models_map[market_name] = selection["models"]
        else:
            allowed_models_map[market_name] = _market_config_value(market_name, "models")
    market_metrics = {
        Market.ML.name: _market_config_value(Market.ML.name, "metric_slot"),
        Market.SPREAD.name: _market_config_value(Market.SPREAD.name, "metric_slot"),
        Market.TOTAL.name: _market_config_value(Market.TOTAL.name, "metric_slot"),
    }
    config_weights_map = {
        Market.ML.name: _market_config_value(Market.ML.name, "weights"),
        Market.SPREAD.name: _market_config_value(Market.SPREAD.name, "weights"),
        Market.TOTAL.name: _market_config_value(Market.TOTAL.name, "weights"),
    }
    resolved_ensemble_meta: dict[str, dict[str, Any]] = {}

    with pd.ExcelWriter(report_path) as writer:
        for model_name in models:
            model_df = all_games_df.copy(deep=True)
            market_frames: list[pd.DataFrame] = []
            market_metadatas: list[tuple[Market, dict[str, Any]]] = []
            params_sources: set[str] = set()
            tuned_metrics: set[str] = set()
            params_run_ids: set[str] = set()

            for market in EXPORT_MARKETS:
                resolution = resolve_model_market_params_with_metadata(
                    model_name,
                    params=model_params,
                    params_file=model_params_file,
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    tuned_metric=tuned_metric,
                    market=market,
                )
                resolved_params = resolution.params
                params_source = resolution.params_source
                tuned_metric_used = resolution.tuned_metric_used
                params_source_label = resolution.params_source_label or params_source
                params_source_run_id = resolution.source_run_id
                params_metric_optimized = resolution.metric_optimized
                params_best_score = resolution.best_score
                params_fingerprint = resolution.params_fingerprint
                params_nonempty = resolution.params_nonempty
                schedule_df = _build_schedule_dataframe(
                    model_df,
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    model=model_name,
                    upcoming_only=upcoming_only,
                    model_params=resolved_params,
                    params_source=params_source,
                    params_source_label=params_source_label,
                    params_source_run_id=params_source_run_id,
                    tuned_metric_used=tuned_metric_used,
                    params_metric_optimized=params_metric_optimized,
                    params_best_score=params_best_score,
                    params_fingerprint=params_fingerprint,
                    params_nonempty=params_nonempty,
                    params_run_id=resolution.source_run_id,
                    params_market=market.name,
                    fit_end_date=fit_end_date,
                    fail_on_health_check=fail_on_health_check,
                    mode=mode,
                )
                schedule_df = _order_schedule_export(schedule_df)
                market_frames.append(schedule_df)

                # Extract health status for metadata
                health_status = schedule_df.attrs.get("_health_status")
                health_details = schedule_df.attrs.get("_health_details")

                played_for_metadata = _filter_games_through(
                    _completed_games(model_df), fit_end_date
                )
                metadata = _build_model_metadata(
                    model_name=model_name,
                    played=played_for_metadata,
                    schedule_df=schedule_df,
                    params_source=params_source,
                    params_source_label=params_source_label,
                    params_source_run_id=params_source_run_id,
                    tuned_metric_used=tuned_metric_used,
                    params_metric_optimized=params_metric_optimized,
                    params_best_score=params_best_score,
                    health_status=health_status,
                    health_details=health_details,
                    params_fingerprint=params_fingerprint,
                    params_nonempty=params_nonempty,
                    params_run_id=resolution.source_run_id,
                    params_market=market.name,
                )
                market_metadatas.append((market, metadata))
                if params_source_label:
                    params_sources.add(str(params_source_label))
                if tuned_metric_used:
                    tuned_metrics.add(str(tuned_metric_used))
                if params_source_run_id:
                    params_run_ids.add(str(params_source_run_id))

                dashboard_rows.extend(
                    _dashboard_rows_for_today(
                        schedule_df,
                        model_name,
                        resolved_as_of_date,
                        model_metadata=metadata,
                    )
                )
                # Set bets_schedule_df from the first model's ML market schedule
                # This ensures it has all required columns and isn't None for ensemble logic
                if bets_schedule_df is None and market == Market.ML:
                    bets_schedule_df = schedule_df

            # Filter empty DataFrames before concat to avoid FutureWarning
            non_empty_frames = [df for df in market_frames if not df.empty]
            if non_empty_frames:
                # Suppress FutureWarning for DataFrame concat with all-NA columns
                import warnings
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=FutureWarning, message=".*DataFrame concatenation.*")
                    model_df_all_markets = pd.concat(non_empty_frames, ignore_index=True)
            else:
                model_df_all_markets = pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)
            if not model_df_all_markets.empty:
                model_df_all_markets = (
                    model_df_all_markets.assign(
                        _sort_dt=pd.to_datetime(model_df_all_markets["date"], errors="coerce"),
                        _market_order=model_df_all_markets["params_market"].map(MARKET_ORDER).fillna(len(MARKET_ORDER)),
                    )
                    .sort_values(
                        ["_sort_dt", "game_id", "_market_order", "away_team", "home_team"]
                    )
                    .drop(columns=["_sort_dt", "_market_order"], errors="ignore")
                )

            params_source_joined = ",".join(sorted(params_sources)) if params_sources else "multi_market"
            params_source_labels = {md.get("params_source_label") for _, md in market_metadatas if md.get("params_source_label")}
            params_metric_opts = {md.get("params_metric_optimized") for _, md in market_metadatas if md.get("params_metric_optimized")}
            params_fingerprints = {md.get("params_fingerprint") for _, md in market_metadatas if md.get("params_fingerprint")}
            params_nonempty_flags = {str(md.get("params_nonempty")) for _, md in market_metadatas if md.get("params_nonempty") is not None}

            combined_played = _filter_games_through(
                _completed_games(model_df), fit_end_date
            )
            combined_metadata = _build_model_metadata(
                model_name=model_name,
                played=combined_played,
                schedule_df=model_df_all_markets,
                params_source=params_source_joined,
                params_source_label=",".join(sorted(params_source_labels)) if params_source_labels else params_source_joined,
                params_source_run_id=",".join(sorted(params_run_ids)) if params_run_ids else None,
                tuned_metric_used=",".join(sorted(tuned_metrics)) if tuned_metrics else None,
                params_metric_optimized=",".join(sorted(params_metric_opts)) if params_metric_opts else None,
                params_best_score=None,
                params_fingerprint=",".join(sorted(params_fingerprints)) if params_fingerprints else None,
                params_nonempty=",".join(sorted(params_nonempty_flags)) if params_nonempty_flags else None,
                params_run_id=",".join(sorted(params_run_ids)) if params_run_ids else None,
                params_market="MULTI",
            )
            combined_metadata["export_markets"] = ",".join(market.name for market in EXPORT_MARKETS)
            combined_metadata[
                "note"
            ] = "Multi-market schedule; rows are duplicated per market. See params_market for grouping and ordering."
            for market, metadata in market_metadatas:
                for key in (
                    "params_market",
                    "params_source",
                    "params_source_label",
                    "params_source_run_id",
                    "tuned_metric_used",
                    "params_metric_optimized",
                    "params_best_score",
                    "params_fingerprint",
                    "params_nonempty",
                    "tuning_run_id",
                    "run_timestamp_utc",
                    "prediction_hash",
                ):
                    combined_metadata[f"{market.name}.{key}"] = metadata.get(key)

            start_row = _write_metadata_section(writer, model_name, combined_metadata)
            model_df_all_markets.to_excel(
                writer,
                sheet_name=model_name,
                index=False,
                startrow=start_row,
            )

        dashboard_df = pd.DataFrame(dashboard_rows)
        if not dashboard_df.empty:
            model_order = {name: idx for idx, name in enumerate(models)}
            sort_dt = (
                dashboard_df["_sort_dt"]
                if "_sort_dt" in dashboard_df.columns
                else pd.to_datetime(dashboard_df["date"], errors="coerce")
            )
            sort_game_id = (
                dashboard_df["_sort_game_id"]
                if "_sort_game_id" in dashboard_df.columns
                else pd.NA
            )
            dashboard_df = dashboard_df.assign(
                _model_order=dashboard_df["model"]
                .map(model_order)
                .fillna(len(model_order)),
                _sort_dt=sort_dt,
                _sort_game_id=sort_game_id,
                _market_order=dashboard_df["params_market"].map(MARKET_ORDER).fillna(len(MARKET_ORDER)),
            ).sort_values(
                ["_sort_dt", "_sort_game_id", "_market_order", "_model_order", "model"]
            )
            dashboard_df = dashboard_df.drop(
                columns=["_model_order", "_sort_dt", "_sort_game_id", "_market_order"],
                errors="ignore",
            )
        dashboard_df = dashboard_df.reindex(columns=DASHBOARD_COLUMNS)
        dashboard_df.to_excel(writer, sheet_name="dashboard", index=False)

        # Build market forecast rows for any BETS sheet generation so that ML/SPREAD/TOTAL
        # data is always available, even with a single model. Ensemble logic still only
        # applies when len(models) > 1, but single-model TOTAL data can be passed through.
        ensemble_applied = False
        market_forecast_rows: dict[str, list[dict[str, Any]]] = {
            Market.ML.name: [],
            Market.SPREAD.name: [],
            Market.TOTAL.name: [],
        }
        if bets_schedule_df is not None:
            if len(models) > 1:
                _validate_market_tuning_inputs(
                    db_path=db_path,
                    sport=sport,
                    season=season,
                    models=models,
                    ensemble_ids=ensemble_ids,
                    strict=strict,
                )
            market_forecast_rows = _build_market_forecasts_for_ensembles(
                all_games_df,  # Pass full schedule for rating computation
                db_path=db_path,
                sport=sport,
                season=season,
                models=models,
                as_of_date=resolved_as_of_date,
                allowed_models_by_market=allowed_models_map if len(models) > 1 else None,
                market_metrics=market_metrics,
                fit_end_date=fit_end_date,
                mode=mode,
            )

            # Ensure all required ensemble columns exist in bets_schedule_df before ensemble writes.
            # This fixes the issue where ensemble code tries to write to columns that don't exist
            # because bets_schedule_df was built from ML-market schedule only.
            if bets_schedule_df is not None:
                ensemble_columns = [
                    "spread_source",
                    "spread_ensemble_components_json",
                    "total_source",
                    "total_ensemble_components_json",
                    "total",
                    "total_sd",
                    "ml_ensemble_components_json",
                ]
                for col in ensemble_columns:
                    if col not in bets_schedule_df.columns:
                        bets_schedule_df[col] = pd.NA

        # -------------------------------------------------------------------------
        # ML ENSEMBLE APPLICATION: Apply ML ensemble to BETS schedule
        # Logs weight resolution, filtering, and per-game application stats
        # -------------------------------------------------------------------------
        try:
            ml_rows = market_forecast_rows.get(Market.ML.name, [])
            _LOG.debug(
                "[ML ensemble] Starting: forecast_rows=%d, bets_schedule_rows=%d",
                len(ml_rows),
                len(bets_schedule_df) if bets_schedule_df is not None else 0,
            )
            if bets_schedule_df is not None and ml_rows:
                forecast_df = pd.DataFrame(ml_rows)
                if not forecast_df.empty:
                    weights, ensemble_models, weight_source, weight_run_id, selection_run_id = _resolve_ensemble_weights(
                        db_path=db_path,
                        sport=sport,
                        season=season,
                        market=Market.ML.name,
                        ensemble_id=ensemble_ids[Market.ML.name],
                        config_weights=config_weights_map.get(Market.ML.name),
                        selection_context=selection_contexts[Market.ML.name],
                        tuning_context=tuning_contexts[Market.ML.name],
                        config_warnings=config_warnings,
                    )

                    # Log weight resolution results
                    _LOG.info(
                        "[ML ensemble] Weights resolved: source=%s, models=%s, weights=%s, run_id=%s",
                        weight_source,
                        ensemble_models,
                        weights,
                        weight_run_id,
                    )

                    # Filter forecast data to only include models the ensemble was trained on
                    if ensemble_models:
                        forecast_df = forecast_df[forecast_df["model_name"].isin(set(ensemble_models))]
                        if forecast_df.empty:
                            config_warnings.append(
                                f"No ML forecasts matched ensemble models {ensemble_models}; ensemble skipped"
                            )
                            raise Exception("No ML forecasts after ensemble model filter")
                    filtered_weights, final_models = _filter_market_weights_for_forecast(
                        weights=weights,
                        forecast_df=forecast_df,
                        market=Market.ML.name,
                    )
                    if not final_models:
                        config_warnings.append(
                            "No ML forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No ML forecasts after weight filtering")
                    forecast_df = forecast_df[forecast_df["model_name"].isin(final_models)]
                    if forecast_df.empty:
                        config_warnings.append(
                            "No ML forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No ML forecasts after weight filtering")
                    weights = filtered_weights

                    # Log post-filtering state
                    unique_ml_models = set(forecast_df["model_name"].dropna().unique())
                    use_ensemble = len(unique_ml_models) > 1
                    _LOG.info(
                        "[ML ensemble] Post-filter: final_models=%s, use_ensemble=%s, filtered_weights=%s",
                        sorted(final_models) if final_models else [],
                        use_ensemble,
                        filtered_weights,
                    )

                    # Counter for tracking games updated
                    _ml_games_updated = 0

                    if use_ensemble:
                        ensemble = MLWeightedAverageEnsemble(
                            sport,
                            season,
                            ensemble_id=ensemble_ids[Market.ML.name],
                            weights=weights,
                        )
                        used_models = sorted(set(forecast_df.get("model_name", [])))
                        cfg_meta = market_config_meta.get(Market.ML.name, {}) if isinstance(market_config_meta, dict) else {}
                        resolved_ensemble_meta[Market.ML.name] = {
                            "ensemble_id": ensemble.ensemble_id,
                            "metric_slot": market_metrics.get(Market.ML.name),
                            "configured_models": ensemble_models or used_models,
                            "configured_weights": config_weights_map.get(Market.ML.name),
                            "weights_source": weight_source or "equal",
                            "source_run_id": weight_run_id,
                            "weights": weights,
                            "used_models": used_models,
                            "config_source": cfg_meta.get("source"),
                            "config_path": cfg_meta.get("path"),
                            "selection_run_id": selection_run_id,
                            "selection_models": (
                                selection_contexts[Market.ML.name]["models"]
                                if selection_contexts[Market.ML.name]
                                else None
                            ),
                            "tuning_run_id": weight_run_id if weight_source == "db_tuned" else None,
                        }
                        # Load calibrator for the ensemble source, if present
                        try:
                            ensemble_cal = load_latest_calibrator(
                                sport=sport,
                                season=season,
                                model=ensemble.ensemble_id,
                                market=Market.ML,
                            )
                        except Exception:
                            ensemble_cal = None

                        # Apply per-game ensemble
                        for gid in pd.unique(bets_schedule_df["game_id"]):
                            try:
                                subset = forecast_df[forecast_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                raw_p, components_json = ensemble.combine(subset)
                                if raw_p is None:
                                    continue
                                final_p = raw_p
                                if ensemble_cal is not None:
                                    try:
                                        transformed = ensemble_cal.transform([raw_p])
                                        if transformed is not None and len(transformed) > 0:
                                            final_p = float(transformed[0])
                                    except Exception:
                                        pass

                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[mask, "home_win_prob"] = final_p
                                bets_schedule_df.loc[mask, "away_win_prob"] = 1.0 - final_p

                                # Set win_prob_source to the ensemble ID (replacing any default "direct" value)
                                bets_schedule_df.loc[mask, "win_prob_source"] = ensemble.ensemble_id
                                bets_schedule_df.loc[mask, "ml_ensemble_components_json"] = components_json
                                ensemble_applied = True
                                _ml_games_updated += 1
                            except Exception:
                                # Best-effort per-game; continue on errors.
                                continue

                        # Log ML ensemble success
                        _LOG.info(
                            "[ML ensemble] SUCCESS: ensemble_id=%s, games_updated=%d, calibrator_loaded=%s",
                            ensemble.ensemble_id,
                            _ml_games_updated,
                            ensemble_cal is not None,
                        )
                    else:
                        # Single-model pass-through (no ensemble blending needed)
                        single_model_name = (
                            list(unique_ml_models)[0] if unique_ml_models else bets_model_name
                        )
                        _LOG.info(
                            "[ML ensemble] Single-model pass-through: model=%s",
                            single_model_name,
                        )
                        _ml_single_games_updated = 0
                        for gid in pd.unique(bets_schedule_df["game_id"]):
                            try:
                                mask = bets_schedule_df["game_id"] == gid
                                if not mask.any():
                                    continue
                                p_home = bets_schedule_df.loc[mask, "home_win_prob"].iloc[0]
                                if _is_missing(p_home):
                                    continue
                                components = {
                                    "models": [single_model_name],
                                    "p_home_win": float(p_home),
                                }
                                bets_schedule_df.loc[mask, "ml_ensemble_components_json"] = json.dumps(
                                    components
                                )
                                _ml_single_games_updated += 1
                            except Exception:
                                continue
                        _LOG.info(
                            "[ML ensemble] Single-model complete: games_updated=%d",
                            _ml_single_games_updated,
                        )
        except Exception as e:
            # Do not fail report generation for ensemble errors.
            _LOG.warning(f"ML ensemble application failed: {e}", exc_info=True)
            pass

        ml_components_missing = (
            bets_schedule_df is not None
            and (
                "ml_ensemble_components_json" not in bets_schedule_df.columns
                or not bets_schedule_df["ml_ensemble_components_json"].notna().any()
            )
        )
        fallback_applied = False
        if ml_components_missing and not ensemble_applied:
            try:
                fallback_rows = market_forecast_rows.get(Market.ML.name, [])
                if fallback_rows:
                    fallback_df = pd.DataFrame(fallback_rows)
                    if not fallback_df.empty:
                        allowed_models = allowed_models_map.get(Market.ML.name)
                        if allowed_models:
                            fallback_df = fallback_df[
                                fallback_df["model_name"].isin(set(allowed_models))
                            ]
                        if not fallback_df.empty:
                            weights, _, _, _, _ = _resolve_ensemble_weights(
                                db_path=db_path,
                                sport=sport,
                                season=season,
                                market=Market.ML.name,
                                ensemble_id=ensemble_ids[Market.ML.name],
                                config_weights=config_weights_map.get(Market.ML.name),
                                selection_context=selection_contexts[Market.ML.name],
                                tuning_context=tuning_contexts[Market.ML.name],
                                config_warnings=config_warnings,
                            )
                            ensemble = MLWeightedAverageEnsemble(
                                sport,
                                season,
                                ensemble_id=ensemble_ids[Market.ML.name],
                                weights=weights,
                            )
                            for gid in pd.unique(fallback_df["game_id"]):
                                subset = fallback_df[fallback_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                _, components_json = ensemble.combine(subset)
                                if components_json is None:
                                    continue
                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[
                                    mask, "ml_ensemble_components_json"
                                ] = components_json
                            fallback_applied = True
            except Exception as exc:  # pragma: no cover - best-effort fallback
                _LOG.warning(
                    "ML ensemble fallback failed: %s", exc, exc_info=False
                )
        if ml_components_missing and not fallback_applied and not ensemble_applied:
            _LOG.warning(
                "ML ensemble fallback ran without an applied ensemble; "
                "ml_ensemble_components_json remains blank."
            )

        # -------------------------------------------------------------------------
        # SPREAD ENSEMBLE APPLICATION: Apply SPREAD ensemble to BETS schedule
        # For multi-model, use ensemble; for single model, pass through directly.
        # Logs weight resolution, filtering, and per-game application stats.
        # -------------------------------------------------------------------------
        spread_ensemble_applied = False
        try:
            spread_rows = market_forecast_rows.get(Market.SPREAD.name, [])
            _LOG.debug(
                "[SPREAD ensemble] Starting: forecast_rows=%d, bets_schedule_rows=%d",
                len(spread_rows),
                len(bets_schedule_df) if bets_schedule_df is not None else 0,
            )
            if bets_schedule_df is not None and spread_rows:
                forecast_df = pd.DataFrame(spread_rows)
                if not forecast_df.empty:
                    weights, ensemble_models, weight_source, weight_run_id, selection_run_id = _resolve_ensemble_weights(
                        db_path=db_path,
                        sport=sport,
                        season=season,
                        market=Market.SPREAD.name,
                        ensemble_id=ensemble_ids[Market.SPREAD.name],
                        config_weights=config_weights_map.get(Market.SPREAD.name),
                        selection_context=selection_contexts[Market.SPREAD.name],
                        tuning_context=tuning_contexts[Market.SPREAD.name],
                        config_warnings=config_warnings,
                    )

                    # Log weight resolution results
                    _LOG.info(
                        "[SPREAD ensemble] Weights resolved: source=%s, models=%s, weights=%s, run_id=%s",
                        weight_source,
                        ensemble_models,
                        weights,
                        weight_run_id,
                    )

                    # Filter forecast data to only include models the ensemble was trained on
                    if ensemble_models:
                        forecast_df = forecast_df[forecast_df["model_name"].isin(set(ensemble_models))]
                        if forecast_df.empty:
                            config_warnings.append(
                                f"No SPREAD forecasts matched ensemble models {ensemble_models}; ensemble skipped"
                            )
                            raise Exception("No SPREAD forecasts after ensemble model filter")
                    filtered_weights, final_models = _filter_market_weights_for_forecast(
                        weights=weights,
                        forecast_df=forecast_df,
                        market=Market.SPREAD.name,
                    )
                    if not final_models:
                        config_warnings.append(
                            "No SPREAD forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No SPREAD forecasts after weight filtering")
                    forecast_df = forecast_df[forecast_df["model_name"].isin(final_models)]
                    if forecast_df.empty:
                        config_warnings.append(
                            "No SPREAD forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No SPREAD forecasts after weight filtering")
                    weights = filtered_weights

                    # Log post-filtering state
                    unique_spread_models = set(forecast_df["model_name"].dropna().unique())
                    use_ensemble = len(unique_spread_models) > 1
                    _LOG.info(
                        "[SPREAD ensemble] Post-filter: final_models=%s, use_ensemble=%s, filtered_weights=%s",
                        sorted(final_models) if final_models else [],
                        use_ensemble,
                        filtered_weights,
                    )

                    if use_ensemble:
                        # Multi-model: use weighted ensemble
                        spread_ensemble = SpreadWeightedAverageEnsemble(
                            sport,
                            season,
                            ensemble_id=ensemble_ids[Market.SPREAD.name],
                            weights=weights,
                        )
                        used_models = sorted(unique_spread_models)
                        cfg_meta = market_config_meta.get(Market.SPREAD.name, {}) if isinstance(market_config_meta, dict) else {}
                        resolved_ensemble_meta[Market.SPREAD.name] = {
                            "ensemble_id": spread_ensemble.ensemble_id,
                            "metric_slot": market_metrics.get(Market.SPREAD.name),
                            "configured_models": ensemble_models or used_models,
                            "configured_weights": config_weights_map.get(Market.SPREAD.name),
                            "weights_source": weight_source or "equal",
                            "source_run_id": weight_run_id,
                            "weights": weights,
                            "used_models": used_models,
                            "config_source": cfg_meta.get("source"),
                            "config_path": cfg_meta.get("path"),
                            "selection_run_id": selection_run_id,
                            "selection_models": (
                                selection_contexts[Market.SPREAD.name]["models"]
                                if selection_contexts[Market.SPREAD.name]
                                else None
                            ),
                            "tuning_run_id": weight_run_id if weight_source == "db_tuned" else None,
                        }
                        spread_games_updated = 0
                        bets_game_ids = pd.unique(bets_schedule_df["game_id"])
                        for gid in bets_game_ids:
                            try:
                                subset = forecast_df[forecast_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                margin_mean_raw, margin_sd_raw, components_json = (
                                    spread_ensemble.combine(subset)
                                )
                                if margin_mean_raw is None:
                                    continue
                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[mask, "margin_mean"] = margin_mean_raw
                                if margin_sd_raw is not None:
                                    bets_schedule_df.loc[mask, "margin_sd"] = margin_sd_raw
                                bets_schedule_df.loc[
                                    mask, "spread_source"
                                ] = spread_ensemble.ensemble_id
                                bets_schedule_df.loc[
                                    mask, "spread_ensemble_components_json"
                                ] = components_json
                                spread_ensemble_applied = True
                                spread_games_updated += 1
                            except Exception:
                                continue

                        # Log SPREAD ensemble success
                        _LOG.info(
                            "[SPREAD ensemble] SUCCESS: ensemble_id=%s, games_updated=%d",
                            spread_ensemble.ensemble_id,
                            spread_games_updated,
                        )
                    else:
                        # Single-model: pass through SPREAD data directly (no ensemble)
                        single_model_name = list(unique_spread_models)[0] if unique_spread_models else bets_model_name
                        _LOG.info(
                            "[SPREAD ensemble] Single-model pass-through: model=%s",
                            single_model_name,
                        )
                        _spread_single_games_updated = 0
                        for gid in pd.unique(bets_schedule_df["game_id"]):
                            try:
                                subset = forecast_df[forecast_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                # Take the first (and only) row for this game
                                row = subset.iloc[0]
                                margin_mean = row.get("margin_mean")
                                margin_sd = row.get("margin_sd")
                                if margin_mean is None or (isinstance(margin_mean, float) and pd.isna(margin_mean)):
                                    continue
                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[mask, "margin_mean"] = margin_mean
                                if margin_sd is not None and not (isinstance(margin_sd, float) and pd.isna(margin_sd)):
                                    bets_schedule_df.loc[mask, "margin_sd"] = margin_sd
                                bets_schedule_df.loc[mask, "spread_source"] = single_model_name
                                # Single model: store components as JSON with single entry
                                components = {"models": [single_model_name], "margin_mean": float(margin_mean)}
                                if margin_sd is not None and not (isinstance(margin_sd, float) and pd.isna(margin_sd)):
                                    components["margin_sd"] = float(margin_sd)
                                bets_schedule_df.loc[mask, "spread_ensemble_components_json"] = json.dumps(components)
                                spread_ensemble_applied = True
                                _spread_single_games_updated += 1
                            except Exception:
                                continue
                        _LOG.info(
                            "[SPREAD ensemble] Single-model complete: games_updated=%d",
                            _spread_single_games_updated,
                        )
        except Exception as e:
            _LOG.warning(f"SPREAD ensemble application failed: {e}", exc_info=True)
            pass

        # -------------------------------------------------------------------------
        # TOTAL ENSEMBLE APPLICATION: Apply TOTAL ensemble to BETS schedule
        # For multi-model, use ensemble; for single model, pass through directly.
        # Logs weight resolution, filtering, and per-game application stats.
        # -------------------------------------------------------------------------
        # single model, pass through the model's TOTAL forecast directly. This ensures
        # leftover ML totals don't seep into BETS when only one model runs.
        total_ensemble_applied = False
        try:
            total_rows = market_forecast_rows.get(Market.TOTAL.name, [])
            _LOG.debug(
                "[TOTAL ensemble] Starting: forecast_rows=%d, bets_schedule_rows=%d",
                len(total_rows),
                len(bets_schedule_df) if bets_schedule_df is not None else 0,
            )
            if bets_schedule_df is not None and total_rows:
                forecast_df = pd.DataFrame(total_rows)
                if not forecast_df.empty:
                    weights, ensemble_models, weight_source, weight_run_id, selection_run_id = _resolve_ensemble_weights(
                        db_path=db_path,
                        sport=sport,
                        season=season,
                        market=Market.TOTAL.name,
                        ensemble_id=ensemble_ids[Market.TOTAL.name],
                        config_weights=config_weights_map.get(Market.TOTAL.name),
                        selection_context=selection_contexts[Market.TOTAL.name],
                        tuning_context=tuning_contexts[Market.TOTAL.name],
                        config_warnings=config_warnings,
                    )

                    # Log weight resolution results
                    _LOG.info(
                        "[TOTAL ensemble] Weights resolved: source=%s, models=%s, weights=%s, run_id=%s",
                        weight_source,
                        ensemble_models,
                        weights,
                        weight_run_id,
                    )

                    # Filter forecast data to only include models the ensemble was trained on
                    if ensemble_models:
                        forecast_df = forecast_df[forecast_df["model_name"].isin(set(ensemble_models))]
                        if forecast_df.empty:
                            config_warnings.append(
                                f"No TOTAL forecasts matched ensemble models {ensemble_models}; ensemble skipped"
                            )
                            raise Exception("No TOTAL forecasts after ensemble model filter")
                    filtered_weights, final_models = _filter_market_weights_for_forecast(
                        weights=weights,
                        forecast_df=forecast_df,
                        market=Market.TOTAL.name,
                    )
                    if not final_models:
                        config_warnings.append(
                            "No TOTAL forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No TOTAL forecasts after weight filtering")
                    forecast_df = forecast_df[forecast_df["model_name"].isin(final_models)]
                    if forecast_df.empty:
                        config_warnings.append(
                            "No TOTAL forecasts remained after weight filtering; ensemble skipped"
                        )
                        raise Exception("No TOTAL forecasts after weight filtering")
                    weights = filtered_weights

                    # Log post-filtering state
                    unique_total_models = set(forecast_df["model_name"].dropna().unique())
                    use_ensemble = len(unique_total_models) > 1
                    _LOG.info(
                        "[TOTAL ensemble] Post-filter: final_models=%s, use_ensemble=%s, filtered_weights=%s",
                        sorted(final_models) if final_models else [],
                        use_ensemble,
                        filtered_weights,
                    )

                    # Counter for tracking games updated
                    _total_games_updated = 0

                    if use_ensemble:
                        # Multi-model: use weighted ensemble
                        total_ensemble = TotalWeightedAverageEnsemble(
                            sport,
                            season,
                            ensemble_id=ensemble_ids[Market.TOTAL.name],
                            weights=weights,
                        )
                        used_models = sorted(unique_total_models)
                        cfg_meta = market_config_meta.get(Market.TOTAL.name, {}) if isinstance(market_config_meta, dict) else {}
                        resolved_ensemble_meta[Market.TOTAL.name] = {
                            "ensemble_id": total_ensemble.ensemble_id,
                            "metric_slot": market_metrics.get(Market.TOTAL.name),
                            "configured_models": ensemble_models or used_models,
                            "configured_weights": config_weights_map.get(Market.TOTAL.name),
                            "weights_source": weight_source or "equal",
                            "source_run_id": weight_run_id,
                            "weights": weights,
                            "used_models": used_models,
                            "config_source": cfg_meta.get("source"),
                            "config_path": cfg_meta.get("path"),
                            "selection_run_id": selection_run_id,
                            "selection_models": (
                                selection_contexts[Market.TOTAL.name]["models"]
                                if selection_contexts[Market.TOTAL.name]
                                else None
                            ),
                            "tuning_run_id": weight_run_id if weight_source == "db_tuned" else None,
                        }
                        # Compute recency adjustment for totals (e.g., to account for January slowdown)
                        total_adjustment = None
                        if db_path and sport and season:
                            total_adjustment = _compute_total_recency_adjustment(
                                db_path,
                                sport,
                                as_of_date=as_of_date,
                                lookback_games=100,
                            )
                        
                        for gid in pd.unique(bets_schedule_df["game_id"]):
                            try:
                                subset = forecast_df[forecast_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                total_mean_raw, total_sd_raw, components_json = (
                                    total_ensemble.combine(subset)
                                )
                                if total_mean_raw is None:
                                    continue
                                
                                # Apply recency adjustment if available
                                if total_adjustment is not None and abs(total_adjustment) > 0.5:
                                    total_mean_raw = total_mean_raw + total_adjustment
                                
                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[mask, "total"] = total_mean_raw
                                if total_sd_raw is not None:
                                    bets_schedule_df.loc[mask, "total_sd"] = total_sd_raw
                                bets_schedule_df.loc[
                                    mask, "total_source"
                                ] = total_ensemble.ensemble_id
                                bets_schedule_df.loc[
                                    mask, "total_ensemble_components_json"
                                ] = components_json
                                total_ensemble_applied = True
                                _total_games_updated += 1
                            except Exception:
                                continue

                        # Log TOTAL ensemble success
                        _LOG.info(
                            "[TOTAL ensemble] SUCCESS: ensemble_id=%s, games_updated=%d, recency_adjustment=%s",
                            total_ensemble.ensemble_id,
                            _total_games_updated,
                            total_adjustment,
                        )
                    else:
                        # Single-model: pass through TOTAL data directly (no ensemble)
                        _LOG.info(
                            "[TOTAL ensemble] Single-model pass-through starting",
                        )
                        _total_single_games_updated = 0
                        # Compute recency adjustment for totals
                        total_adjustment = None
                        if db_path and sport and season:
                            total_adjustment = _compute_total_recency_adjustment(
                                db_path,
                                sport,
                                as_of_date=as_of_date,
                                lookback_games=100,
                            )
                        
                        single_model_name = list(unique_total_models)[0] if unique_total_models else bets_model_name
                        for gid in pd.unique(bets_schedule_df["game_id"]):
                            try:
                                subset = forecast_df[forecast_df["game_id"] == gid]
                                if subset.empty:
                                    continue
                                # Take the first (and only) row for this game
                                row = subset.iloc[0]
                                total_mean = row.get("total_mean")
                                total_sd = row.get("total_sd")
                                if total_mean is None or (isinstance(total_mean, float) and pd.isna(total_mean)):
                                    continue
                                
                                # Apply recency adjustment if available
                                if total_adjustment is not None and abs(total_adjustment) > 0.5:
                                    total_mean = total_mean + total_adjustment
                                
                                mask = bets_schedule_df["game_id"] == gid
                                bets_schedule_df.loc[mask, "total"] = total_mean
                                if total_sd is not None and not (isinstance(total_sd, float) and pd.isna(total_sd)):
                                    bets_schedule_df.loc[mask, "total_sd"] = total_sd
                                bets_schedule_df.loc[mask, "total_source"] = single_model_name
                                # Single model: store components as JSON with single entry
                                components = {"models": [single_model_name], "total_mean": float(total_mean)}
                                if total_sd is not None and not (isinstance(total_sd, float) and pd.isna(total_sd)):
                                    components["total_sd"] = float(total_sd)
                                bets_schedule_df.loc[mask, "total_ensemble_components_json"] = json.dumps(components)
                                total_ensemble_applied = True
                                _total_single_games_updated += 1
                            except Exception:
                                continue
                        _LOG.info(
                            "[TOTAL ensemble] Single-model complete: model=%s, games_updated=%d, recency_adjustment=%s",
                            single_model_name,
                            _total_single_games_updated,
                            total_adjustment,
                        )
        except Exception as e:
            _LOG.warning(f"TOTAL ensemble application failed: {e}", exc_info=True)
            pass

        # If the ensemble was applied successfully to any games, set the BETS model label
        # and ensure META reflects the ensemble as the bets model.
        if ensemble_applied and bets_schedule_df is not None:
            try:
                bets_schedule_df["model"] = ensemble.ensemble_id
                bets_model_name = ensemble.ensemble_id
            except Exception:
                # Best-effort assignment; ignore failures.
                pass

        ml_ensemble_id = ensemble_ids.get(Market.ML.name, "ensemble_ml_v1")
        spread_ensemble_id = ensemble_ids.get(Market.SPREAD.name, "ensemble_spread_v1")
        total_ensemble_id = ensemble_ids.get(Market.TOTAL.name, "ensemble_total_v1")

        # CRITICAL: Always use the ensemble source IDs for SPREAD and TOTAL markets.
        # These IDs are the canonical identifiers for where predictions came from,
        # even when individual market forecasts aren't applied to the BETS schedule.
        spread_source_id = spread_ensemble_id
        total_source_id = total_ensemble_id
        ml_source_id = (
            ml_ensemble_id if ensemble_applied else (bets_model_name or "direct")
        )
        _LOG.info(f"[get_schedule] Sources: ML={ml_source_id}, SPREAD={spread_source_id}, TOTAL={total_source_id}, ensemble_applied={ensemble_applied}")
        review_run_id = _deterministic_review_run_id(
            sport=sport,
            season=season,
            as_of_date=resolved_as_of_date,
            ml_source_id=ml_source_id,
            spread_source_id=spread_source_id,
            total_source_id=total_source_id,
        )

        # -------------------------------------------------------------------------
        # BETS DATAFRAME CONSTRUCTION: Build the final BETS sheet data
        # -------------------------------------------------------------------------
        _LOG.info(
            "[BETS sheet] Starting _build_bets_dataframe: input_rows=%d, model=%s, review_run_id=%s",
            len(bets_schedule_df) if bets_schedule_df is not None else 0,
            bets_model_name,
            review_run_id,
        )
        bets_df = _build_bets_dataframe(
            bets_schedule_df if bets_schedule_df is not None else pd.DataFrame(),
            model_name=bets_model_name,
            as_of_date=resolved_as_of_date,
            review_run_id=review_run_id,
            db_path=db_path,
            sport=sport,
            season=season,
            spread_ensemble_applied=spread_ensemble_applied,
            total_ensemble_applied=total_ensemble_applied,
        )
        _LOG.info(
            "[BETS sheet] _build_bets_dataframe complete: output_rows=%d, unique_games=%d",
            len(bets_df),
            bets_df["game_id"].nunique() if "game_id" in bets_df.columns and not bets_df.empty else 0,
        )
        
        # Populate ensemble source IDs in BETS dataframe for all rows
        # This ensures the correct source is persisted to the database
        if not bets_df.empty:
            # For SPREAD rows, always set to ensemble ID (overwrite any pass-through or fallback values)
            if "market_type" in bets_df.columns and "spread_source" in bets_df.columns:
                mask = bets_df["market_type"] == "spread"
                count = mask.sum()
                if count > 0:
                    before_sources = bets_df.loc[mask, "spread_source"].unique()
                    _LOG.info(f"[BETS] SPREAD rows BEFORE: {before_sources}, count={count}")
                    bets_df.loc[mask, "spread_source"] = spread_source_id
                    after_sources = bets_df.loc[mask, "spread_source"].unique()
                    _LOG.info(f"[BETS] SPREAD rows AFTER: {after_sources}, spread_source_id={spread_source_id}")
            
            # For TOTAL rows, always set to ensemble ID (overwrite any pass-through or fallback values)
            if "market_type" in bets_df.columns and "total_source" in bets_df.columns:
                mask = bets_df["market_type"] == "total"
                count = mask.sum()
                if count > 0:
                    before_sources = bets_df.loc[mask, "total_source"].unique()
                    _LOG.info(f"[BETS] TOTAL rows BEFORE: {before_sources}, count={count}")
                    bets_df.loc[mask, "total_source"] = total_source_id
                    after_sources = bets_df.loc[mask, "total_source"].unique()
                    _LOG.info(f"[BETS] TOTAL rows AFTER: {after_sources}, total_source_id={total_source_id}")
        
        bets_df.to_excel(writer, sheet_name="BETS", index=False)
        _LOG.info(
            "[BETS sheet] Written to Excel: rows=%d, sheet=BETS",
            len(bets_df),
        )

        param_sources = _collect_market_param_sources(
            db_path=db_path,
            sport=sport,
            season=season,
            models=models,
        )
        ensemble_sources = _collect_ensemble_weight_sources(
            db_path=db_path,
            sport=sport,
            season=season,
            ensemble_ids=ensemble_ids,
        )
        config_path = next(
            (meta.get("path") for meta in market_config_meta.values() if isinstance(meta, dict) and meta.get("path")),
            None,
        )
        config_sha = next(
            (meta.get("sha256") for meta in market_config_meta.values() if isinstance(meta, dict) and meta.get("sha256")),
            None,
        )
        config_sources_json = json.dumps(market_config_meta, sort_keys=True, default=str)
        meta_rows = [
            {"key": "review_run_id", "value": review_run_id},
            {"key": "sport", "value": sport},
            {"key": "season", "value": season},
            {"key": "as_of_date", "value": resolved_as_of_date.isoformat()},
            {"key": "bets_model", "value": bets_model_name},
            {"key": "workbook_kind", "value": "schedule_with_bets"},
            {"key": "created_at_utc", "value": datetime.now(timezone.utc).isoformat()},
            {
                "key": "ensemble_config_path",
                "value": config_path,
            },
            {
                "key": "ensemble_config_sha256",
                "value": config_sha,
            },
            {
                "key": "ensemble_config_warnings",
                "value": ", ".join(config_warnings) if config_warnings else None,
            },
            {
                "key": "ensemble_config_sources_json",
                "value": config_sources_json,
            },
            {
                "key": "active_model_market_params_source_json",
                "value": json.dumps(param_sources, sort_keys=True),
            },
            {
                "key": "active_ensemble_market_weights_source_json",
                "value": json.dumps(ensemble_sources, sort_keys=True),
            },
            {
                "key": "resolved_ensemble_membership_json",
                "value": json.dumps(resolved_ensemble_meta, sort_keys=True, default=str),
            },
        ]
        meta_df = pd.DataFrame(meta_rows)
        meta_df.to_excel(writer, sheet_name="META", index=False)
        _LOG.debug(
            "[META sheet] Written to Excel: rows=%d, sheet=META",
            len(meta_df),
        )

    # -------------------------------------------------------------------------
    # WORKBOOK POST-PROCESSING: Apply formulas, formatting, and save
    # -------------------------------------------------------------------------
    _LOG.debug("[Workbook] Loading workbook for post-processing: %s", report_path)
    wb = load_workbook(report_path)
    if "BETS" in wb.sheetnames:
        ws = wb["BETS"]
        apply_ev_formulas(ws, use_price=True)
        apply_model_prob_formulas_for_bets_sheet(ws)
        validate_bets_formulas(ws)
        # Apply UX and helper formatting (add-only, best-effort)
        try:
            apply_bets_sheet_formatting(ws)
        except Exception:
            pass
        validate_no_ellipsis_formulas(wb)
        # Format the `stake` column as US dollars (column header: "stake").
        try:
            header = next(ws.iter_rows(min_row=1, max_row=1))
            stake_col = None
            for cell in header:
                if cell.value is not None and str(cell.value).strip().lower() == "stake":
                    stake_col = cell.column_letter
                    break
            if stake_col:
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{stake_col}{r}"]
                    cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
        except Exception:
            # Best-effort formatting; do not fail the report generation on errors.
            pass

        ws.protection.sheet = False
    if "META" in wb.sheetnames:
        ws = wb["META"]
        ws.sheet_state = "hidden"
    wb.save(report_path)
    _LOG.info(
        "[Workbook] Saved successfully: path=%s",
        report_path,
    )
    return report_path
