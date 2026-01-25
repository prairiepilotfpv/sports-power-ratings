"""Reporting helpers for bets: daily/weekly/monthly summaries and edge/CLV summaries."""

from __future__ import annotations

import sqlite3
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
try:
    from openpyxl.worksheet.sparklines import SparkLine, SparkLineList
except ImportError:  # older openpyxl versions omit sparklines
    SparkLine = None
    SparkLineList = None


def _edge_bucket(edge: float | None) -> str:
    if edge is None:
        return "<0%"
    if edge >= 0.05:
        return ">5%"
    if edge >= 0.02:
        return "2-5%"
    if edge >= 0.0:
        return "0-2%"
    return "<0%"


def daily_report(db_path: str | Path, *, sport: str, season: str) -> List[Dict[str, Any]]:
    """Aggregate bets by game date with simple stake/PnL/EV rollups."""
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        q = """
            SELECT g.date,
                   b.status,
                   b.stake,
                   COALESCE(b.profit, 0),
                   o.edge,
                   b.id
            FROM bets b
            LEFT JOIN games g ON b.game_id = g.game_id
            LEFT JOIN opportunities o ON b.source_opportunity_id = o.id
            WHERE g.sport = ? AND g.season = ?
        """
        rows = cur.execute(q, (sport, season)).fetchall()
        agg: Dict[str, Dict[str, Any]] = {}
        for game_date, status, stake, profit, edge, _bid in rows:
            if game_date is None:
                continue
            if isinstance(game_date, str):
                try:
                    game_date = pd.to_datetime(game_date).date()
                except Exception:
                    continue
            key = game_date
            rec = agg.setdefault(
                key,
                {
                    "date": game_date,
                    "total_bets": 0,
                    "total_stake": 0.0,
                    "avg_edge": None,
                    "settled_count": 0,
                    "pending_count": 0,
                    "realized_pl": 0.0,
                    "unrealized_ev": 0.0,
                },
            )
            rec["total_bets"] += 1
            stake_val = float(stake or 0.0)
            rec["total_stake"] += stake_val
            if edge is not None:
                if rec["avg_edge"] is None:
                    rec["avg_edge"] = float(edge) * stake_val
                else:
                    rec["avg_edge"] += float(edge) * stake_val
            if status == "settled":
                rec["settled_count"] += 1
                rec["realized_pl"] += float(profit or 0.0)
            else:
                rec["pending_count"] += 1
        # finalize weighted avg_edge
        for rec in agg.values():
            if rec["avg_edge"] is not None and rec["total_stake"]:
                rec["avg_edge"] = float(rec["avg_edge"]) / float(rec["total_stake"])
        return sorted(agg.values(), key=lambda r: r["date"])
    finally:
        conn.close()


def edge_bucket_report(db_path: str | Path, *, sport: str, season: str) -> List[Dict[str, Any]]:
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        q = """
            SELECT o.edge AS edge, COUNT(b.id) AS count, SUM(b.stake) AS total_stake,
                   SUM(CASE WHEN b.status='settled' THEN COALESCE(b.profit,0) ELSE 0 END) AS profit
            FROM bets b
            LEFT JOIN opportunities o ON b.source_opportunity_id = o.id
            LEFT JOIN games g ON b.game_id = g.game_id
            WHERE g.sport = ? AND g.season = ?
            GROUP BY o.edge
        """
        rows = cur.execute(q, (sport, season)).fetchall()
        buckets: Dict[str, Dict[str, Any]] = {}
        for edge, count, stake_sum, profit_sum in rows:
            bucket = _edge_bucket(edge)
            b = buckets.setdefault(bucket, {"count": 0, "stake": 0.0, "profit": 0.0})
            b["count"] += int(count or 0)
            b["stake"] += float(stake_sum or 0.0)
            b["profit"] += float(profit_sum or 0.0)
        result = []
        for label, data in buckets.items():
            roi = None
            if data["stake"]:
                roi = data["profit"] / data["stake"]
            result.append({"edge_bucket": label, "count": data["count"], "stake": data["stake"], "profit": data["profit"], "roi": roi})
        order = [">5%", "2-5%", "0-2%", "<0%"]
        result.sort(key=lambda r: order.index(r["edge_bucket"]) if r["edge_bucket"] in order else 99)
        return result
    finally:
        conn.close()


def clv_summary(db_path: str | Path, *, sport: str, season: str) -> Dict[str, Any]:
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        q = """
            WITH latest_clv AS (
                SELECT game_id, market_type, selection, close_line, close_odds
                FROM (
                    SELECT game_id, market_type, selection, close_line, close_odds, captured_at, id,
                           ROW_NUMBER() OVER(PARTITION BY game_id, market_type, selection ORDER BY datetime(captured_at) DESC, id DESC) AS rn
                    FROM clv_snapshots
                ) WHERE rn = 1
            )
            SELECT AVG(COALESCE(b.clv_close_odds, lc.close_odds)), AVG(COALESCE(b.clv_close_line, lc.close_line))
            FROM bets b
            LEFT JOIN games g ON b.game_id = g.game_id
            LEFT JOIN latest_clv lc ON lc.game_id = b.game_id AND lc.market_type = b.market_type AND lc.selection = b.selection
            WHERE g.sport = ? AND g.season = ? AND COALESCE(b.clv_close_odds, lc.close_odds) IS NOT NULL
        """
        row = cur.execute(q, (sport, season)).fetchone()
        return {"avg_clv_close_odds": float(row[0]) if row and row[0] is not None else None, "avg_clv_close_line": float(row[1]) if row and row[1] is not None else None}
    finally:
        conn.close()


def write_report_csv(rows: List[Dict[str, Any]], output_path: str | Path) -> Path:
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    df.to_csv(p, index=False)
    return p


def write_report_xlsx(rows: List[Dict[str, Any]], output_path: str | Path, sheet_name: str = "report") -> Path:
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    df.to_excel(p, sheet_name=sheet_name, index=False)
    return p


def write_full_report_xlsx(
    db_path: str | Path,
    *,
    sport: str,
    season: str,
    rpt_type: str,
    rows: List[Dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """Write a workbook containing the main report plus edge buckets and CLV sheets.

    - `rpt_type` should be one of `daily`, `weekly`, or `monthly` and will be
      used as the main sheet name.
    """
    p = Path(output_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    main_df = pd.DataFrame(rows)
    edge_rows = edge_bucket_report(db_path, sport=sport, season=season)
    edge_df = pd.DataFrame(edge_rows)
    clv = clv_summary(db_path, sport=sport, season=season)
    clv_df = pd.DataFrame([clv]) if clv else pd.DataFrame()

    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(p, engine="openpyxl") as writer:
        # Main sheet (daily/weekly/monthly) — write title row then data starting at row 2
        sheet_name = rpt_type if rpt_type in ("daily", "weekly", "monthly") else "report"
        if not main_df.empty:
            main_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        else:
            pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

        # Edge buckets — write title row then data at row 2
        if not edge_df.empty:
            edge_df.to_excel(writer, sheet_name="edge_buckets", index=False, startrow=1)
        else:
            pd.DataFrame(edge_rows).to_excel(writer, sheet_name="edge_buckets", index=False, startrow=1)

        # CLV summary — write title row then data at row 2
        if not clv_df.empty:
            clv_df.to_excel(writer, sheet_name="clv", index=False, startrow=1)

        # After writing, add header/title rows and basic number formats
        ws_main = writer.sheets[sheet_name]
        title = f"{sport.upper()} {season} {rpt_type.title()} Report"
        ws_main.cell(row=1, column=1, value=title)
        ws_main.cell(row=1, column=1).font = Font(bold=True)

        # Edge buckets formatting
        ws_edge = writer.sheets.get("edge_buckets")
        if ws_edge is not None:
            ws_edge.cell(row=1, column=1, value="Edge Buckets")
            ws_edge.cell(row=1, column=1).font = Font(bold=True)
            # find columns and apply numeric formats where applicable
            header_row = [c.value for c in next(ws_edge.iter_rows(min_row=2, max_row=2))]
            # mapping: stake -> currency, profit -> currency, roi -> percent
            for idx, h in enumerate(header_row, start=1):
                key = (h or "").lower()
                if "stake" in key or "profit" in key:
                    for cell in ws_edge.iter_rows(min_row=3, min_col=idx, max_col=idx):
                        for c in cell:
                            try:
                                c.number_format = "\u00A4#,##0.00"
                            except Exception:
                                pass
                if "roi" in key:
                    for cell in ws_edge.iter_rows(min_row=3, min_col=idx, max_col=idx):
                        for c in cell:
                            try:
                                c.number_format = "0.0%"
                            except Exception:
                                pass

            # autofit edge_buckets columns
            for col_idx in range(1, ws_edge.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws_edge[col_letter]:
                    try:
                        val = "" if cell.value is None else str(cell.value)
                    except Exception:
                        val = ""
                    max_len = max(max_len, len(val))
                ws_edge.column_dimensions[col_letter].width = max(10, max_len + 2)

            # conditional formatting: ROI color scale, Profit data bar
            header_map = { (h or "").lower(): idx for idx, h in enumerate(header_row, start=1) }
            roi_idx = header_map.get("roi")
            profit_idx = header_map.get("profit")
            stake_idx = header_map.get("stake")

            if roi_idx:
                roi_col = get_column_letter(roi_idx)
                ws_edge.conditional_formatting.add(
                    f"{roi_col}3:{roi_col}{ws_edge.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=0,
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )
            if profit_idx:
                profit_col = get_column_letter(profit_idx)
                ws_edge.conditional_formatting.add(
                    f"{profit_col}3:{profit_col}{ws_edge.max_row}",
                    DataBarRule(start_type="num", start_value=0, end_type="max", color="63BE7B"),
                )

            # sparkline for stake trend across buckets (single sparkline)
            if SparkLine and SparkLineList and stake_idx and ws_edge.max_row >= 3:
                stake_col = get_column_letter(stake_idx)
                spark_col = get_column_letter(ws_edge.max_column + 1)
                ws_edge.cell(row=2, column=ws_edge.max_column + 1, value="stake_spark")
                spark = SparkLine(range=f"{stake_col}3:{stake_col}{ws_edge.max_row}", location=f"{spark_col}3")
                spark_group = SparkLineList(type="column", sparklines=[spark])
                if ws_edge._sparklines is None:
                    ws_edge._sparklines = []
                ws_edge._sparklines.append(spark_group)

        # CLV formatting
        ws_clv = writer.sheets.get("clv")
        if ws_clv is not None:
            ws_clv.cell(row=1, column=1, value="CLV Summary")
            ws_clv.cell(row=1, column=1).font = Font(bold=True)
            # apply numeric format to known columns
            clv_headers = [c.value for c in next(ws_clv.iter_rows(min_row=2, max_row=2))]
            for idx, h in enumerate(clv_headers, start=1):
                key = (h or "").lower()
                if "odds" in key or "line" in key:
                    for cell in ws_clv.iter_rows(min_row=3, min_col=idx, max_col=idx):
                        for c in cell:
                            try:
                                c.number_format = "\u00A4#,##0.00"
                            except Exception:
                                pass

            # autofit clv columns
            for col_idx in range(1, ws_clv.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws_clv[col_letter]:
                    try:
                        val = "" if cell.value is None else str(cell.value)
                    except Exception:
                        val = ""
                    max_len = max(max_len, len(val))
                ws_clv.column_dimensions[col_letter].width = max(10, max_len + 2)

            # color scale for odds/line if present
            header_map = { (h or "").lower(): idx for idx, h in enumerate(clv_headers, start=1) }
            for key in ("avg_clv_close_odds", "avg_clv_close_line"):
                idx = header_map.get(key)
                if idx:
                    col = get_column_letter(idx)
                    ws_clv.conditional_formatting.add(
                        f"{col}3:{col}{ws_clv.max_row}",
                        ColorScaleRule(
                            start_type="min",
                            start_color="F8696B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="63BE7B",
                        ),
                    )

        # autofit main sheet columns
        for col_idx in range(1, ws_main.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws_main[col_letter]:
                try:
                    val = "" if cell.value is None else str(cell.value)
                except Exception:
                    val = ""
                max_len = max(max_len, len(val))
            ws_main.column_dimensions[col_letter].width = max(10, max_len + 2)

        # Small dashboard with key identifiers (detailed KPIs omitted for stability)
        ws_dash = writer.book.create_sheet("dashboard")
        writer.sheets["dashboard"] = ws_dash
        ws_dash.cell(row=1, column=1, value="Dashboard").font = Font(bold=True)
        ws_dash.cell(row=2, column=1, value="Sport").font = Font(bold=True)
        ws_dash.cell(row=2, column=2, value=sport.upper())
        ws_dash.cell(row=3, column=1, value="Season").font = Font(bold=True)
        ws_dash.cell(row=3, column=2, value=season)

        # PnL scenarios sheet
        try:
            _write_pnl_sheet(writer, db_path=db_path, sport=sport, season=season)
        except Exception:
            pass

        if "PnL_Scenarios" not in writer.book.sheetnames:
            ws_pnl = writer.book.create_sheet("PnL_Scenarios")
            writer.sheets["PnL_Scenarios"] = ws_pnl
            ws_pnl.cell(row=1, column=1, value="PnL Scenarios")

    return p


def _write_pnl_sheet(writer, *, db_path: str | Path, sport: str, season: str):
    """Optional PnL scenarios sheet with Kelly sizing; best-effort."""
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        q = """
            WITH latest_clv AS (
                SELECT game_id, market_type, selection, close_line, close_odds
                FROM (
                    SELECT game_id, market_type, selection, close_line, close_odds, captured_at, id,
                           ROW_NUMBER() OVER(PARTITION BY game_id, market_type, selection ORDER BY datetime(captured_at) DESC, id DESC) AS rn
                    FROM clv_snapshots
                ) WHERE rn = 1
            )
            SELECT b.id, b.stake, b.odds, b.line, b.book,
                   COALESCE(b.clv_close_odds, lc.close_odds) AS clv_close_odds,
                   COALESCE(b.clv_close_line, lc.close_line) AS clv_close_line,
                   o.model_prob, o.edge, g.date, b.game_id, b.market_type, b.selection
            FROM bets b
            LEFT JOIN opportunities o ON b.source_opportunity_id = o.id
            LEFT JOIN games g ON b.game_id = g.game_id
            LEFT JOIN latest_clv lc ON lc.game_id = b.game_id AND lc.market_type = b.market_type AND lc.selection = b.selection
            WHERE g.sport = ? AND g.season = ?
        """
        data = cur.execute(q, (sport, season)).fetchall()
    finally:
        conn.close()

    from src.utils.odds import payout_per_unit, american_to_implied

    pnl_rows = []
    bankroll = 1000.0
    for r in data:
        bid, stake, odds, line, bookcol, clv_odds, clv_line, model_prob, edge, gdate, gid, mtype, sel = r
        stake = float(stake or 0.0)
        implied = None
        if odds is not None:
            try:
                implied = american_to_implied(int(odds))
            except Exception:
                implied = None
        if model_prob is None:
            model_prob = implied
        b = None
        if odds is not None:
            b = payout_per_unit(int(odds))
        elif implied:
            b = (1.0 / implied) - 1.0
        kelly = None
        if b and model_prob is not None:
            qv = 1.0 - float(model_prob)
            try:
                val = (b * float(model_prob) - qv) / b
                kelly = max(0.0, val)
            except Exception:
                kelly = 0.0
        recommended = (kelly or 0.0) * bankroll
        expected_ev = (edge or 0.0) * stake if edge is not None else None
        pnl_rows.append(
            {
                "bet_id": bid,
                "date": gdate,
                "game_id": gid,
                "market_type": mtype,
                "selection": sel,
                "odds": odds,
                "line": line,
                "stake": stake,
                "implied_prob": implied,
                "model_prob": model_prob,
                "edge": edge,
                "kelly_fraction": kelly,
                "kelly_recommended_stake": recommended,
                "expected_ev": expected_ev,
                "clv_close_odds": clv_odds,
                "clv_close_line": clv_line,
            }
        )

    if pnl_rows:
        pnl_df = pd.DataFrame(pnl_rows)
    else:
        pnl_df = pd.DataFrame(
            columns=
            [
                "bet_id",
                "date",
                "game_id",
                "market_type",
                "selection",
                "odds",
                "line",
                "stake",
                "implied_prob",
                "model_prob",
                "edge",
                "kelly_fraction",
                "kelly_recommended_stake",
                "expected_ev",
                "clv_close_odds",
                "clv_close_line",
            ]
        )

    pnl_df.to_excel(writer, sheet_name="PnL_Scenarios", index=False)
    ws_pnl = writer.sheets["PnL_Scenarios"]
    for col_idx in range(1, ws_pnl.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws_pnl[col_letter]:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws_pnl.column_dimensions[col_letter].width = max(10, max_len + 2)


def weekly_report(db_path: str | Path, *, sport: str, season: str, start: str | None = None, end: str | None = None) -> List[Dict[str, Any]]:
    """Aggregate daily report into weekly buckets (weeks starting on Monday)."""
    rows = daily_report(db_path, sport=sport, season=season)
    if not rows:
        return []
        # Dashboard sheet: key KPIs
        try:
            daily_rows = daily_report(db_path, sport=sport, season=season)
            total_bets = sum(r.get("total_bets", 0) for r in daily_rows)
            total_stake = sum(r.get("total_stake", 0.0) for r in daily_rows)
            realized_pl = sum(r.get("realized_pl", 0.0) for r in daily_rows)
            pending_ev = sum(r.get("unrealized_ev", 0.0) for r in daily_rows)
            avg_edge = None
            if total_stake:
                weighted = sum((r.get("avg_edge") or 0.0) * (r.get("total_stake") or 0.0) for r in daily_rows)
                avg_edge = float(weighted / total_stake)

            settled = sum(r.get("settled_count", 0) for r in daily_rows)
            win_rate = float(settled) / float(total_bets) if settled and total_bets else None

            dashboard = [
                ("Sport", sport.upper()),
                ("Season", season),
                ("Total Bets", total_bets),
                ("Total Stake", total_stake),
                ("Avg Edge", avg_edge),
                ("Realized P/L", realized_pl),
                ("Unrealized EV", pending_ev),
                ("Settled Count", settled),
                ("Win Rate (approx)", win_rate),
            ]

            ws_dash = writer.book.create_sheet("dashboard")
            writer.sheets["dashboard"] = ws_dash
            ws_dash.cell(row=1, column=1, value="Dashboard").font = Font(bold=True)
            for i, (k, v) in enumerate(dashboard, start=2):
                ws_dash.cell(row=i, column=1, value=k).font = Font(bold=True)
                cell = ws_dash.cell(row=i, column=2, value=v)
                if k in ("Total Stake", "Realized P/L", "Unrealized EV"):
                    try:
                        cell.number_format = "\u00A4#,##0.00"
                    except Exception:
                        pass
                if k in ("Avg Edge", "Win Rate (approx)"):
                    try:
                        cell.number_format = "0.0%"
                        if v is not None:
                            cell.value = float(v)
                    except Exception:
                        pass

            for col_idx in range(1, ws_dash.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws_dash[col_letter]:
                    try:
                        val = "" if cell.value is None else str(cell.value)
                    except Exception:
                        val = ""
                    max_len = max(max_len, len(val))
                ws_dash.column_dimensions[col_letter].width = max(10, max_len + 2)

            # best-effort PnL scenarios sheet
            _write_pnl_sheet(writer, db_path=db_path, sport=sport, season=season)
        except Exception:
            pass

        # ensure PnL sheet exists even if helper skipped
        if "PnL_Scenarios" not in writer.book.sheetnames:
            ws_pnl = writer.book.create_sheet("PnL_Scenarios")
            writer.sheets["PnL_Scenarios"] = ws_pnl
            ws_pnl.cell(row=1, column=1, value="PnL Scenarios")
